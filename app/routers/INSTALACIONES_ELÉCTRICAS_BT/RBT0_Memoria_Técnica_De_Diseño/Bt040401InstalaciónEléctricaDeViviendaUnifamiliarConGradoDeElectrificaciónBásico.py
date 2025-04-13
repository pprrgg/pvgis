#!/usr/bin/env python3

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from io import BytesIO
import os
import matplotlib.pyplot as plt
import pandas as pd
from types import SimpleNamespace
import subprocess
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata


import numpy as np
import sys

router = APIRouter()


@router.post(f"/Bt040401")
def Bt040401InstalaciónEléctricaDeViviendaUnifamiliarConGradoDeElectrificaciónBásico(
    data: dict,
):
    data = comun.texfile.limpiar_valores(data)
    print(data)
    # parametros
    pp = SimpleNamespace()
    # tablas y figuras en latex
    uu = SimpleNamespace()
    #
    vv = SimpleNamespace()
    xx = SimpleNamespace()
    yy = SimpleNamespace()
    # datos genericos
    zz = SimpleNamespace()

    zz.nombre = sys._getframe().f_code.co_name
    zz.ruta_script = os.path.dirname(
        os.path.abspath(__file__)
    )  # Obtener la carpeta sin el nombre del script

    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)
    zz.data = data

    output_buffer = comun.main.main(pp, uu, vv, xx, yy, zz)

    with open(f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf", "wb") as f:
        f.write(output_buffer.getvalue())

    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


def file_tex(pp, uu, vv, xx, yy, zz):

    return r"""

\documentclass[a4paper,10pt]{article}
%\documentclass[a4paper,10pt,twocolumn]{article}
\usepackage[top = 2cm, bottom = 2cm, left = 2cm, right = 2cm, asymmetric]{geometry} % Especificar los márgenes según la norma
\usepackage[spanish]{babel}    

\usepackage{graphicx}  % Para incluir imágenes si es necesario
\usepackage{amsmath, amssymb}  % Para expresiones matemáticas
\usepackage{fancyhdr}  % Para personalizar encabezados
\usepackage{chngcntr}  % Para cambiar la numeración de apartados y subsecciones
\usepackage{tocbibind}  % Para incluir bibliografía en la tabla de contenidos
\usepackage{appendix}  % Para el formato de anexos
\usepackage{lipsum}  % Para generar texto de relleno (dummy text)
\usepackage{geometry}  % Para personalizar los márgenes
\usepackage{multicol}  % Para columnas
\usepackage{titlesec}  % Para personalizar los títulos de las secciones
\usepackage[utf8]{inputenc}    
\usepackage[T1]{fontenc}       
\usepackage[pdftex,pdfencoding=auto, colorlinks=true, linkcolor=black, citecolor=black, filecolor=black, urlcolor=black]{hyperref}
\usepackage{tocloft}           
\usepackage{booktabs}          
\usepackage{float} % Paquete necesario para usar la opción [H]
\usepackage{array}
\usepackage{longtable}
\usepackage{circuitikz}
\usepackage{tikz}
\usepackage{tikz-cd}

\usepackage{tikz-qtree}
\usepackage{smartdiagram}

\usepackage{qrcode}            % Paquete para generar QR
\usepackage{tabularx} % Agregar en el preámbulo
\usepackage[absolute,overlay]{textpos} % Para posicionar objetos libremente

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{zz.titulo}  % Nombre del documento en el encabezado izquierdo
\fancyhead[C]{}  % Centro vacío
\fancyhead[R]{\thepage}  % Numeración de páginas en el encabezado derecho
% \renewcommand{\thepage}{\arabic{page}}  % Asegura la numeración de páginas en números arábigos

% Definición de los anexos
\newcommand{\annex}[1]{\section{Anexo #1} \addcontentsline{toc}{section}{Anexo #1}}
\addto\captionsspanish{%
  \renewcommand{\tablename}{Tabla}%
  \renewcommand{\listtablename}{Índice de tablas}%
}
\title{{\small Ref.:\uppercase{zz.codigo}}\\{\textbf{zz.titulo}}}
\author{
pp.ParticipantesP1NombreRaznSocial
}
\date{\today}
\newcommand{\MostrarVariablesAlFinal}{}

















\begin{document}
\begin{Form}

















%% --- PORTADA ---
%\begin{titlepage}
%    \centering
%    {\Large \textbf{TÍTULO DEL INFORME}}\\[1.5cm]
%    %\includegraphics[width=0.4\textwidth]{logo.png}\\[1cm] % Logo de la empresa/institución
%    
%    \textbf{Autor:} Nombre del Autor \\[0.5cm]
%    \textbf{Fecha:} \today \\[1cm]
%    \textbf{Código del documento:} REF-XXXX-YYYY \\[1cm]
%    \textbf{Versión:} 1.0 \\[1cm]
%    \textbf{Empresa / Institución:} Nombre \\[2cm]
%    \vfill
%\end{titlepage}
	\maketitle
    %
	%\begin{abstract}

    %\end{abstract}
% --- ÍNDICE ---
	\tableofcontents  % Tabla de contenidos
	%\listoffigures    % Lista de figuras
	%\listoftables     % Lista de tablas
\newpage
























































\section{MEMORIA TÉCNICA DE DISEÑO}

\subsection{Titular de la instalación:}

Se especificarán nombre y apellidos o Razón Social, así como DNI/CIF, según proceda.

Asimismo, se consignarán los datos de situación de la instalación.

\subsection{Datos de la instalación:}

\begin{itemize}
    \item \textbf{Tensión:} 230 V.
    \item \textbf{Potencia máxima admisible:} 9.200 W.
    \item \textbf{Magnetotérmico o ICPM:} 40 A.
    \item \textbf{Resistencia de tierra:} Consignar el valor medido (ejemplo: 0,7 \si{\ohm}).
    \item \textbf{Empresa suministradora:} IBERDROLA.
    \item \textbf{Instalación:} Consignar según proceda (ejemplo: N).
    \item \textbf{Uso a que se destina:} Vivienda unifamiliar.
    \item \textbf{Superficie:} 170 m².
    \item \textbf{Aforo:} No procede.
\end{itemize}

La potencia máxima admisible dependerá de la sección de la D.I., calculada con un cos$\phi$ = 1, así como del IGA instalado. (Ver cálculos justificativos en el apartado correspondiente).

El ICPM no tiene por qué coincidir con el calibre del IGA, dado que el usuario de la instalación puede elegir libremente su potencia de contratación.

La resistencia de tierra puede en principio obtener el valor máximo de 800 \si{\ohm}, para tensión de contacto de 24 V y sensibilidad del diferencial de 30 mA, no obstante se suelen exigir valores inferiores a 20 \si{\ohm}.

Para viviendas no es necesario consignar el aforo previsto.

\subsection{Acometida:}

Consignar el punto de conexión, tipo de acometida, sección y material de la misma, de acuerdo con los datos suministrados por la compañía suministradora.

\subsection{C.G.P. ó C/C de seguridad:}

\begin{itemize}
    \item \textbf{Tipo:} CGPM.
    \item \textbf{In Base:} 63 A.
    \item \textbf{In cartucho:} 40 A.
\end{itemize}

\subsection{Línea general de alimentación o derivación individual:}

\begin{itemize}
    \item \textbf{Tipo:} Subterránea.
    \item \textbf{Sección:} 16 mm². Cu.
\end{itemize}

\subsection{Módulo de medida:}

\begin{itemize}
    \item \textbf{Tipo:} Monofásico simple tarifa.
    \item \textbf{Situación:} Hornacina en cerramiento perimetral de parcela.
\end{itemize}

Al utilizar un CGPM, ésta ha de estar de acuerdo con las normas de la Compañía suministradora, para viviendas unifamiliares se coloca una caja preparada para albergar contador trifásico de doble tarifa, así como el correspondiente reloj. En este caso, al no ser necesario, se instala contador monofásico de simple tarifa.

\subsection{Protección magnetotérmica/diferencial:}

\begin{itemize}
    \item \textbf{Interruptor General Automático:} In = 40 A.
    \item \textbf{Interruptor diferencial:} 40 A, 30 mA.
\end{itemize}

La potencia mínima de cálculo para viviendas de G.E. elevado será de 9.200 W (IGA = 40 A), pero podría también ser de 11.500 W (IGA = 50 A) y de 14.490 W (IGA = 63 A).

\subsection{Puesta a tierra:}

\begin{itemize}
    \item \textbf{Nº de electrodos:} Consignar el número de electrodos (ejemplo: 3).
    \item \textbf{Tipo:} Consignar el tipo de electrodo (ejemplo: pica).
    \item \textbf{Línea de enlace:} S = 35 mm² Cu.
    \item \textbf{Línea principal:} S = 16 mm² Cu.
\end{itemize}

En el tema de Anejos al Curso se adjunta método de cálculo rápido de la P.A.T., según la NTE-IEP.

\subsection{Relación nominal de receptores y su potencia:}

\begin{tabular}{|l|l|}
\hline
\textbf{RECEPTORES} & \textbf{POTENCIA (W)} \\
\hline
C1 Alumbrado & 2.250 \\
C2 TC uso general + frigorífico + extractor & 3.450 \\
C3 Cocina, Homo & 4.050 \\
C4 Lavadora, lavavajillas y termo eléctrico. & 4.600 \\
C5 TC Baño, Cocina & 3.680 \\
C7 TC Generales & 3.450 \\
C8 Calefac. Elect. & 5.750 \\
C10 Secadora & 2.587 \\
C11 Automatización & 2.300 \\
\hline
\end{tabular}

El número de tomas de cada circuito está limitado por la tabla 1 de la ITC-BT-25 del REBT.

La potencia total no tiene por qué coincidir con la potencia máxima admisible de la instalación.

\subsection{2.2.10. Cálculos justificativos:}

\subsubsection{Cálculo de la instalación interior}

\begin{tabular}{|l|l|l|l|}
\hline
\textbf{Circuito de utilización} & \textbf{PIA (A)} & \textbf{S (mm²)} & \textbf{diam. (mm)} \\
\hline
C1: Iluminación (1) & 10 & 1,5 & 16 \\
C2: Tomas uso general (1) & 16 & 2,5 & 20 \\
C3: Cocina y horno & 25 & 6 & 25 \\
C4: Lavadora, lavavajillas y termo eléctrico & 20 & 4 & 20 \\
C5: Tomas baños y cocina & 16 & 2,5 & 20 \\
C7: Tomas uso general (2) & 16 & 2,5 & 20 \\
C8: Calefacción & 25 & 6 & 25 \\
C10: Secadora & 16 & 2,5 & 20 \\
C11: Automatización & 10 & 1,5 & 16 \\
\hline
\end{tabular}

\subsubsection{Cálculo de la Derivación individual}

\begin{tabular}{|l|l|l|l|l|l|l|l|} % 8 columnas
\hline
\textbf{Potencia (W)} & \textbf{Tensión (V)} & \textbf{cos} & \textbf{Longitud (m)} & \textbf{conductiv. e (\%)} & \textbf{I (A)} & \textbf{S1 (mm²)} & \textbf{S2 (mm²)} \\
\hline
9.200 & 230 & 1 & 25 & 48 & 1,5 & 40,0 & 12,08 \\
\hline
\end{tabular}

\begin{tabular}{|l|l|l|}
\hline
\textbf{Sadop. (mm²)} & \textbf{I admisible(A)} \\
\hline
16 & 100 \\
\hline
\end{tabular}

Para el cálculo de la instalación interior, se han tomado los valores de la Tabla 1 de la ITC-BT-25, los puntos de utilización mínimos en cada estancia deben cumplir la Tabla 2 de la citada instrucción.

Asimismo, en el caso de la instalación interior, no se ha tenido en cuenta el cálculo por caída de tensión, por considerar que en circunstancias normales no será determinante.

Para el cálculo de la Derivación Individual, se ha hecho el cálculo por intensidad (S1) y por caída de tensión (S2), tomando el mayor de ambas. Conviene recordar que reglamentariamente este valor debe ser mayor o igual a 6 mm².

Como se puede apreciar, se ha tomado 48 como valor de la conductividad del cobre y no 56 como habitualmente se viene haciendo; esto es debido a que el valor 56 corresponde a una temperatura en el conductor de 20°C, y el valor 48 a la de 70°C, que es la máxima temperatura que alcanzaría el cable. Evidentemente, para la corriente de servicio el cable nunca alcanzaría los 70°C; pero al ser éste el valor más restrictivo nos valdrá en todos los casos.

Una vez asignado el valor de la sección (en este caso 16 mm²), bastará con comprobar su intensidad máxima admisible en la Tabla 1 de la ITC-BT-07.

Las fórmulas utilizadas para los cálculos se encuentran recogidas en el tema de Anejos al Curso.

\subsection{Memoria descriptiva}

Vivienda unifamiliar de 170 m² de superficie útil, con grado de electrificación básico (9.200 W), con 9 circuitos, instalación interior con conductores unipolares de cobre, aislamiento de PVC, tipo H072-R, instalados bajo tubo en montaje empotrado. Derivación individual con conductores unipolares de cobre, aislamiento de EPR, tipo RZ1-K, instalados bajo tubo en montaje subterráneo.

Para canalizaciones subterráneas, según ITC-BT-07, el cable a utilizar será unipolar o tetrapolar, pero siempre con aislamiento de 0,6/1 kV.

\subsection{Croquis de trazado:}

El croquis de trazado de la instalación deberá contener aquellos aspectos significativos para la correcta interpretación de la misma.

Se acompaña un croquis general en el que se recogen las distancias aproximadas por las que discurrirán las canalizaciones empotradas.

Asimismo, se acompaña croquis también general, de la red equipotencial en cuartos de baño.

Evidentemente, este apartado se podrá ampliar con toda aquella información que se entienda aclaratoria, en mi opinión particular, se debería incluir además un plano de planta de la instalación.






















% --- 7. REFERENCIAS BIBLIOGRÁFICAS ---
\begin{thebibliography}{99}
    \bibitem{normaUNE} UNE 157001:2002. \textit{Criterios generales para la elaboración de documentos técnicos}.
\end{thebibliography}






































































\newpage
\onecolumn


	% P A R A M E T R O S 
	%---------------------
	\TextField[name=I0 ,width=0cm]{}
	\TextField[name=Bt,width=0cm]{}
	%==========================================





\ifdefined\MostrarVariablesAlFinal
%HojaDeDatos



\subsection{algunas chuladas muuuu  simples}
\subsubsection{diagrama en arbol}

\begin{figure}[H]
\begin{tikzpicture}[grow'=right]
\Tree [.A 
        [.{C1 fdasfd}
         [.r ]
         [.r ]
         [.r ]
           ] 
        [.C2 ] 
        [.C3 ] 
      ]

\end{tikzpicture}
\end{figure}



\subsubsection{smartdiagram}
\smartdiagram[sequence diagram]{Waren\\-eingang\\fsadf\\fasdfsafd, Einlagerung,
  Kommission\\-ierung, Verpackung, Versand}

  

\subsubsection{smartdiagram}
  \smartdiagram[descriptive diagram]{
    {Set up,The set up operation consist of..},
    {Run, {After having set up the program, you must run..}},
    {Analyse, You must check what did with analytical tools like..},
    {Modify, {After the analysis, you can still modify or add..}},
    }


    

\subsubsection{smartdiagram}
    \smartdiagram[bubble diagram]{
Build a program,Set up,Run,Analyze,Modify~/\\ Add,Check
}





\subsubsection{smartdiagram 2}
\smartdiagramset{priority arrow width=2cm,
priority arrow height advance=2.25cm,
priority arrow head extend=0.3cm}

\smartdiagram[priority descriptive diagram]{PGfasdfasdfasdfsad F,Ti\textit{k}Z,Smartdiagram}




\subsubsection{smartdiagram}
\smartdiagram[circular diagram]{Set up,Run,Analyse,kljkhkjh,Modify~/ Add}





\subsubsection{smartdiagram}
\smartdiagram[flow diagram:horizontal]{Inicio, Entrada de datos, Procesamiento, Salida de datos, Fin}





\fi
\end{Form}
\end{document}

"""


def parametros():

    nombre_funcion = os.path.basename(__file__)[:-3]
    codigo, _ = comun.texfile.separar_y_mayusculizar(nombre_funcion)
    print(codigo)
    return {
        codigo: [
            ["Parámetro", "Valor"],
            ["Potencia contratada [kW]", 66],
            ["Duración indicativa de la actuación [años]", 15],
            ["Préstamo recibido [€]", 50000],
            ["Tasa de interés del préstamo [\\%]", 5],
            ["beeee [\\%]", 5],
        ],
        "Coordenadas": [
            ["Parámetro", "Valor"],
            ["lat", 36.6641681434112],
            ["lng", -4.45863940137516],
        ],
        "Participantes": [
            ["Parámetro", "P1", "P2", "P3"],
            ["Rol", "", "", ""],
            ["DNI/NIF/NIE", "", "", ""],
            ["Nombre/Razón Social", "Edita Tu nombre", "", ""],
            ["Titulación", "", "", ""],
            ["Dirección", "", "", ""],
            ["Población", "", "", ""],
            ["Teléfono", "", "", ""],
            ["Correo Electrónico", "", "", ""],
        ],
    }


class codigo_para_replace_en_el_latex:

    class calculo:

        def C00Ejemplo(pp, uu, vv, xx, yy, zz):
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )

            # Crear un DataFrame con números aleatorios
            np.random.seed(42)  # Para reproducibilidad
            df = pd.DataFrame(np.random.rand(4, 4))

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["gsdfg"] = total_row

            df = df.copy()

            return pp, uu, vv, xx, yy, zz

    class tabla:

        def T00df2tablas(pp, uu, vv, xx, yy, zz):
            # tabla de todos los df
            for var_name, value in pp.__dict__.items():
                # print(f"pp.{var_name}")
                # Verificar si la variable es un DataFrame
                if isinstance(value, pd.DataFrame):
                    setattr(uu, var_name, f"{(value.to_latex())}")
            # convierte todos los df de uu a tablas
            return pp, uu, vv, xx, yy, zz

        def T01Ejemplo(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'} {__class__.__name__}.{sys._getframe().f_code.co_name}")

            df = pd.DataFrame(np.random.rand(5, 4), columns=["A", "B", "C", "D"])

            _, titulo = comun.texfile.separar_y_mayusculizar(
                sys._getframe().f_code.co_name
            )
            # Generar la tabla en formato LaTeX
            latex_code = df.to_latex(index=False)

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.4\\textwidth}}{{!}}
                {{
                {latex_code}
                }}
                \\caption{{{titulo}}}
            \\end{{table}}
            """
            setattr(
                uu, f"tabla{sys._getframe().f_code.co_name}", latex_code_with_resizebox
            )

            return pp, uu, vv, xx, yy, zz

    class figura:

        def F00Ejemplo(pp, uu, vv, xx, yy, zz):
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            # 1. Crear un archivo temporal para la imagen
            _, titulo = comun.texfile.separar_y_mayusculizar(
                sys._getframe().f_code.co_name
            )
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image:
                temp_image_path1 = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal

            df = pd.DataFrame(np.random.rand(4, 4))
            # .......
            #
            fig, ax = plt.subplots(figsize=(6, 5))
            df.plot(ax=ax)
            #
            # .......

            ax.grid(axis="y", linestyle="--", alpha=0.7)
            plt.xticks(rotation=50)
            ax.set_ylabel("€")  # Etiqueta en el eje y

            plt.savefig(temp_image_path1, format="png", dpi=300)

            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + """\\begin{figure}[H]
                """
                + f"""
                \\includegraphics[width=.5\\textwidth]{{{temp_image_path1}}}
                \\caption{{{titulo}}}
                """
                + """
                \\label{fig:dfasdfsa}
                \\end{figure}
                """
            )
            setattr(uu, f"figura{sys._getframe().f_code.co_name}", latex_code)
            return pp, uu, vv, xx, yy, zz

    class imagen:

        def TodasLasImagnesDeLaCarpetaImg(pp, uu, vv, xx, yy, zz):
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            carpetaimagenes = Path(f"{zz.ruta_script}/{zz.codigo}/img")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.png")] + [
                f.name for f in carpetaimagenes.glob("*.jpg")
            ]

            for imagen in imagenes:
                _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional

                latex_code = (
                    ""
                    + """\\begin{figure}[H]
                    """
                    + f"""
                    \\includegraphics[width=0.9\\textwidth]{{{Path(f"{zz.ruta_script}/img/{imagen}")}}} 
                    \\caption{{{titulo}}}
                    """
                    + """
                    \\label{fig:dfasdfsa}
                    \\end{figure}
                    """
                )

                setattr(uu, f"imagen{imagen}", latex_code)
            return pp, uu, vv, xx, yy, zz


########################################################################################################
########################################################################################################


class comun:
    class main:

        def main(pp, uu, vv, xx, yy, zz):
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            # segun si es un formulario o un documento nuevo

            pp, uu, vv, xx, yy, zz = comun.main.data2parametros(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.calulos(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.tablas(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.figuras(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.imagenes(pp, uu, vv, xx, yy, zz)

            output_buffer = comun.main.replaceLatex(pp, uu, vv, xx, yy, zz)

            # print(uu)
            return output_buffer

        def data2parametros(pp, uu, vv, xx, yy, zz):
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            # Para cada hoja en el JSON
            # Convertir cada valor del diccionario en un DataFrame
            for key, value in zz.data.items():
                key_sanitized = comun.texfile.sanitize_name(
                    key
                )  # Limpiar el nombre de la clave
                df = pd.DataFrame(
                    value[1:], columns=value[0]
                )  # Primera fila como encabezado
                df.set_index(
                    df.columns[0], inplace=True
                )  # Usar la primera columna como índice
                setattr(pp, f"df{key_sanitized}", df)  # Guardar en ss con nombre limpio

                # Crear variables por columna
                for col in df.columns:
                    col_sanitized = comun.texfile.sanitize_name(
                        col
                    )  # Limpiar nombre de columna
                    setattr(
                        pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                    )  # DataFrame con solo esa columna

                    # Crear variables individuales por celda
                    for idx, val in df[col].items():
                        idx_sanitized = comun.texfile.sanitize_name(
                            idx
                        )  # Limpiar índice (fila)
                        if pd.isna(val):
                            val = "."
                        setattr(
                            pp, f"{key_sanitized}{col_sanitized}{idx_sanitized}", val
                        )
            return pp, uu, vv, xx, yy, zz

        def calulos(pp, uu, vv, xx, yy, zz):
            """calculos, valores calculados en yy"""
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            listacalculos = [
                m
                for m in dir(codigo_para_replace_en_el_latex.calculo)
                if callable(getattr(codigo_para_replace_en_el_latex.calculo, m))
                and not m.startswith("__")
            ]
            listacalculos = [s for s in listacalculos if not s.startswith("__")]
            for calculo in sorted(listacalculos):
                metodo = getattr(codigo_para_replace_en_el_latex.calculo, calculo, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)

            return pp, uu, vv, xx, yy, zz

        def figuras(pp, uu, vv, xx, yy, zz):
            """figuras"""
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            listafiguras = [
                m
                for m in dir(codigo_para_replace_en_el_latex.figura)
                if callable(getattr(codigo_para_replace_en_el_latex.figura, m))
                and not m.startswith("__")
            ]
            listafiguras = [s for s in listafiguras if not s.startswith("__")]
            for figura in listafiguras:
                metodo = getattr(codigo_para_replace_en_el_latex.figura, figura, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def imagenes(pp, uu, vv, xx, yy, zz):
            """imagenes"""
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            listaimagenes = [
                m
                for m in dir(codigo_para_replace_en_el_latex.imagen)
                if callable(getattr(codigo_para_replace_en_el_latex.imagen, m))
                and not m.startswith("__")
            ]
            listaimagenes = [s for s in listaimagenes if not s.startswith("__")]

            for imagen in listaimagenes:
                metodo = getattr(codigo_para_replace_en_el_latex.imagen, imagen, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def tablas(pp, uu, vv, xx, yy, zz):
            """tablas"""
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            listatablas = [
                m
                for m in dir(codigo_para_replace_en_el_latex.tabla)
                if callable(getattr(codigo_para_replace_en_el_latex.tabla, m))
                and not m.startswith("__")
            ]
            listatablas = [s for s in listatablas if not s.startswith("__")]
            for tabla in listatablas:
                # print(tabla)
                metodo = getattr(
                    codigo_para_replace_en_el_latex.tabla, tabla, None
                )  # Obtener el método o None si no existe
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)

            return pp, uu, vv, xx, yy, zz

        def replaceLatex(pp, uu, vv, xx, yy, zz):
            """
            Compila un archivo .tex a PDF, reemplaza un texto y devuelve el PDF en memoria.
            :param ruta_tex: Ruta al archivo .tex original.
            :return: Objeto BytesIO con el contenido del PDF.
            """

            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )

            pdf_path = Path(zz.nombre).with_suffix(".pdf")
            ruta_tex = f"{os.path.splitext(pdf_path)[0]}.tex"

            nombre_archivo = os.path.basename(ruta_tex)  # Nombre del archivo sin ruta

            # Crear un directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                ruta_tex_temp = os.path.join(temp_dir, nombre_archivo)

                contenido = file_tex(pp, uu, vv, xx, yy, zz)
                # tabla final con todas los parametros  y variables
                HojaDeDatos = "\n"

                for var_name, value in pp.__dict__.items():
                    print(f"pp.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\npp. {var_name}\n\npp.{var_name}"

                for var_name, value in uu.__dict__.items():
                    print(f"uu.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\nuu. {var_name}\n\nuu.{var_name}"

                for var_name, value in vv.__dict__.items():
                    print(f"vv.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\nvv. {var_name}\n\nvv.{var_name}"

                for var_name, value in xx.__dict__.items():
                    print(f"xx.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\nxx. {var_name}\n\nxx.{var_name}"

                for var_name, value in yy.__dict__.items():
                    print(f"yy.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\nyy. {var_name}\n\nyy.{var_name}"

                # for var_name, value in zz.__dict__.items():
                #     print(f"zz.{var_name}")
                #     HojaDeDatos = HojaDeDatos + f"\n\nzz. {var_name}\n\nzz.{var_name}"

                contenido = contenido.replace(f"HojaDeDatos", HojaDeDatos)

                if 1:
                    # Reemplazar directamente todas las variables en el texto
                    for var_name, value in pp.__dict__.items():
                        # print(f"pp.{var_name}")
                        contenido = contenido.replace(
                            f"pp.{var_name}",
                            f"{str(value)}",
                        )
                    for var_name, value in uu.__dict__.items():
                        # print(f"uu.{var_name}")
                        contenido = contenido.replace(f"uu.{var_name}", f"{str(value)}")

                    for var_name, value in vv.__dict__.items():
                        # print(f"vv.{var_name}")
                        contenido = contenido.replace(f"vv.{var_name}", f"{str(value)}")

                    for var_name, value in xx.__dict__.items():
                        # print(f"xx.{var_name}")
                        contenido = contenido.replace(f"xx.{var_name}", f"{str(value)}")

                    for var_name, value in yy.__dict__.items():
                        # print(f"yy.{var_name}")
                        contenido = contenido.replace(f"yy.{var_name}", f"{str(value)}")

                    for var_name, value in zz.__dict__.items():
                        # print(f"zz.{var_name}")
                        contenido = contenido.replace(f"zz.{var_name}", f"{str(value)}")

                with open(
                    "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.tex", "w", encoding="utf-8"
                ) as file:
                    file.write(contenido)

                # Guardar los cambios en el archivo temporal
                with open(ruta_tex_temp, "w", encoding="utf-8") as file:
                    file.write(contenido)

                # Función para ejecutar pdflatex dos veces
                def ejecutar_pdflatex():
                    print(
                        f"\n{100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                    )
                    if 1:
                        subprocess.run(
                            ["pdflatex", "-output-directory", temp_dir, ruta_tex_temp],
                            check=True,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )

                    return True

                # # para que se cre el indice debe estar descomentado
                # Ejecutar pdflatex dos veces para generar la TOC correctamente
                if not (ejecutar_pdflatex() and ejecutar_pdflatex()):
                    return None  # Si la compilación falla, se detiene el proceso
                # ejecutar_pdflatex()

                # Ruta del PDF generado
                pdf_generado = os.path.join(
                    temp_dir, nombre_archivo.replace(".tex", ".pdf")
                )

                # Leer el PDF en memoria
                with open(pdf_generado, "rb") as file:
                    pdf_en_memoria = BytesIO(file.read())

            return pdf_en_memoria

            #######################

        def xlsxhojaconvalorespordefecto(ruta_script):
            """ """
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            # si es un nuevo archivo tex copia los parametros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            codigo, titulo = comun.texfile.separar_y_mayusculizar(nombre)
            #
            ruta_archivo_excel = f"{ruta_script}/{codigo}.xlsx"
            if 1:
                df = pd.DataFrame(parametros(), columns=["Parámetro", "Valor"])
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    data = parametros()

                    # Crear un libro de trabajo de Excel
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja predeterminada creada por openpyxl
                    libro.remove(libro.active)

                    # Crear hojas con los nombres de los keys y escribir los datos
                    for sheet_name, sheet_data in data.items():
                        hoja = libro.create_sheet(
                            title=sheet_name
                        )  # Crear la hoja con el nombre del key
                        for row in sheet_data:
                            hoja.append(row)  # Agregar cada fila a la hoja

                    # Guardar el archivo Excel
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    # print(ruta_archivo_excel)
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # Cargar todas las hojas del archivo Excel
            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )

            # Crear el diccionario data
            data = {
                sheet_name: [df.columns.tolist()]
                + df.values.tolist()  # Incluir cabecera y datos
                for sheet_name, df in sheets.items()
            }

            return data

    class texfile:

        def normalizar_cadena(texto):
            if isinstance(texto, str):  # Asegurar que es un string
                texto = texto.lower()
                texto = "".join(
                    c
                    for c in unicodedata.normalize("NFD", texto)
                    if unicodedata.category(c) != "Mn"
                )
                # Capitalizar la primera letra después de un espacio
                texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
                # Eliminar espacios
                texto = re.sub(r"\s+", "", texto)
                # Mantener solo caracteres permitidos
                texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
            return texto

        def limpiar_valores(data):
            for key, values in data.items():
                for i in range(
                    1, len(values)
                ):  # Saltamos la primera fila (encabezados)
                    for j in range(
                        1, len(values[i])
                    ):  # Recorremos todas las columnas (excepto encabezados)
                        if isinstance(
                            values[i][j], str
                        ):  # Verificar si es una cadena antes de dividir
                            values[i][j] = (
                                values[i][j].split(";")[0].strip()
                            )  # Tomar solo el primer valor y limpiar espacios
            return data

        def separar_y_mayusculizar(texto):
            # Usamos una expresión regular para separar las mayúsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            # Separar la primera parte (mantenerla como está)
            primera_parte = partes[0]

            # El resto de las partes lo convertimos a mayúsculas y las unimos con espacios
            # resto = " ".join([parte.upper() for parte in partes[1:]])
            resto = " ".join([parte for parte in partes[1:]])

            return primera_parte, resto

        # Función para limpiar nombres de variables
        def sanitize_name(name):
            return re.sub(
                r"[^a-zA-Z0-9]", "", str(name)
            )  # Elimina cualquier carácter no alfanumérico


if __name__ == "__main__":

    # Copio la ruta del scripot para crear el excel
    ruta_script = os.path.dirname(
        os.path.abspath(__file__)
    )  # Obtener la carpeta sin el nombre del script
    #  pero vuelvo a cambiar para ser compatible con las direcciones de fastapi y la carpeta static para figuras

    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    os.chdir("../../../../")  # Cambia dos niveles arriba

    print(os.getcwd())  # Imprime el directorio de trabajo actual

    data = comun.main.xlsxhojaconvalorespordefecto(ruta_script)

    nombre_funcion = os.path.basename(__file__)[:-3]
    globals()[nombre_funcion](data)
