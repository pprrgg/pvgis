#!/usr/bin/env python3


def file_tex():
    return r"""


\documentclass[a4paper,10pt]{article}
%\documentclass[a4paper,10pt,twocolumn]{article}
\usepackage[top = 2cm, bottom = 2cm, left = 2cm, right = 2cm, asymmetric]{geometry} % Especificar los márgenes según la norma
\usepackage[spanish]{babel}    

\usepackage{graphicx}  % Para incluir imágenes si es necesario
\usepackage{amsmath, amssymb}  % Para expresiones matemáticas
\usepackage{fancyhdr}  % Para personalizar encabezados
\usepackage{chngcntr}  % Para cambiar la numeración de apartados y subsecciones
\usepackage{tocbibind}  % Para incluir bibliografía en la tabla de contenidos
\usepackage{appendix}  % Para el formato de anexos
\usepackage{lipsum}  % Para generar texto de relleno (dummy text)
\usepackage{geometry}  % Para personalizar los márgenes
\usepackage{multicol}  % Para columnas
\usepackage{titlesec}  % Para personalizar los títulos de las secciones
\usepackage[utf8]{inputenc}    
\usepackage[T1]{fontenc}       
\usepackage[pdftex,pdfencoding=auto, colorlinks=true, linkcolor=black, citecolor=black, filecolor=black, urlcolor=black]{hyperref}
\usepackage{tocloft}           
\usepackage{booktabs}          % Tablas profesionales
\usepackage{float} % Paquete necesario para usar la opción [H]
\usepackage{array}
\usepackage{longtable}
\usepackage{circuitikz}
\usepackage{tikz}
\usepackage{tikz-cd}
\usepackage{siunitx}



\usepackage{qrcode}            % Paquete para generar QR
\usepackage{tabularx} % Agregar en el preámbulo
\usepackage[absolute,overlay]{textpos} % Para posicionar objetos libremente

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{ss.titulo}  % Nombre del documento en el encabezado izquierdo
\fancyhead[C]{}  % Centro vacío
\fancyhead[R]{\thepage}  % Numeración de páginas en el encabezado derecho
% \renewcommand{\thepage}{\arabic{page}}  % Asegura la numeración de páginas en números arábigos

% Definición de los anexos
\newcommand{\annex}[1]{\section{Anexo #1} \addcontentsline{toc}{section}{Anexo #1}}
\addto\captionsspanish{%
  \renewcommand{\tablename}{Tabla}%
  \renewcommand{\listtablename}{Índice de tablas}%
}
\title{{\small Ref.:\uppercase{ss.codigo}}\\{\textbf{ss.titulo}}}

\author{
    \begin{minipage}{0.35\textwidth}
    \centering
    ss.iiParticipantesRolP1\\
    ss.iiParticipantesNombrerazonSocialP1 \\
    ss.iiParticipantesTitulacionP1\\
    ss.iiParticipantesTelefonoP1\\
    ss.iiParticipantesCorreoElectronicoP1\\
    \end{minipage}%
    \hfill
    \begin{minipage}{0.35\textwidth}
    \centering
    ss.iiParticipantesRolP2\\
    ss.iiParticipantesNombrerazonSocialP2 \\
    ss.iiParticipantesTitulacionP2\\
    ss.iiParticipantesTelefonoP2\\
    ss.iiParticipantesCorreoElectronicoP2\\    
    \end{minipage}%
    \hfill
    \begin{minipage}{0.35\textwidth}
    \centering
    ss.iiParticipantesRolP3\\
    ss.iiParticipantesNombrerazonSocialP3 \\
    ss.iiParticipantesTitulacionP3\\
    ss.iiParticipantesTelefonoP3\\
    ss.iiParticipantesCorreoElectronicoP3\\    
    \end{minipage}%
}

\date{\today}
\newcommand{\MostrarVariablesAlFinal}{}


\begin{document}
\begin{Form}

	% Mostrar el título
	\maketitle
	% \onecolumn

	\begin{abstract}

En este documento se ha realizado un análisis completo de un sistema eléctrico con presencia de armónicos. A partir de los datos proporcionados, se calculó la corriente fundamental (\(I_1 \approx \SI{160}{A}\)) y el THD de corriente (\(\text{THD}_i \approx 32.8\%\)), evidenciando una distorsión significativa debido a los armónicos de orden 3 y 9. Además, se comparó el coseno de phi (\(\cos \phi = 0.853\)) con el factor de potencia (\(FP = 0.811\)), observándose que el FP es menor debido al efecto de los armónicos, los cuales incrementan la corriente eficaz total. Finalmente, se generó una representación gráfica de la señal de corriente, mostrando la distorsión causada por los armónicos. Estos resultados destacan la importancia de implementar medidas de mitigación de armónicos para mejorar la calidad de la energía y el factor de potencia del sistema.
        	\end{abstract}

	% Iniciar el formato de dos columnas

	% Reemplazamos el entorno IEEEkeywords por una lista de palabras clave
	\noindent\textbf{Palabras clave:} Potencia activa, Potencia reactiva, Potencia aparente, Corriente fundamental, Armónicos de corriente, THD (Distorsión Armónica Total), Factor de potencia, Coseno de phi (\(\cos \phi\)), Calidad de la energía, Distorsión armónica, Mitigación de armónicos, Análisis de circuitos, Sistemas eléctricos, Corriente eficaz, Frecuencia fundamental, Armónicos de orden 3 y 9


	\tableofcontents  % Tabla de contenidos
	\listoffigures    % Lista de figuras
	%\listoftables     % Lista de tablas

	\newpage



    








































\section{Datos iniciales}
Los datos proporcionados son los siguientes:
\begin{itemize}
    \item Tensión (\(U\)): \SI{230}{V}
    \item Potencia activa (\(P\)): \SI{31.40}{kW}
    \item Potencia reactiva (\(Q\)): \SI{19.19}{kVAr}
    \item Armónicos de corriente:
        \begin{itemize}
            \item \(f_3 = 20\%\)
            \item \(f_5 = 0\%\)
            \item \(f_7 = 0\%\)
            \item \(f_9 = 26\%\)
        \end{itemize}
\end{itemize}



\section{Cálculo de la corriente fundamental (\(I_1\))}
La corriente fundamental se calcula a partir de la potencia aparente (\(S\)) y la tensión (\(U\)).




\subsection{Paso 1: Cálculo de la potencia aparente (\(S\))}
\[
S = \sqrt{P^2 + Q^2} = \sqrt{(31.40)^2 + (19.19)^2} = \sqrt{985.96 + 368.2561} = \sqrt{1354.2161} \approx \SI{36.80}{kVA}
\]

\subsection{Paso 2: Cálculo de la corriente aparente (\(I\))}
\[
I = \frac{S}{U} = \frac{36.80 \times 10^3}{230} \approx \SI{160}{A}
\]

\subsection{Paso 3: Corriente fundamental (\(I_1\))}
La corriente fundamental es aproximadamente igual a la corriente aparente:
\[
I_1 \approx \SI{160}{A}
\]

\section{Cálculo del THD de corriente}
El THD de corriente se calcula como:
\[
\text{THD}_i = \frac{\sqrt{I_3^2 + I_5^2 + I_7^2 + I_9^2}}{I_1} \times 100 \%
\]

\subsection{Paso 1: Cálculo de las corrientes armónicas}
\[
I_3 = I_1 \times \frac{f_3}{100} = 160 \times 0.20 = \SI{32}{A}
\]
\[
I_5 = I_1 \times \frac{f_5}{100} = 160 \times 0 = \SI{0}{A}
\]
\[
I_7 = I_1 \times \frac{f_7}{100} = 160 \times 0 = \SI{0}{A}
\]
\[
I_9 = I_1 \times \frac{f_9}{100} = 160 \times 0.26 = \SI{41.6}{A}
\]

\subsection{Paso 2: Cálculo del THD de corriente}
\[
\text{THD}_i = \frac{\sqrt{32^2 + 0^2 + 0^2 + 41.6^2}}{160} \times 100 \% = \frac{\sqrt{1024 + 0 + 0 + 1730.56}}{160} \times 100 \% 
\]
\[
\text{THD}_i = \frac{\sqrt{2754.56}}{160} \times 100 \% = \frac{52.48}{160} \times 100 \% \approx 32.8 \%
\]

ss.f01figuraejemplo


\section{Cálculo y comparación del coseno de phi (\(\cos \phi\)) y el factor de potencia (FP)}

\subsection{Cálculo del coseno de phi (\(\cos \phi\))}
\[
\cos \phi = \frac{P}{S} = \frac{31.40}{36.80} = 0.853
\]

\subsection{Cálculo del factor de potencia (FP)}
El factor de potencia se calcula como:
\[
FP = \frac{P}{U \cdot I_{\text{ef}}}
\]

\subsubsection{Paso 1: Cálculo de la corriente eficaz total (\(I_{\text{ef}}\))}
\[
I_{\text{ef}} = \sqrt{I_1^2 + I_3^2 + I_5^2 + I_7^2 + I_9^2} = \sqrt{160^2 + 32^2 + 0^2 + 0^2 + 41.6^2}
\]
\[
I_{\text{ef}} = \sqrt{25600 + 1024 + 0 + 0 + 1730.56} = \sqrt{28354.56} \approx \SI{168.4}{A}
\]


\subsubsection{Paso 2: Cálculo del factor de potencia (FP)}
\[
FP = \frac{31400}{230 \times 168.4} = \frac{31400}{38732} \approx 0.811
\]

\subsection{Comparación y análisis}
\begin{itemize}
    \item \(\cos \phi = 0.853\)
    \item \(FP = 0.811\)
\end{itemize}

El factor de potencia (FP) es menor que el \(\cos \phi\) debido a la presencia de armónicos, que aumentan la corriente eficaz total y reducen el factor de potencia. La distorsión armónica (THD del 32.8\%) justifica esta diferencia.






























\section{Dimensionado del ﬁltro activo necesario. }

\subsection{Datos proporcionados}
Los datos iniciales para el dimensionamiento del filtro activo son los siguientes:
\begin{itemize}
    \item THDv (Tasa de Distorsión Armónica de Tensión): \SI{2.6}{\%}
    \item THDi (Tasa de Distorsión Armónica de Corriente): \SI{32.8}{\%}
    \item Corriente fundamental (\(I_1\)): \SI{160}{A}
    \item Armónicos significativos:
        \begin{itemize}
            \item 3er armónico: \SI{20}{\%}
            \item 9no armónico: \SI{26}{\%}
        \end{itemize}
\end{itemize}

\subsection{Cálculo de la corriente armónica total a compensar}
La corriente armónica total (\(I_{\text{arm}}\)) se calcula a partir del THDi y la corriente fundamental:
\[
I_{\text{arm}} = I_1 \times \frac{\text{THDi}}{100} = 160 \times \frac{32.8}{100} = \SI{52.48}{A}
\]

\subsection{Selección del filtro activo}
Para asegurar un correcto funcionamiento, se añade un margen de seguridad del 30\% a la corriente armónica:
\[
I_{\text{filtro}} = I_{\text{arm}} \times 1.3 = 52.48 \times 1.3 \approx \SI{68.22}{A}
\]

\subsection{Modelo recomendado}
Según el catálogo de \textbf{Circutor}, se recomienda utilizar un filtro activo con una capacidad de \SI{75}{A} o superior. Algunos modelos adecuados son:
\begin{itemize}
    \item \textbf{Circutor AFQevo}: Disponible en capacidades de \SI{30}{A}, \SI{50}{A}, \SI{75}{A}, \SI{100}{A}, etc.
    \item \textbf{Circutor AFQmini}: Para aplicaciones más pequeñas, pero en este caso no sería suficiente.
\end{itemize}

\subsection{Consideraciones adicionales}
\begin{itemize}
    \item El valor de THDv (\SI{2.6}{\%}) está dentro de los límites aceptables según la norma \textbf{IEEE 519} (THDv < \SI{5}{\%} para sistemas de baja tensión).
    \item El filtro activo debe estar configurado para mitigar específicamente los armónicos de orden 3 y 9.
    \item Además de reducir los armónicos, el filtro activo mejorará el factor de potencia, acercándolo a 1.
\end{itemize}

\subsection{Recomendación final}
Se recomienda instalar un \textbf{filtro activo Circutor AFQevo de \SI{75}{A}}, ya que:
\begin{itemize}
    \item Cubre la corriente armónica a compensar (\SI{52.48}{A}) con un margen de seguridad.
    \item Es capaz de mitigar los armónicos de orden 3 y 9.
    \item Mejora el factor de potencia y reduce el THDi a valores aceptables.
\end{itemize}












\section{Modelo de ﬁltro activo de la “Serie AFQm” de la casa Circutor, para
una instalación trifásica con neutro, si se considera que la fase analizada es la
que ene un mayor THDi y una mayor corriente fundamental, y los valores
considerados, son los más desfavorables alcanzados de forma habitual para el
periodo analizado.}
\subsection{Datos de la fase analizada}
Los valores más desfavorables obtenidos durante el periodo analizado (7 días) son:
\begin{itemize}
    \item THDi (Tasa de Distorsión Armónica de Corriente): \SI{32.8}{\%}
    \item Corriente fundamental (\(I_1\)): \SI{160}{A}
    \item Armónicos significativos:
        \begin{itemize}
            \item 3er armónico: \SI{20}{\%}
            \item 9no armónico: \SI{26}{\%}
        \end{itemize}
\end{itemize}

\subsection{Cálculo de la corriente armónica total a compensar}
La corriente armónica total (\(I_{\text{arm}}\)) se calcula como:
\[
I_{\text{arm}} = I_1 \times \frac{\text{THDi}}{100} = 160 \times \frac{32.8}{100} = \SI{52.48}{A}
\]

\subsection{Margen de seguridad}
Se añade un margen de seguridad del 30\%:
\[
I_{\text{filtro}} = I_{\text{arm}} \times 1.3 = 52.48 \times 1.3 \approx \SI{68.22}{A}
\]

\subsection{Selección del modelo de la Serie AFQm}
El modelo recomendado es el \textbf{Circutor AFQm 100-4/75}:
\begin{itemize}
    \item Capacidad nominal: \SI{75}{A}
    \item Adecuado para sistemas trifásicos con neutro
    \item Mitiga armónicos de orden 3 y 9
    \item Mejora el factor de potencia y reduce el THDi
\end{itemize}

\subsection{Recomendación final}
Se recomienda instalar un \textbf{filtro activo Circutor AFQm 100-4/75} para esta instalación trifásica con neutro, ya que:
\begin{itemize}
    \item Cubre la corriente armónica a compensar (\SI{68.22}{A}) con un margen de seguridad.
    \item Mitiga los armónicos de orden 3 y 9.
    \item Mejora la calidad de la energía y el factor de potencia.
\end{itemize}



















































\section{Bibliografía}
\begin{itemize}
    \item \href{https://www.ingenierosformacion.com/}{F0JV118LXMT1, Curso 'Diseño avanzado de instalaciones eléctricas de Baja Tensión. COGITI, José Luis Rodríguez' }
    \item \href{https://circutor.com/soporte/formacion/notebooks/armonicos-electricos/}{CIRCUTOR. Técnicas de compensación y filtrado de perturbaciones armónicas }
    \item \href{https://circutor.com/articulos/armonicos-electricos-definicion-problematica-soluciones/}{CIRCUTOR. Armónicos eléctricos: definición  problemática  soluciones }
\end{itemize}








    






























	% P A R A M E T R O S 
	%---------------------

	\TextField[name=I0 ,width=0cm]{}
	\TextField[name=Bt,width=0cm]{}


	%==========================================




\end{Form}

\end{document}


"""


class codigo_al_final_del_latex:
    class parametros:
        def parametros():
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            parametros = r"""
            [
            {
                "Descripción": " Potencia contratada",
                "valor": 66,
                "Unidad": "kW"
            },
            {
                "Descripción": "Duración indicativa de la actuación",
                "valor": 15,
                "Unidad": "años"
            },
            {
                "Descripción": "Préstamo recibido",
                "valor": 50000,
                "Unidad": "€"
            },
            {
                "Descripción": "Tasa de interés del préstamo",
                "valor": 5,
                "Unidad": "\\%"
            },
            {
                "Descripción": "jojoj",
                "valor": 5,
                "Unidad": "\\%"
            }
            ]
            """
            datos = json.loads(parametros)
            df = pd.DataFrame(datos)
            df.iloc[:, 0] = "" + df.iloc[:, 0] + r" [" + df.iloc[:, 2] + r"]"
            return df

    class calculos:
        def c01parametros(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            print(99999999999999999999999999999999999999999999999999999999999999999)
            print(ss.nombre)
            print(99999999999999999999999999999999999999999999999999999999999999999)
            for param, value in ss.data[ss.codigo]:
                param = zz.texfile.normalizar_cadena(param)
                print(f"ss.{param} = {value}")

                # Intentar convertir value a float, si no es posible, dejarlo igual
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass  # Si no se puede convertir, mantiene el valor original

                setattr(ss, param, value)

            for param, value in ss.data[ss.codigo][
                1:
            ]:  # Empezamos desde el índice 1 para omitir los encabezados

                # Intentamos convertir a float si es posible
                try:
                    float_value = float(value)
                    setattr(ss, param, float_value)
                    globals()[param] = getattr(ss, param)

                except (ValueError, TypeError):
                    # Si no es convertible a float, lo dejamos como está
                    pass
            return ss

        def c02calculoejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Crear un DataFrame con números aleatorios
            np.random.seed(42)  # Para reproducibilidad
            df = pd.DataFrame(np.random.rand(4, 4))

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["gsdfg"] = total_row

            df = df.copy()

            return ss

    class tablas:
        """"""

        def t03totalfactura(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import pandas as pd

            data_totales = {
                "Concepto": [
                    "Subtotal",
                    "Impuesto eléctrico",
                    "Alquiler contador",
                    "Subtotal",
                    "IVA ",
                    "TOTAL FACTURA",
                ],
                " ": [
                    "           ",
                    "0,5\\%",
                    "30 días",
                    "",
                    "5\\%",
                    "",
                ],
                "€/mes": [19.37, 0.10, 0.81, 20.28, 1.01, 21.29],
            }

            df = pd.DataFrame(data_totales)
            # df.set_index("Energía consumida", inplace=True)  # Convertir la primera columna en índice

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["TOTAL"] = total_row
            # Formatear solo las columnas deseadas a 2 decimales antes de exportar

            df["€/mes"] = df["€/mes"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{2.5cm}}|" + "l" * (len(df.columns) - 1),
            )

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.45\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Total factura}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

        def t02tablaPotenciacontratada(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import pandas as pd

            # Crear el DataFrame
            datos = {
                "Potencia contratada": [
                    "PUNTA",
                    "VALLE",
                    "Margen comercialización fijo",
                ],
                "kW": [
                    ss.PotenciaContratadakw,
                    ss.PotenciaContratadakw,
                    ss.PotenciaContratadakw,
                ],
                "€/kW/año": [26.164043, 1.143132, 3.113000],
                "€/mes": [10.75, 0.47, 1.28],
            }

            df = pd.DataFrame(datos)
            # df.set_index("Potencia contratada", inplace=True)  # Convertir la primera columna en índice

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["TOTAL"] = total_row
            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]
            df.loc["gsdfg"] = total_row

            # Formatear solo las columnas deseadas a 2 decimales antes de exportar
            df["kW"] = df["kW"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )
            df["€/mes"] = df["€/mes"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{2.5cm}}|" + "l" * (len(df.columns) - 1),
            )

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.45\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Potencia}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

        def t02tablaEnergíaconsumida(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import pandas as pd

            # Crear el DataFrame con los datos
            datos = {
                "Energía consumida": [
                    "PUNTA",
                    "LLANO",
                    "VALLE",
                    "Coste energía",
                    "Compensación excedentes FV",
                ],
                "kWh": [13.5, 10.5, 108.0, 132.0, "-"],
                "€/kWh": [0.074409, 0.028470, 0.003034, 0.150000, "-"],
                "€/mes": [1.00, 0.30, 0.33, 19.80, -14.56],
            }

            df = pd.DataFrame(datos)

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]
            df.loc["gsdfg"] = total_row

            # Formatear solo las columnas deseadas a 2 decimales antes de exportar
            df["kWh"] = df["kWh"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )
            df["€/mes"] = df["€/mes"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{2.5cm}}|" + "l" * (len(df.columns) - 1),
            )

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.45\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Energía consumida}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

        def t00tablaparametros(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            """poner en el valor el simbolo sin acentos guiones ni nada, que sera el valor para los calculos"""

            # df = solo_desarrollo.parametros.parametros()

            par = ss.data[next(iter(ss.data))]
            df = pd.DataFrame(par[1:], columns=par[0])

            # print(df)

            # Exportar a LaTeX
            latex_code = df.to_latex(index=False, escape=False)

            df.iloc[:, 1] = df.iloc[:, 1].apply(
                lambda x: f"{float(x):.2f}".replace(".", ",") if type(x) == float else x
            )  # Dos decimales

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{5cm}}|" + "l" * (len(df.columns) - 1),
                position="H",
                caption="t00tablaparametros",
            )

            return latex_code

        def t99tablaresultados(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            titulo = "hjklhlkjhkljjhhj"
            """poner en el valor el simbolo sin acentos guiones ni nada, que sera el valor para los calculos"""
            parametros = r"""
            [
            {
                "Descripción": "Período de recuperación",
                "valor": 11,
                "Unidad": "años"
            },
            {
                "Descripción": "Valor Actual Neto",
                "valor": 59063.66,
                "Unidad": "€"
            },
            {
                "Descripción": "Tasa Interna de Retorno",
                "valor": 13.09,
                "Unidad": "\\%"
            }
            ]
            """
            datos = json.loads(parametros)
            df = pd.DataFrame(datos)
            df.iloc[:, 0] = "" + df.iloc[:, 0] + r" [" + df.iloc[:, 2] + r"]"

            # Exportar a LaTeX
            latex_code = df.to_latex(index=False, escape=False)

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{3cm}}|" + "l" * (len(df.columns) - 1),
                position="H",
                caption="t99tablaresultados",
            )

            return latex_code

        def xxt01tablaejemplo(ss):
            print(f"\n {100*'+'}{__class__.__name__}.{sys._getframe().f_code.co_name}")

            df = ss.df.copy()

            # Generar la tabla en formato LaTeX
            latex_code = df.to_latex(index=False)

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.4\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Tabla escalada al ancho de la página}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

    class figuras:
        """"""

        def f01figuraejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # 1. Crear un archivo temporal para la imagen

            titulo = "Señal de corriente con armónicos"
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image:
                temp_image_path2 = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal





            import numpy as np
            import matplotlib.pyplot as plt

            # Parámetros de la señal
            f1 = 50  # Frecuencia fundamental (Hz)
            I1 = 160  # Amplitud de la corriente fundamental (A)
            I3 = 32  # Amplitud del 3er armónico (A)
            I5 = 0  # Amplitud del 5to armónico (A)
            I7 = 0  # Amplitud del 7mo armónico (A)
            I9 = 41.6  # Amplitud del 9no armónico (A)

            # Tiempo de muestreo
            t = np.linspace(0, 0.02, 1000)  # Un ciclo de 20 ms (50 Hz)

            # Señales individuales
            fundamental = I1 * np.sin(2 * np.pi * f1 * t)
            armonico_3 = I3 * np.sin(2 * np.pi * 3 * f1 * t)
            armonico_5 = I5 * np.sin(2 * np.pi * 5 * f1 * t)
            armonico_7 = I7 * np.sin(2 * np.pi * 7 * f1 * t)
            armonico_9 = I9 * np.sin(2 * np.pi * 9 * f1 * t)

            # Señal total (suma de la fundamental y los armónicos)
            corriente_total = fundamental + armonico_3 + armonico_5 + armonico_7 + armonico_9

            # Gráfica
            plt.figure(figsize=(10, 6))
            plt.plot(t, corriente_total, label="Corriente total", color="blue")
            plt.plot(t, fundamental, label="Fundamental (50 Hz)", linestyle="--", color="red")
            plt.plot(t, armonico_3, label="3er armónico (150 Hz)", linestyle=":", color="green")
            plt.plot(t, armonico_9, label="9no armónico (450 Hz)", linestyle="-.", color="purple")

            # Configuración de la gráfica
            # plt.title("Señal de corriente con armónicos")
            plt.xlabel("Tiempo [s]")
            plt.ylabel("Corriente [A]")
            plt.grid(True)
            plt.legend(loc="upper right")
            plt.show()






            plt.savefig(temp_image_path2, format="png", dpi=300)

            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + """\\begin{figure}[H]
                """
                + f"""
                \\includegraphics[width=.985\\textwidth]{{{temp_image_path2}}}
                \\caption{{{titulo}}}
                """
                + """
                \\label{fig:dfassssdfsa}
                \\end{figure}
                """
            )

            return latex_code

        def f02figuraejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # 1. Crear un archivo temporal para la imagen
            titulo = "Flujo de Caja y Acumulado"
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image:
                temp_image_path1 = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal

            df = pd.DataFrame(np.random.rand(4, 4))
            # .......
            #
            fig, ax = plt.subplots(figsize=(6, 5))
            df.plot(ax=ax)
            #
            # .......

            ax.grid(axis="y", linestyle="--", alpha=0.7)
            plt.xticks(rotation=50)
            ax.set_ylabel("€")  # Etiqueta en el eje y

            plt.savefig(temp_image_path1, format="png", dpi=300)

            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + """\\begin{figure}[H]
                """
                + f"""
                \\includegraphics[width=.5\\textwidth]{{{temp_image_path1}}}
                \\caption{{{titulo}}}
                """
                + """
                \\label{fig:dfasdfsa}
                \\end{figure}
                """
            )

            return latex_code


class zz:
    class main:

        def parametros_latex(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Para cada hoja
            for key, values in ss.data.items():
                # Crear un DataFrame a partir de los datos
                df = pd.DataFrame(
                    values[1:], columns=values[0]
                )  # Crear DataFrame con la primera fila como encabezado
                df = df.set_index(df.columns[0])
                variable = f"hh {key}"
                variable = zz.texfile.normalizar_cadena(variable)
                setattr(ss, variable, df)  # Asignar como atributo de ss

                # Crear un DataFrame a partir de las columnas del DataFrame
                for col in df.columns:
                    variable = f"cc {key} {col}"
                    variable = zz.texfile.normalizar_cadena(variable)
                    setattr(ss, variable, df[[col]])

                # crear las variables con claves formadas por <indice><columna>
                for index, row in df.iterrows():
                    for col in df.columns:
                        variable = f"ii {key} {index} {col}"
                        variable = zz.texfile.normalizar_cadena(variable)
                        setattr(ss, variable, row[col])
                print(ss.__dict__)
            return ss

        def parametros_formulario(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # excel a ss sapacio de estado
            for param, value in ss.data[ss.nombre]:
                param = zz.texfile.normalizar_cadena(param)
                # Intentar convertir value a float, si no es posible, dejarlo igual
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass  # Si no se puede convertir, mantiene el valor original
                setattr(ss, param, value)
            return ss

        def main(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # segun si es un formulario o un documento nuevo
            pdf_path = Path(ss.nombre).with_suffix(".pdf")

            # if os.path.exists(f"{os.path.splitext(pdf_path)[0]}.tex"):
            if ss.archivo_nuevo:
                # print("El archivo .tex existe.")
                tex_path = f"{os.path.splitext(pdf_path)[0]}.tex"
                ss = zz.main.parametros_latex(ss)

                """calculos"""
                listacalculos = [
                    m
                    for m in dir(codigo_al_final_del_latex.calculos)
                    if callable(getattr(codigo_al_final_del_latex.calculos, m))
                    and not m.startswith("__")
                ]
                listacalculos = [s for s in listacalculos if not s.startswith("xx")]
                for calculo in sorted(listacalculos):
                    metodo = getattr(codigo_al_final_del_latex.calculos, calculo, None)
                    ss = metodo(ss)

                output_buffer = zz.texfile.replaceLatex(tex_path, ss)

            else:
                print("El archivo .tex no existe.")
                # Leer el archivo PDF desde el disco
                ss = zz.main.parametros_formulario(ss)

                """calculos"""
                listacalculos = [
                    m
                    for m in dir(codigo_al_final_del_latex.calculos)
                    if callable(getattr(codigo_al_final_del_latex.calculos, m))
                    and not m.startswith("__")
                ]
                listacalculos = [s for s in listacalculos if not s.startswith("xx")]
                for calculo in sorted(listacalculos):
                    metodo = getattr(codigo_al_final_del_latex.calculos, calculo, None)
                    ss = metodo(ss)

                with open(pdf_path, "rb") as file:
                    pdf_content = file.read()
                output_buffer = BytesIO(pdf_content)
                output_buffer.seek(0)
                output_buffer = zz.formulario.rellenoFormulario(output_buffer, ss)

            return output_buffer

            #######################

        def xlsxhojaconvalorespordefecto(df):
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # si es un nuevo archivo tex copia los parametros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            codigo, titulo = separar_y_mayusculizar(nombre)
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria

            #
            ruta_archivo_excel = f"./{codigo}.xlsx"
            if 1:
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja por defecto (llamada 'Sheet')
                    hoja_por_defecto = libro["Sheet"]
                    libro.remove(hoja_por_defecto)

                    hoja = libro.create_sheet(title=codigo)

                    # Seleccionar la hoja activa
                    # hoja = libro.active
                    # Guardar el nuevo libro de trabajo
                    libro.save(ruta_archivo_excel)

                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    hoja = libro[codigo]

                    # Borrar todo el contenido de la hoja
                    for fila in hoja.iter_rows():
                        for celda in fila:
                            celda.value = None

                    # Insertar los datos del DataFrame en la hoja de Excel
                    # Insertar la cabecera
                    hoja["A1"] = "Parámetro"
                    hoja["B1"] = "Valor"

                    # Insertar los datos del DataFrame en la hoja de Excel
                    for i, (parametro, valor) in enumerate(
                        zip(df["Descripción"], df["valor"]), start=2
                    ):
                        # hoja[f"A{i}"] = parametro.replace("_", "").replace(" ", "")
                        hoja[f"A{i}"] = parametro
                        hoja[f"B{i}"] = valor

                    # Guardar el archivo con los cambios
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    # print(ruta_archivo_excel)
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # Cargar el archivo Excel

            # Cargar todas las hojas del archivo Excel
            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )

            # Crear el diccionario data
            data = {
                sheet_name: [df.columns.tolist()]
                + df.values.tolist()  # Incluir cabecera y datos
                for sheet_name, df in sheets.items()
            }

            return data

    class formulario:
        def rellenoFormulario(pdf_buffer, ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            pdf_document = fitz.open(stream=pdf_buffer, filetype="pdf")
            output_buffer = BytesIO()
            variables = vars(ss)
            # print("CAMPOS DEL FORMULARIO")
            # Recorrer todas las páginas del PDF
            for page_num in range(len(pdf_document)):
                page = pdf_document.load_page(page_num)
                # Iterar sobre los widgets (campos de formulario)
                for widget in page.widgets():
                    field_name = widget.field_name  # Obtener el ss.nombre del campo
                    field_name = zz.texfile.normalizar_cadena(field_name)
                    # print(field_name)
                    if (
                        field_name in variables
                    ):  # Si el ss.nombre del campo está en el SimpleNamespace
                        widget.field_value = str(
                            variables[field_name]
                        )  # Asignar el valor
                        try:
                            widget.update()  # Actualizar el widget
                        except Exception as e:
                            """"""
                            print(f"Error al actualizar el campo {field_name}: {e}")
            pdf_document.save(output_buffer)
            pdf_document.close()
            output_buffer.seek(0)
            return output_buffer

    class texfile:

        def replaceLatex(ruta_tex, ss):
            """
            Compila un archivo .tex a PDF, reemplaza un texto y devuelve el PDF en memoria.
            :param ruta_tex: Ruta al archivo .tex original.
            :return: Objeto BytesIO con el contenido del PDF.
            """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            ruta_tex_absoluta = os.path.abspath(
                ruta_tex
            )  # Obtener la ruta absoluta del archivo original
            nombre_archivo = os.path.basename(ruta_tex)  # Nombre del archivo sin ruta

            # Crear un directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                ruta_tex_temp = os.path.join(temp_dir, nombre_archivo)

                contenido = file_tex()

                ##########################################

                if 1:
                    """figuras"""
                    listafiguras = [
                        m
                        for m in dir(codigo_al_final_del_latex.figuras)
                        if callable(getattr(codigo_al_final_del_latex.figuras, m))
                        and not m.startswith("__")
                    ]
                    listafiguras = [s for s in listafiguras if not s.startswith("xx")]
                    for figura in listafiguras:
                        metodo = getattr(
                            codigo_al_final_del_latex.figuras, figura, None
                        )
                        contenido = contenido.replace(f"ss.{figura}", metodo(ss))

                if 1:
                    """tablas"""
                    listatablas = [
                        m
                        for m in dir(codigo_al_final_del_latex.tablas)
                        if callable(getattr(codigo_al_final_del_latex.tablas, m))
                        and not m.startswith("__")
                    ]
                    listatablas = [s for s in listatablas if not s.startswith("xx")]
                    for tabla in listatablas:
                        # print(tabla)
                        metodo = getattr(
                            codigo_al_final_del_latex.tablas, tabla, None
                        )  # Obtener el método o None si no existe
                        contenido = contenido.replace(f"ss.{tabla}", metodo(ss))

                if 1:
                    """variables"""
                    # Reemplazo usando .replace() para cada variable
                    for variable, valor in vars(ss).items():
                        if (
                            valor is None
                            or (isinstance(valor, str) and valor.strip() == "")
                            or (isinstance(valor, pd.DataFrame) and valor.empty)
                            or (isinstance(valor, pd.Series) and valor.isna().all())
                        ):
                            valor = "%"

                        contenido = contenido.replace(f"ss.{variable}", str(valor))
                        print(f"ss.{variable}")

                # solo depuracion
                # solo depuracion
                # solo depuracion
                # solo depuracion
                with open(
                    "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.tex", "w", encoding="utf-8"
                ) as file:
                    file.write(contenido)
                # solo depuracion
                # solo depuracion
                # solo depuracion
                # solo depuracion

                ##########################################

                # Guardar los cambios en el archivo temporal
                with open(ruta_tex_temp, "w", encoding="utf-8") as file:
                    file.write(contenido)

                # Función para ejecutar pdflatex dos veces
                def ejecutar_pdflatex():
                    print(
                        f"\n{100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                    )
                    if 1:
                        subprocess.run(
                            ["pdflatex", "-output-directory", temp_dir, ruta_tex_temp],
                            check=True,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )

                    return True

                # # para que se cre el indice debe estar descomentado
                # Ejecutar pdflatex dos veces para generar la TOC correctamente
                if not (ejecutar_pdflatex() and ejecutar_pdflatex()):
                    return None  # Si la compilación falla, se detiene el proceso
                # ejecutar_pdflatex()

                # Ruta del PDF generado
                pdf_generado = os.path.join(
                    temp_dir, nombre_archivo.replace(".tex", ".pdf")
                )

                # Leer el PDF en memoria
                with open(pdf_generado, "rb") as file:
                    pdf_en_memoria = BytesIO(file.read())

            return pdf_en_memoria

        def normalizar_cadena(texto):
            if isinstance(texto, str):  # Asegurar que es un string
                texto = texto.lower()
                texto = "".join(
                    c
                    for c in unicodedata.normalize("NFD", texto)
                    if unicodedata.category(c) != "Mn"
                )
                # Capitalizar la primera letra después de un espacio
                texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
                # Eliminar espacios
                texto = re.sub(r"\s+", "", texto)
                # Mantener solo caracteres permitidos
                texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
            return texto


from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from PyPDF2 import PdfReader, PdfWriter
from io import BytesIO
import os
import matplotlib.pyplot as plt
import matplotlib

matplotlib.use("Agg")

import io
import base64
import pandas as pd
import fitz  # PyMuPDF
from types import SimpleNamespace
import subprocess
import shutil
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata

import numpy as np
import sys


router = APIRouter()


def separar_y_mayusculizar(texto):
    # Usamos una expresión regular para separar las mayúsculas
    partes = re.findall(r"[A-Z][^A-Z]*", texto)

    # Separar la primera parte (mantenerla como está)
    primera_parte = partes[0]

    # El resto de las partes lo convertimos a mayúsculas y las unimos con espacios
    resto = " ".join([parte.upper() for parte in partes[1:]])

    return primera_parte, resto


router = APIRouter()


@router.post("/Btx040")
def Btx040FiltroDeArmónicos(
    data: dict,
):

    # parra que en la primera carga, los campos que tgiene valores separados por ; para los selectores coja el primer valor por defecto, tambien 
    # debera hacerlo si se da a recalcular sin seleccionar un valor 
    for key, values in data.items():
        for i in range(1, len(values)):  # Saltamos la primera fila (encabezados)
            if isinstance(values[i][1], str):  # Verificar si es una cadena antes de dividir
                values[i][1] = values[i][1].split(';')[0]  # Tomar solo el primer valor
    # parra que en la primera carga, los campos que tgiene valores separados por ; para los selectores coja el primer valor por defecto, tambien 
    # debera hacerlo si se da a recalcular sin seleccionar un valor 
    for key, values in data.items():
        for i in range(1, len(values)):  # Saltamos la primera fila (encabezados)
            if isinstance(values[i][1], str):  # Verificar si es una cadena antes de dividir
                values[i][1] = values[i][1].split(';')[0]  # Tomar solo el primer valor
    print(data)

    ss = SimpleNamespace()

    # Ejemplo de uso
    ss.nombre = sys._getframe().f_code.co_name

    ss.codigo, ss.titulo = separar_y_mayusculizar(ss.nombre)

    ss.archivo_nuevo = True

    ss.data = data

    output_buffer = zz.main.main(ss)
    # solo depuracion
    # solo depuracion
    # solo depuracion
    with open(f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf", "wb") as f:
        f.write(output_buffer.getvalue())
    # solo depuracion
    # solo depuracion
    # solo depuracion
    # devuelve a fastapi el archivo
    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


if __name__ == "__main__":

    # Cambiar al directorio del script
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    df = codigo_al_final_del_latex.parametros.parametros()
    data = zz.main.xlsxhojaconvalorespordefecto(df)
    nombre_funcion = os.path.basename(__file__)[:-3]
    globals()[nombre_funcion](data)
