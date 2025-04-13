#!/usr/bin/env python3


def file_tex():
    return r"""


\documentclass[a4paper,10pt,twocolumn]{article}
\usepackage[top = 2cm, bottom = 2cm, left = 2cm, right = 2cm, asymmetric]{geometry} % Especificar los márgenes según la norma
\usepackage[spanish]{babel}    

\usepackage{graphicx}  % Para incluir imágenes si es necesario
\usepackage{amsmath, amssymb}  % Para expresiones matemáticas
\usepackage{fancyhdr}  % Para personalizar encabezados
\usepackage{chngcntr}  % Para cambiar la numeración de apartados y subsecciones
\usepackage{tocbibind}  % Para incluir bibliografía en la tabla de contenidos
\usepackage{appendix}  % Para el formato de anexos
\usepackage{lipsum}  % Para generar texto de relleno (dummy text)
\usepackage{geometry}  % Para personalizar los márgenes
\usepackage{multicol}  % Para columnas
\usepackage{titlesec}  % Para personalizar los títulos de las secciones
\usepackage[utf8]{inputenc}    
\usepackage[T1]{fontenc}       
\usepackage[pdftex,pdfencoding=auto, colorlinks=true, linkcolor=black, citecolor=black, filecolor=black, urlcolor=black]{hyperref}
\usepackage{tocloft}           
\usepackage{booktabs}          % Tablas profesionales
\usepackage{float} % Paquete necesario para usar la opción [H]
\usepackage{array}
\usepackage{longtable}
\usepackage{circuitikz}
\usepackage{tikz}
\usepackage{tikz-cd}



\usepackage{qrcode}            % Paquete para generar QR
\usepackage{tabularx} % Agregar en el preámbulo
\usepackage[absolute,overlay]{textpos} % Para posicionar objetos libremente

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{ss.titulo}  % Nombre del documento en el encabezado izquierdo
\fancyhead[C]{}  % Centro vacío
\fancyhead[R]{\thepage}  % Numeración de páginas en el encabezado derecho
% \renewcommand{\thepage}{\arabic{page}}  % Asegura la numeración de páginas en números arábigos

% Definición de los anexos
\newcommand{\annex}[1]{\section*{Anexo #1} \addcontentsline{toc}{section}{Anexo #1}}
\addto\captionsspanish{%
  \renewcommand{\tablename}{Tabla}%
  \renewcommand{\listtablename}{Índice de tablas}%
}
\title{{\small Ref.:\uppercase{ss.codigo}}\\{\textbf{ss.titulo}}}

\author{
    \begin{minipage}{0.35\textwidth}
    \centering
    ss.iiParticipantesRolP1\\
    ss.iiParticipantesNombrerazonSocialP1 \\
    ss.iiParticipantesTitulacionP1\\
    ss.iiParticipantesTelefonoP1\\
    ss.iiParticipantesCorreoElectronicoP1\\
    \end{minipage}%
    \hfill
    \begin{minipage}{0.35\textwidth}
    \centering
    ss.iiParticipantesRolP2\\
    ss.iiParticipantesNombrerazonSocialP2 \\
    ss.iiParticipantesTitulacionP2\\
    ss.iiParticipantesTelefonoP2\\
    ss.iiParticipantesCorreoElectronicoP2\\    
    \end{minipage}%
    \hfill
    \begin{minipage}{0.35\textwidth}
    \centering
    ss.iiParticipantesRolP3\\
    ss.iiParticipantesNombrerazonSocialP3 \\
    ss.iiParticipantesTitulacionP3\\
    ss.iiParticipantesTelefonoP3\\
    ss.iiParticipantesCorreoElectronicoP3\\    
    \end{minipage}%
}

\date{\today}
\newcommand{\MostrarVariablesAlFinal}{}


\begin{document}
\begin{Form}

	% Mostrar el título
	\maketitle
	% \onecolumn



	% Iniciar el formato de dos columnas

	% Reemplazamos el entorno IEEEkeywords por una lista de palabras clave

	\tableofcontents  % Tabla de contenidos
	\listoffigures    % Lista de figuras
	\listoftables     % Lista de tablas

    \begin{abstract}
		Este es un ejemplo de resumen para un artículo en formato general. Aquí se debe proporcionar una visión general del contenido del artículo.
	\end{abstract}

	\noindent\textbf{Palabras clave:} Ejemplo, LaTeX, Formato General, Documentación, IEEE.

	\newpage



    











\section{Introducción}
Como resultado de una auditoría energética realizada con un analizador de redes, se ha obtenido un gráfico que representa la energía reactiva total consumida por una instalación trifásica a 400 V, con una potencia contratada de 60 kW. El valor medio del factor de potencia (\(\cos \phi\)) registrado es de 0.85. En este informe se determinará la batería de condensadores más adecuada, el número de transformadores de intensidad necesarios y el valor del parámetro \( C/k \).

\section{Datos Iniciales}
\begin{itemize}
    \item Tensión del sistema (\( V \)): 400 V (trifásico).
    \item Potencia contratada (\( P \)): 60 kW.
    \item Factor de potencia inicial (\( \cos \phi_1 \)): 0.85.
    \item Factor de potencia deseado (\( \cos \phi_2 \)): 0.95 (valor típico para compensación).
    \item Corriente nominal del interruptor automático (IGA): 160 A, regulado a 125 A.
    \item Composición de la batería de condensadores: 7.5 + 7.5 + 7.5 kVAR.
\end{itemize}

\section{Cálculo de la Potencia Reactiva}
La potencia reactiva (\( Q \)) necesaria para compensar el factor de potencia se calcula mediante la siguiente fórmula:

\[
Q = P \cdot (\tan \phi_1 - \tan \phi_2)
\]

Donde:
\begin{itemize}
    \item \( \tan \phi_1 = \tan(\arccos(0.85)) \approx 0.6197 \).
    \item \( \tan \phi_2 = \tan(\arccos(0.95)) \approx 0.3287 \).
\end{itemize}

Sustituyendo los valores:

\[
Q = 60 \, \text{kW} \cdot (0.6197 - 0.3287) \approx 17.46 \, \text{kVAR}
\]

\section{Selección de la Batería de Condensadores}
Consultando el catálogo de baterías de condensadores, se ha determinado que la composición más adecuada para la instalación es:

\[
7.5 + 7.5 + 7.5 \, \text{kVAR} = 22.5 \, \text{kVAR}
\]

Esta batería supera ligeramente la potencia reactiva calculada (17.46 kVAR), lo que asegura una compensación adecuada.

\section{Transformadores de Intensidad}
Para determinar el número de transformadores de intensidad necesarios, se debe considerar la corriente del sistema. La corriente (\( I \)) se calcula como:

\[
I = \frac{P}{\sqrt{3} \cdot V \cdot \cos \phi_2}
\]

Sustituyendo los valores:

\[
I = \frac{60 \, \text{kW}}{\sqrt{3} \cdot 400 \, \text{V} \cdot 0.95} \approx 91.27 \, \text{A}
\]

Dado que el interruptor automático (IGA) está regulado a 125 A, se recomienda utilizar transformadores de intensidad con una relación de 125/5 A. La batería de condensadores requiere un transformador de intensidad por fase, por lo que se necesitan \textbf{3 transformadores de intensidad}.

\section{Cálculo del Parámetro \( C/k \)}
El parámetro \( C/k \) se utiliza para ajustar la compensación de reactiva en función de la capacidad de los condensadores y la constante del sistema. Se calcula como:

\[
C/k = \frac{Q}{V^2 \cdot 2 \pi f}
\]

Donde:
\begin{itemize}
    \item \( Q = 22.5 \, \text{kVAR} \) (capacidad total de la batería).
    \item \( V = 400 \, \text{V} \).
    \item \( f = 50 \, \text{Hz} \).
\end{itemize}

Sustituyendo los valores:

\[
C/k = \frac{22.5 \, \text{kVAR}}{(400 \, \text{V})^2 \cdot 2 \pi \cdot 50 \, \text{Hz}} \approx 0.447 \, \text{µF}
\]

\section{Conclusión}
\begin{itemize}
    \item La batería de condensadores más adecuada para la instalación es de 22.5 kVAR, con una composición de 7.5 + 7.5 + 7.5 kVAR.
    \item Se necesitan \textbf{3 transformadores de intensidad} con una relación de 125/5 A.
    \item El valor del parámetro \( C/k \) calculado es de aproximadamente \textbf{0.447 µF}.
\end{itemize}


\newpage


\section{Introducción}
Este informe tiene como objetivo estimar el valor de reactiva en un sistema eléctrico y determinar los componentes necesarios para la compensación de energía reactiva, incluyendo la batería de condensadores, los transformadores de intensidad y el valor de \( C/k \).

\section{Datos Iniciales}
Se dispone de los siguientes datos para el cálculo:
\begin{itemize}
    \item Potencia activa (\( P \)): 100 kW
    \item Factor de potencia inicial (\( \cos \phi_1 \)): 0.75
    \item Factor de potencia deseado (\( \cos \phi_2 \)): 0.95
    \item Tensión del sistema (\( V \)): 400 V
    \item Frecuencia (\( f \)): 50 Hz
\end{itemize}

\section{Cálculo de la Potencia Reactiva}
La potencia reactiva (\( Q \)) necesaria para compensar el factor de potencia se calcula mediante la siguiente fórmula:

\[
Q = P \cdot (\tan \phi_1 - \tan \phi_2)
\]

Donde:
\begin{itemize}
    \item \( \tan \phi_1 \) es la tangente del ángulo correspondiente al factor de potencia inicial.
    \item \( \tan \phi_2 \) es la tangente del ángulo correspondiente al factor de potencia deseado.
\end{itemize}

Sustituyendo los valores:

\[
\tan \phi_1 = \tan(\arccos(0.75)) \approx 0.8819
\]
\[
\tan \phi_2 = \tan(\arccos(0.95)) \approx 0.3287
\]
\[
Q = 100 \, \text{kW} \cdot (0.8819 - 0.3287) \approx 55.32 \, \text{kVAR}
\]

\section{Dimensionamiento de la Batería de Condensadores}
La batería de condensadores debe ser capaz de suministrar la potencia reactiva calculada. Se selecciona una batería de condensadores de 60 kVAR para cumplir con los requisitos.

\section{Transformadores de Intensidad}
Los transformadores de intensidad deben ser dimensionados en función de la corriente del sistema. La corriente (\( I \)) se calcula como:

\[
I = \frac{P}{\sqrt{3} \cdot V \cdot \cos \phi_2}
\]

Sustituyendo los valores:

\[
I = \frac{100 \, \text{kW}}{\sqrt{3} \cdot 400 \, \text{V} \cdot 0.95} \approx 152.11 \, \text{A}
\]

Se recomienda utilizar transformadores de intensidad con una relación de 150/5 A.

\section{Cálculo de \( C/k \)}
El valor de \( C/k \) se utiliza para ajustar la compensación de reactiva en función de la relación entre la capacidad de los condensadores y la constante del sistema. Se calcula como:

\[
C/k = \frac{Q}{V^2 \cdot 2 \pi f}
\]

Sustituyendo los valores:

\[
C/k = \frac{60 \, \text{kVAR}}{(400 \, \text{V})^2 \cdot 2 \pi \cdot 50 \, \text{Hz}} \approx 1.19 \, \text{µF}
\]

\section{Conclusión}
Se ha estimado una potencia reactiva de 55.32 kVAR, por lo que se recomienda instalar una batería de condensadores de 60 kVAR. Los transformadores de intensidad deben tener una relación de 150/5 A, y el valor de \( C/k \) calculado es de aproximadamente 1.19 µF.
















\section{Bibliografía}
\begin{itemize}
    \item \href{https://library.e.abb.com/public/897462d590876b5fc125791a003bd1e0/1TXA007107G0701_CT8.pdf}{ABB. Corrección del factor de potencia y filtrado de armónicos en las instalaciones eléctricas}
\end{itemize}








    






























	% P A R A M E T R O S 
	%---------------------

	\TextField[name=I0 ,width=0cm]{}
	\TextField[name=Bt,width=0cm]{}


	%==========================================




\end{Form}

\end{document}


"""


class codigo_al_final_del_latex:
    class parametros:
        def parametros():
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            parametros = r"""
            [
            {
                "Descripción": " Potencia contratada",
                "valor": 66,
                "Unidad": "kW"
            },
            {
                "Descripción": "Duración indicativa de la actuación",
                "valor": 15,
                "Unidad": "años"
            },
            {
                "Descripción": "Préstamo recibido",
                "valor": 50000,
                "Unidad": "€"
            },
            {
                "Descripción": "Tasa de interés del préstamo",
                "valor": 5,
                "Unidad": "\\%"
            },
            {
                "Descripción": "jojoj",
                "valor": 5,
                "Unidad": "\\%"
            }
            ]
            """
            datos = json.loads(parametros)
            df = pd.DataFrame(datos)
            df.iloc[:, 0] = "" + df.iloc[:, 0] + r" [" + df.iloc[:, 2] + r"]"
            return df

    class calculos:
        def c01parametros(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            print(99999999999999999999999999999999999999999999999999999999999999999)
            print(ss.nombre)
            print(99999999999999999999999999999999999999999999999999999999999999999)
            for param, value in ss.data[ss.codigo]:
                param = zz.texfile.normalizar_cadena(param)
                print(f"ss.{param} = {value}")

                # Intentar convertir value a float, si no es posible, dejarlo igual
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass  # Si no se puede convertir, mantiene el valor original

                setattr(ss, param, value)

            for param, value in ss.data[ss.codigo][
                1:
            ]:  # Empezamos desde el índice 1 para omitir los encabezados

                # Intentamos convertir a float si es posible
                try:
                    float_value = float(value)
                    setattr(ss, param, float_value)
                    globals()[param] = getattr(ss, param)

                except (ValueError, TypeError):
                    # Si no es convertible a float, lo dejamos como está
                    pass
            return ss

        def c02calculoejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Crear un DataFrame con números aleatorios
            np.random.seed(42)  # Para reproducibilidad
            df = pd.DataFrame(np.random.rand(4, 4))

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["gsdfg"] = total_row

            df = df.copy()

            return ss

    class tablas:
        """"""

        def t03totalfactura(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import pandas as pd

            data_totales = {
                "Concepto": [
                    "Subtotal",
                    "Impuesto eléctrico",
                    "Alquiler contador",
                    "Subtotal",
                    "IVA ",
                    "TOTAL FACTURA",
                ],
                " ": [
                    "           ",
                    "0,5\\%",
                    "30 días",
                    "",
                    "5\\%",
                    "",
                ],
                "€/mes": [19.37, 0.10, 0.81, 20.28, 1.01, 21.29],
            }

            df = pd.DataFrame(data_totales)
            # df.set_index("Energía consumida", inplace=True)  # Convertir la primera columna en índice

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["TOTAL"] = total_row
            # Formatear solo las columnas deseadas a 2 decimales antes de exportar

            df["€/mes"] = df["€/mes"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{2.5cm}}|" + "l" * (len(df.columns) - 1),
            )

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.45\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Total factura}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

        def t02tablaPotenciacontratada(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import pandas as pd

            # Crear el DataFrame
            datos = {
                "Potencia contratada": [
                    "PUNTA",
                    "VALLE",
                    "Margen comercialización fijo",
                ],
                "kW": [
                    ss.PotenciaContratadakw,
                    ss.PotenciaContratadakw,
                    ss.PotenciaContratadakw,
                ],
                "€/kW/año": [26.164043, 1.143132, 3.113000],
                "€/mes": [10.75, 0.47, 1.28],
            }

            df = pd.DataFrame(datos)
            # df.set_index("Potencia contratada", inplace=True)  # Convertir la primera columna en índice

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["TOTAL"] = total_row
            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]
            df.loc["gsdfg"] = total_row

            # Formatear solo las columnas deseadas a 2 decimales antes de exportar
            df["kW"] = df["kW"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )
            df["€/mes"] = df["€/mes"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{2.5cm}}|" + "l" * (len(df.columns) - 1),
            )

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.45\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Potencia}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

        def t02tablaEnergíaconsumida(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import pandas as pd

            # Crear el DataFrame con los datos
            datos = {
                "Energía consumida": [
                    "PUNTA",
                    "LLANO",
                    "VALLE",
                    "Coste energía",
                    "Compensación excedentes FV",
                ],
                "kWh": [13.5, 10.5, 108.0, 132.0, "-"],
                "€/kWh": [0.074409, 0.028470, 0.003034, 0.150000, "-"],
                "€/mes": [1.00, 0.30, 0.33, 19.80, -14.56],
            }

            df = pd.DataFrame(datos)

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]
            df.loc["gsdfg"] = total_row

            # Formatear solo las columnas deseadas a 2 decimales antes de exportar
            df["kWh"] = df["kWh"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )
            df["€/mes"] = df["€/mes"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{2.5cm}}|" + "l" * (len(df.columns) - 1),
            )

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.45\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Energía consumida}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

        def t00tablaparametros(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            """poner en el valor el simbolo sin acentos guiones ni nada, que sera el valor para los calculos"""

            # df = solo_desarrollo.parametros.parametros()

            par = ss.data[next(iter(ss.data))]
            df = pd.DataFrame(par[1:], columns=par[0])

            # print(df)

            # Exportar a LaTeX
            latex_code = df.to_latex(index=False, escape=False)

            df.iloc[:, 1] = df.iloc[:, 1].apply(
                lambda x: f"{float(x):.2f}".replace(".", ",") if type(x) == float else x
            )  # Dos decimales

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{5cm}}|" + "l" * (len(df.columns) - 1),
                position="H",
                caption="t00tablaparametros",
            )

            return latex_code

        def t99tablaresultados(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            titulo = "hjklhlkjhkljjhhj"
            """poner en el valor el simbolo sin acentos guiones ni nada, que sera el valor para los calculos"""
            parametros = r"""
            [
            {
                "Descripción": "Período de recuperación",
                "valor": 11,
                "Unidad": "años"
            },
            {
                "Descripción": "Valor Actual Neto",
                "valor": 59063.66,
                "Unidad": "€"
            },
            {
                "Descripción": "Tasa Interna de Retorno",
                "valor": 13.09,
                "Unidad": "\\%"
            }
            ]
            """
            datos = json.loads(parametros)
            df = pd.DataFrame(datos)
            df.iloc[:, 0] = "" + df.iloc[:, 0] + r" [" + df.iloc[:, 2] + r"]"

            # Exportar a LaTeX
            latex_code = df.to_latex(index=False, escape=False)

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{3cm}}|" + "l" * (len(df.columns) - 1),
                position="H",
                caption="t99tablaresultados",
            )

            return latex_code

        def xxt01tablaejemplo(ss):
            print(f"\n {100*'+'}{__class__.__name__}.{sys._getframe().f_code.co_name}")

            df = ss.df.copy()

            # Generar la tabla en formato LaTeX
            latex_code = df.to_latex(index=False)

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.4\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Tabla escalada al ancho de la página}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

    class figuras:
        """"""

        def f01figuraejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # 1. Crear un archivo temporal para la imagen
            titulo = "Flujo de Caja y Acumulado"
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image:
                temp_image_path2 = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal

            df = pd.DataFrame(np.random.rand(4, 4))
            # .......
            #
            fig, ax = plt.subplots(figsize=(6, 5))
            df.plot(ax=ax)
            #
            # .......

            ax.grid(axis="y", linestyle="--", alpha=0.7)
            plt.xticks(rotation=50)
            ax.set_ylabel("€")  # Etiqueta en el eje y

            plt.savefig(temp_image_path2, format="png", dpi=300)

            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + """\\begin{figure}[H]
                """
                + f"""
                \\includegraphics[width=.5\\textwidth]{{{temp_image_path2}}}
                \\caption{{{titulo}}}
                """
                + """
                \\label{fig:dfassssdfsa}
                \\end{figure}
                """
            )

            return latex_code

        def f02figuraejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # 1. Crear un archivo temporal para la imagen
            titulo = "Flujo de Caja y Acumulado"
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image:
                temp_image_path1 = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal

            df = pd.DataFrame(np.random.rand(4, 4))
            # .......
            #
            fig, ax = plt.subplots(figsize=(6, 5))
            df.plot(ax=ax)
            #
            # .......

            ax.grid(axis="y", linestyle="--", alpha=0.7)
            plt.xticks(rotation=50)
            ax.set_ylabel("€")  # Etiqueta en el eje y

            plt.savefig(temp_image_path1, format="png", dpi=300)

            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + """\\begin{figure}[H]
                """
                + f"""
                \\includegraphics[width=.5\\textwidth]{{{temp_image_path1}}}
                \\caption{{{titulo}}}
                """
                + """
                \\label{fig:dfasdfsa}
                \\end{figure}
                """
            )

            return latex_code


class zz:
    class main:

        def parametros_latex(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Para cada hoja
            for key, values in ss.data.items():
                # Crear un DataFrame a partir de los datos
                df = pd.DataFrame(
                    values[1:], columns=values[0]
                )  # Crear DataFrame con la primera fila como encabezado
                df = df.set_index(df.columns[0])
                variable = f"hh {key}"
                variable = zz.texfile.normalizar_cadena(variable)
                setattr(ss, variable, df)  # Asignar como atributo de ss

                # Crear un DataFrame a partir de las columnas del DataFrame
                for col in df.columns:
                    variable = f"cc {key} {col}"
                    variable = zz.texfile.normalizar_cadena(variable)
                    setattr(ss, variable, df[[col]])

                # crear las variables con claves formadas por <indice><columna>
                for index, row in df.iterrows():
                    for col in df.columns:
                        variable = f"ii {key} {index} {col}"
                        variable = zz.texfile.normalizar_cadena(variable)
                        setattr(ss, variable, row[col])
                print(ss.__dict__)
            return ss

        def parametros_formulario(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # excel a ss sapacio de estado
            for param, value in ss.data[ss.nombre]:
                param = zz.texfile.normalizar_cadena(param)
                # Intentar convertir value a float, si no es posible, dejarlo igual
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass  # Si no se puede convertir, mantiene el valor original
                setattr(ss, param, value)
            return ss

        def main(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # segun si es un formulario o un documento nuevo
            pdf_path = Path(ss.nombre).with_suffix(".pdf")

            # if os.path.exists(f"{os.path.splitext(pdf_path)[0]}.tex"):
            if ss.archivo_nuevo:
                # print("El archivo .tex existe.")
                tex_path = f"{os.path.splitext(pdf_path)[0]}.tex"
                ss = zz.main.parametros_latex(ss)

                """calculos"""
                listacalculos = [
                    m
                    for m in dir(codigo_al_final_del_latex.calculos)
                    if callable(getattr(codigo_al_final_del_latex.calculos, m))
                    and not m.startswith("__")
                ]
                listacalculos = [s for s in listacalculos if not s.startswith("xx")]
                for calculo in sorted(listacalculos):
                    metodo = getattr(codigo_al_final_del_latex.calculos, calculo, None)
                    ss = metodo(ss)

                output_buffer = zz.texfile.replaceLatex(tex_path, ss)

            else:
                print("El archivo .tex no existe.")
                # Leer el archivo PDF desde el disco
                ss = zz.main.parametros_formulario(ss)

                """calculos"""
                listacalculos = [
                    m
                    for m in dir(codigo_al_final_del_latex.calculos)
                    if callable(getattr(codigo_al_final_del_latex.calculos, m))
                    and not m.startswith("__")
                ]
                listacalculos = [s for s in listacalculos if not s.startswith("xx")]
                for calculo in sorted(listacalculos):
                    metodo = getattr(codigo_al_final_del_latex.calculos, calculo, None)
                    ss = metodo(ss)

                with open(pdf_path, "rb") as file:
                    pdf_content = file.read()
                output_buffer = BytesIO(pdf_content)
                output_buffer.seek(0)
                output_buffer = zz.formulario.rellenoFormulario(output_buffer, ss)

            return output_buffer

            #######################

        def xlsxhojaconvalorespordefecto(df):
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # si es un nuevo archivo tex copia los parametros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            codigo, titulo = separar_y_mayusculizar(nombre)
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria

            #
            ruta_archivo_excel = f"./{codigo}.xlsx"
            if 1:
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja por defecto (llamada 'Sheet')
                    hoja_por_defecto = libro["Sheet"]
                    libro.remove(hoja_por_defecto)

                    hoja = libro.create_sheet(title=codigo)

                    # Seleccionar la hoja activa
                    # hoja = libro.active
                    # Guardar el nuevo libro de trabajo
                    libro.save(ruta_archivo_excel)

                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    hoja = libro[codigo]

                    # Borrar todo el contenido de la hoja
                    for fila in hoja.iter_rows():
                        for celda in fila:
                            celda.value = None

                    # Insertar los datos del DataFrame en la hoja de Excel
                    # Insertar la cabecera
                    hoja["A1"] = "Parámetro"
                    hoja["B1"] = "Valor"

                    # Insertar los datos del DataFrame en la hoja de Excel
                    for i, (parametro, valor) in enumerate(
                        zip(df["Descripción"], df["valor"]), start=2
                    ):
                        # hoja[f"A{i}"] = parametro.replace("_", "").replace(" ", "")
                        hoja[f"A{i}"] = parametro
                        hoja[f"B{i}"] = valor

                    # Guardar el archivo con los cambios
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    # print(ruta_archivo_excel)
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # Cargar el archivo Excel

            # Cargar todas las hojas del archivo Excel
            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )

            # Crear el diccionario data
            data = {
                sheet_name: [df.columns.tolist()]
                + df.values.tolist()  # Incluir cabecera y datos
                for sheet_name, df in sheets.items()
            }

            return data

    class formulario:
        def rellenoFormulario(pdf_buffer, ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            pdf_document = fitz.open(stream=pdf_buffer, filetype="pdf")
            output_buffer = BytesIO()
            variables = vars(ss)
            # print("CAMPOS DEL FORMULARIO")
            # Recorrer todas las páginas del PDF
            for page_num in range(len(pdf_document)):
                page = pdf_document.load_page(page_num)
                # Iterar sobre los widgets (campos de formulario)
                for widget in page.widgets():
                    field_name = widget.field_name  # Obtener el ss.nombre del campo
                    field_name = zz.texfile.normalizar_cadena(field_name)
                    # print(field_name)
                    if (
                        field_name in variables
                    ):  # Si el ss.nombre del campo está en el SimpleNamespace
                        widget.field_value = str(
                            variables[field_name]
                        )  # Asignar el valor
                        try:
                            widget.update()  # Actualizar el widget
                        except Exception as e:
                            """"""
                            print(f"Error al actualizar el campo {field_name}: {e}")
            pdf_document.save(output_buffer)
            pdf_document.close()
            output_buffer.seek(0)
            return output_buffer

    class texfile:

        def replaceLatex(ruta_tex, ss):
            """
            Compila un archivo .tex a PDF, reemplaza un texto y devuelve el PDF en memoria.
            :param ruta_tex: Ruta al archivo .tex original.
            :return: Objeto BytesIO con el contenido del PDF.
            """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            ruta_tex_absoluta = os.path.abspath(
                ruta_tex
            )  # Obtener la ruta absoluta del archivo original
            nombre_archivo = os.path.basename(ruta_tex)  # Nombre del archivo sin ruta

            # Crear un directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                ruta_tex_temp = os.path.join(temp_dir, nombre_archivo)

                contenido = file_tex()

                ##########################################

                if 1:
                    """figuras"""
                    listafiguras = [
                        m
                        for m in dir(codigo_al_final_del_latex.figuras)
                        if callable(getattr(codigo_al_final_del_latex.figuras, m))
                        and not m.startswith("__")
                    ]
                    listafiguras = [s for s in listafiguras if not s.startswith("xx")]
                    for figura in listafiguras:
                        metodo = getattr(
                            codigo_al_final_del_latex.figuras, figura, None
                        )
                        contenido = contenido.replace(f"ss.{figura}", metodo(ss))

                if 1:
                    """tablas"""
                    listatablas = [
                        m
                        for m in dir(codigo_al_final_del_latex.tablas)
                        if callable(getattr(codigo_al_final_del_latex.tablas, m))
                        and not m.startswith("__")
                    ]
                    listatablas = [s for s in listatablas if not s.startswith("xx")]
                    for tabla in listatablas:
                        # print(tabla)
                        metodo = getattr(
                            codigo_al_final_del_latex.tablas, tabla, None
                        )  # Obtener el método o None si no existe
                        contenido = contenido.replace(f"ss.{tabla}", metodo(ss))

                if 1:
                    """variables"""
                    # Reemplazo usando .replace() para cada variable
                    for variable, valor in vars(ss).items():
                        if (
                            valor is None
                            or (isinstance(valor, str) and valor.strip() == "")
                            or (isinstance(valor, pd.DataFrame) and valor.empty)
                            or (isinstance(valor, pd.Series) and valor.isna().all())
                        ):
                            valor = "%"

                        contenido = contenido.replace(f"ss.{variable}", str(valor))
                        print(f"ss.{variable}")

                # solo depuracion
                # solo depuracion
                # solo depuracion
                # solo depuracion
                with open(
                    "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.tex", "w", encoding="utf-8"
                ) as file:
                    file.write(contenido)
                # solo depuracion
                # solo depuracion
                # solo depuracion
                # solo depuracion

                ##########################################

                # Guardar los cambios en el archivo temporal
                with open(ruta_tex_temp, "w", encoding="utf-8") as file:
                    file.write(contenido)

                # Función para ejecutar pdflatex dos veces
                def ejecutar_pdflatex():
                    print(
                        f"\n{100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                    )
                    if 1:
                        subprocess.run(
                            ["pdflatex", "-output-directory", temp_dir, ruta_tex_temp],
                            check=True,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )

                    return True

                # # para que se cre el indice debe estar descomentado
                # Ejecutar pdflatex dos veces para generar la TOC correctamente
                if not (ejecutar_pdflatex() and ejecutar_pdflatex()):
                    return None  # Si la compilación falla, se detiene el proceso
                # ejecutar_pdflatex()

                # Ruta del PDF generado
                pdf_generado = os.path.join(
                    temp_dir, nombre_archivo.replace(".tex", ".pdf")
                )

                # Leer el PDF en memoria
                with open(pdf_generado, "rb") as file:
                    pdf_en_memoria = BytesIO(file.read())

            return pdf_en_memoria

        def normalizar_cadena(texto):
            if isinstance(texto, str):  # Asegurar que es un string
                texto = texto.lower()
                texto = "".join(
                    c
                    for c in unicodedata.normalize("NFD", texto)
                    if unicodedata.category(c) != "Mn"
                )
                # Capitalizar la primera letra después de un espacio
                texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
                # Eliminar espacios
                texto = re.sub(r"\s+", "", texto)
                # Mantener solo caracteres permitidos
                texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
            return texto


from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from PyPDF2 import PdfReader, PdfWriter
from io import BytesIO
import os
import matplotlib.pyplot as plt
import matplotlib

matplotlib.use("Agg")

import io
import base64
import pandas as pd
import fitz  # PyMuPDF
from types import SimpleNamespace
import subprocess
import shutil
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata

import numpy as np
import sys


router = APIRouter()


def separar_y_mayusculizar(texto):
    # Usamos una expresión regular para separar las mayúsculas
    partes = re.findall(r"[A-Z][^A-Z]*", texto)

    # Separar la primera parte (mantenerla como está)
    primera_parte = partes[0]

    # El resto de las partes lo convertimos a mayúsculas y las unimos con espacios
    resto = " ".join([parte.upper() for parte in partes[1:]])

    return primera_parte, resto


router = APIRouter()


@router.post("/Btx030")
def Btx030BateríaDeCondensadores(
    data: dict,
):

    # parra que en la primera carga, los campos que tgiene valores separados por ; para los selectores coja el primer valor por defecto, tambien 
    # debera hacerlo si se da a recalcular sin seleccionar un valor 
    for key, values in data.items():
        for i in range(1, len(values)):  # Saltamos la primera fila (encabezados)
            if isinstance(values[i][1], str):  # Verificar si es una cadena antes de dividir
                values[i][1] = values[i][1].split(';')[0]  # Tomar solo el primer valor
    # parra que en la primera carga, los campos que tgiene valores separados por ; para los selectores coja el primer valor por defecto, tambien 
    # debera hacerlo si se da a recalcular sin seleccionar un valor 
    for key, values in data.items():
        for i in range(1, len(values)):  # Saltamos la primera fila (encabezados)
            if isinstance(values[i][1], str):  # Verificar si es una cadena antes de dividir
                values[i][1] = values[i][1].split(';')[0]  # Tomar solo el primer valor
    print(data)

    ss = SimpleNamespace()

    # Ejemplo de uso
    ss.nombre = sys._getframe().f_code.co_name

    ss.codigo, ss.titulo = separar_y_mayusculizar(ss.nombre)

    ss.archivo_nuevo = True

    ss.data = data

    output_buffer = zz.main.main(ss)
    # solo depuracion
    # solo depuracion
    # solo depuracion
    with open(f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf", "wb") as f:
        f.write(output_buffer.getvalue())
    # solo depuracion
    # solo depuracion
    # solo depuracion
    # devuelve a fastapi el archivo
    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


if __name__ == "__main__":

    # Cambiar al directorio del script
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    df = codigo_al_final_del_latex.parametros.parametros()
    data = zz.main.xlsxhojaconvalorespordefecto(df)
    nombre_funcion = os.path.basename(__file__)[:-3]
    globals()[nombre_funcion](data)
