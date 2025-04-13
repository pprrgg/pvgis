#!/usr/bin/env python3


def file_tex():
    return r"""


\documentclass[a4paper,10pt]{article}
%\documentclass[a4paper,10pt,twocolumn]{article}
\usepackage[top = 2cm, bottom = 2cm, left = 2cm, right = 2cm, asymmetric]{geometry} % Especificar los márgenes según la norma
\usepackage[spanish]{babel}    

\usepackage{graphicx}  % Para incluir imágenes si es necesario
\usepackage{amsmath, amssymb}  % Para expresiones matemáticas
\usepackage{fancyhdr}  % Para personalizar encabezados
\usepackage{chngcntr}  % Para cambiar la numeración de apartados y subsecciones
\usepackage{tocbibind}  % Para incluir bibliografía en la tabla de contenidos
\usepackage{appendix}  % Para el formato de anexos
\usepackage{lipsum}  % Para generar texto de relleno (dummy text)
\usepackage{geometry}  % Para personalizar los márgenes
\usepackage{multicol}  % Para columnas
\usepackage{titlesec}  % Para personalizar los títulos de las secciones
\usepackage[utf8]{inputenc}    
\usepackage[T1]{fontenc}       
\usepackage[pdftex,pdfencoding=auto, colorlinks=true, linkcolor=black, citecolor=black, filecolor=black, urlcolor=black]{hyperref}
\usepackage{tocloft}           
\usepackage{booktabs}          % Tablas profesionales
\usepackage{float} % Paquete necesario para usar la opción [H]
\usepackage{array}
\usepackage{longtable}
\usepackage{circuitikz}
\usepackage{tikz}
\usepackage{tikz-cd}



\usepackage{qrcode}            % Paquete para generar QR
\usepackage{tabularx} % Agregar en el preámbulo
\usepackage[absolute,overlay]{textpos} % Para posicionar objetos libremente

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{ss.titulo}  % Nombre del documento en el encabezado izquierdo
\fancyhead[C]{}  % Centro vacío
\fancyhead[R]{\thepage}  % Numeración de páginas en el encabezado derecho
% \renewcommand{\thepage}{\arabic{page}}  % Asegura la numeración de páginas en números arábigos

% Definición de los anexos
\newcommand{\annex}[1]{\section{Anexo #1} \addcontentsline{toc}{section}{Anexo #1}}
\addto\captionsspanish{%
  \renewcommand{\tablename}{Tabla}%
  \renewcommand{\listtablename}{Índice de tablas}%
}
%\title{{\small Ref.:\uppercase{ss.codigo}}\\{\textbf{ss.titulo}}}
\title{{\small Ref.:\uppercase{ss.codigo}}\\{\textbf{Instalación de Comunicación de SAI Legrand}}}




\author{Francisco Manuel Roman Gamez}
\date{\today}
\newcommand{\MostrarVariablesAlFinal}{}

\begin{document}

\begin{Form}

	% Mostrar el título



\maketitle

	\begin{abstract}

En este documento se ha realizado un análisis completo de un sistema SAI Legrand para determinar la referencia y modelo de los 
elementos necesarios para realizar la instalación de
comunicación.
   	\end{abstract}

\section*{Requisitos de la Instalación}
\begin{itemize}
    \item Temperatura mínima de la sala: 12°C
    \item Equipamiento de detección: Sensor de humo y sensor de temperatura
    \item Integración: Tarjeta de red interna en el SAI
\end{itemize}

\section*{Componentes Seleccionados}
\begin{table}[H]
    \centering
    \begin{tabular}{lllll}
        \toprule
        \textbf{Cant.} & \textbf{Elemento} & \textbf{Ref.} & \textbf{Modelo} & \textbf{Justificación} \\
        \midrule
        1 & Interfaz de red & 3 109 31 & CS141B SK & 
        \begin{minipage}{5cm}
            \footnotesize
            Tarjeta interna que cumple con el requisito de integración en el SAI. Soporta protocolos SNMP para monitorización remota y conexión Ethernet 10/100Mbit (pág. 111 del catálogo).
        \end{minipage} \\
        \midrule
        1 & Administrador sensores & 3 108 99 & SensorManager & 
        \begin{minipage}{5cm}
            \footnotesize
            Necesario para gestionar múltiples sensores (hasta 8 entradas). Permite configuración de escalas para adaptar sensores genéricos (pág. 112, características técnicas).
        \end{minipage} \\
        \midrule
        1 & Sensor temperatura & 3 109 00 & SM\_T & 
        \begin{minipage}{5cm}
            \footnotesize
            Rango de -25°C a 100°C cubre los 12°C requeridos. Compatible con SensorManager e incluye cable de 5m (pág. 112, tabla de especificaciones).
        \end{minipage} \\
        \midrule
        1 & Sensor humo & 3 109 02 & Sensor puerta & 
        \begin{minipage}{5cm}
            \footnotesize
            Solución adaptable mediante contacto digital. Compatible con CS141B SK y SensorManager (pág. 112, nota de compatibilidad).
        \end{minipage} \\
        \midrule
        1 & Software gestión & 3 108 79 & UPS Management & 
        \begin{minipage}{5cm}
            \footnotesize
            Software completo para monitorización remota. Requiere agente RCCMD para control de equipos (pág. 113, sección de protección centralizada).
        \end{minipage} \\
        \midrule
        1 & Agente software & 3 108 85 & RCCMD & 
        \begin{minipage}{5cm}
            \footnotesize
            Licencia necesaria para ejecutar comandos remotos en los ordenadores conectados al SAI (pág. 113, descripción de funcionalidades).
        \end{minipage} \\
        \bottomrule
    \end{tabular}
    \caption{Especificación técnica}
    \label{tab:componentes}
\end{table}

\section*{Notas Adicionales}
\begin{itemize}
    \item Todos los componentes seleccionados son compatibles entre sí según el catálogo Legrand 2023.
    \item El sensor de humo requiere configuración especial como entrada digital en el SensorManager.
    \item La interfaz CS141B SK incluye 1 licencia RCCMD (pág. 111), pero se recomienda adicional para redundancia.
\end{itemize}















































\section{Bibliografía}
\begin{itemize}
    \item \href{https://www.ingenierosformacion.com/}{0LCL0JYCGW5C, Curso 'Avanzado en instalaciones eléctricas industriales. Industria 4.0. COGITI, José Luis Rodríguez' }
    \item \href{https://www.legrand.es/sites/g/files/ocwmcr651/files/2022-10/Catalogo-SAI-UPS-Legrand.pdf}{LEGRAND .SAI SISTEMA DE ALIMENTACIÓN ININTERRUMPIDA.}
    \item \href{https://www.generex.de}{GENEREX. UPS Management}
\end{itemize}








    






























	% P A R A M E T R O S 
	%---------------------

	\TextField[name=I0 ,width=0cm]{}
	\TextField[name=Bt,width=0cm]{}


	%==========================================




\end{Form}

\end{document}


"""


class codigo_al_final_del_latex:
    class parametros:
        def parametros():
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            parametros = r"""
            [
            {
                "Descripción": " Potencia contratada",
                "valor": 66,
                "Unidad": "kW"
            },
            {
                "Descripción": "Duración indicativa de la actuación",
                "valor": 15,
                "Unidad": "años"
            },
            {
                "Descripción": "Préstamo recibido",
                "valor": 50000,
                "Unidad": "€"
            },
            {
                "Descripción": "Tasa de interés del préstamo",
                "valor": 5,
                "Unidad": "\\%"
            },
            {
                "Descripción": "jojoj",
                "valor": 5,
                "Unidad": "\\%"
            }
            ]
            """
            datos = json.loads(parametros)
            df = pd.DataFrame(datos)
            df.iloc[:, 0] = "" + df.iloc[:, 0] + r" [" + df.iloc[:, 2] + r"]"
            return df

    class calculos:
        def c01parametros(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            print(99999999999999999999999999999999999999999999999999999999999999999)
            print(ss.nombre)
            print(99999999999999999999999999999999999999999999999999999999999999999)
            for param, value in ss.data[ss.codigo]:
                param = zz.texfile.normalizar_cadena(param)
                print(f"ss.{param} = {value}")

                # Intentar convertir value a float, si no es posible, dejarlo igual
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass  # Si no se puede convertir, mantiene el valor original

                setattr(ss, param, value)

            for param, value in ss.data[ss.codigo][
                1:
            ]:  # Empezamos desde el índice 1 para omitir los encabezados

                # Intentamos convertir a float si es posible
                try:
                    float_value = float(value)
                    setattr(ss, param, float_value)
                    globals()[param] = getattr(ss, param)

                except (ValueError, TypeError):
                    # Si no es convertible a float, lo dejamos como está
                    pass
            return ss

        def c02calculoejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Crear un DataFrame con números aleatorios
            np.random.seed(42)  # Para reproducibilidad
            df = pd.DataFrame(np.random.rand(4, 4))

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["gsdfg"] = total_row

            df = df.copy()

            return ss

    class tablas:
        """"""

        def t03totalfactura(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import pandas as pd

            data_totales = {
                "Concepto": [
                    "Subtotal",
                    "Impuesto eléctrico",
                    "Alquiler contador",
                    "Subtotal",
                    "IVA ",
                    "TOTAL FACTURA",
                ],
                " ": [
                    "           ",
                    "0,5\\%",
                    "30 días",
                    "",
                    "5\\%",
                    "",
                ],
                "€/mes": [19.37, 0.10, 0.81, 20.28, 1.01, 21.29],
            }

            df = pd.DataFrame(data_totales)
            # df.set_index("Energía consumida", inplace=True)  # Convertir la primera columna en índice

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["TOTAL"] = total_row
            # Formatear solo las columnas deseadas a 2 decimales antes de exportar

            df["€/mes"] = df["€/mes"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{2.5cm}}|" + "l" * (len(df.columns) - 1),
            )

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.45\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Total factura}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

        def t02tablaPotenciacontratada(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import pandas as pd

            # Crear el DataFrame
            datos = {
                "Potencia contratada": [
                    "PUNTA",
                    "VALLE",
                    "Margen comercialización fijo",
                ],
                "kW": [
                    ss.PotenciaContratadakw,
                    ss.PotenciaContratadakw,
                    ss.PotenciaContratadakw,
                ],
                "€/kW/año": [26.164043, 1.143132, 3.113000],
                "€/mes": [10.75, 0.47, 1.28],
            }

            df = pd.DataFrame(datos)
            # df.set_index("Potencia contratada", inplace=True)  # Convertir la primera columna en índice

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["TOTAL"] = total_row
            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]
            df.loc["gsdfg"] = total_row

            # Formatear solo las columnas deseadas a 2 decimales antes de exportar
            df["kW"] = df["kW"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )
            df["€/mes"] = df["€/mes"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{2.5cm}}|" + "l" * (len(df.columns) - 1),
            )

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.45\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Potencia}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

        def t02tablaEnergíaconsumida(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import pandas as pd

            # Crear el DataFrame con los datos
            datos = {
                "Energía consumida": [
                    "PUNTA",
                    "LLANO",
                    "VALLE",
                    "Coste energía",
                    "Compensación excedentes FV",
                ],
                "kWh": [13.5, 10.5, 108.0, 132.0, "-"],
                "€/kWh": [0.074409, 0.028470, 0.003034, 0.150000, "-"],
                "€/mes": [1.00, 0.30, 0.33, 19.80, -14.56],
            }

            df = pd.DataFrame(datos)

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]
            df.loc["gsdfg"] = total_row

            # Formatear solo las columnas deseadas a 2 decimales antes de exportar
            df["kWh"] = df["kWh"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )
            df["€/mes"] = df["€/mes"].apply(
                lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else x
            )

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{2.5cm}}|" + "l" * (len(df.columns) - 1),
            )

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.45\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Energía consumida}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

        def t00tablaparametros(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            """poner en el valor el simbolo sin acentos guiones ni nada, que sera el valor para los calculos"""

            # df = solo_desarrollo.parametros.parametros()

            par = ss.data[next(iter(ss.data))]
            df = pd.DataFrame(par[1:], columns=par[0])

            # print(df)

            # Exportar a LaTeX
            latex_code = df.to_latex(index=False, escape=False)

            df.iloc[:, 1] = df.iloc[:, 1].apply(
                lambda x: f"{float(x):.2f}".replace(".", ",") if type(x) == float else x
            )  # Dos decimales

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{5cm}}|" + "l" * (len(df.columns) - 1),
                position="H",
                caption="t00tablaparametros",
            )

            return latex_code

        def t99tablaresultados(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            titulo = "hjklhlkjhkljjhhj"
            """poner en el valor el simbolo sin acentos guiones ni nada, que sera el valor para los calculos"""
            parametros = r"""
            [
            {
                "Descripción": "Período de recuperación",
                "valor": 11,
                "Unidad": "años"
            },
            {
                "Descripción": "Valor Actual Neto",
                "valor": 59063.66,
                "Unidad": "€"
            },
            {
                "Descripción": "Tasa Interna de Retorno",
                "valor": 13.09,
                "Unidad": "\\%"
            }
            ]
            """
            datos = json.loads(parametros)
            df = pd.DataFrame(datos)
            df.iloc[:, 0] = "" + df.iloc[:, 0] + r" [" + df.iloc[:, 2] + r"]"

            # Exportar a LaTeX
            latex_code = df.to_latex(index=False, escape=False)

            # Exportar a LaTeX
            latex_code = df.to_latex(
                index=False,
                escape=False,
                decimal=",",
                column_format=f"p{{3cm}}|" + "l" * (len(df.columns) - 1),
                position="H",
                caption="t99tablaresultados",
            )

            return latex_code

        def xxt01tablaejemplo(ss):
            print(f"\n {100*'+'}{__class__.__name__}.{sys._getframe().f_code.co_name}")

            df = ss.df.copy()

            # Generar la tabla en formato LaTeX
            latex_code = df.to_latex(index=False)

            # Insertar el entorno resizebox
            latex_code_with_resizebox = f"""
            \\begin{{table}}[h!]
                \\resizebox{{0.4\\textwidth}}{{!}}{{%
                {latex_code}
                }}
                \\caption{{Tabla escalada al ancho de la página}}
            \\end{{table}}
            """

            return latex_code_with_resizebox

    class figuras:
        """"""

        def f01figuraejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # 1. Crear un archivo temporal para la imagen

            titulo = "Señal de corriente con armónicos"
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image:
                temp_image_path2 = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal





            import numpy as np
            import matplotlib.pyplot as plt

            # Parámetros de la señal
            f1 = 50  # Frecuencia fundamental (Hz)
            I1 = 160  # Amplitud de la corriente fundamental (A)
            I3 = 32  # Amplitud del 3er armónico (A)
            I5 = 0  # Amplitud del 5to armónico (A)
            I7 = 0  # Amplitud del 7mo armónico (A)
            I9 = 41.6  # Amplitud del 9no armónico (A)

            # Tiempo de muestreo
            t = np.linspace(0, 0.02, 1000)  # Un ciclo de 20 ms (50 Hz)

            # Señales individuales
            fundamental = I1 * np.sin(2 * np.pi * f1 * t)
            armonico_3 = I3 * np.sin(2 * np.pi * 3 * f1 * t)
            armonico_5 = I5 * np.sin(2 * np.pi * 5 * f1 * t)
            armonico_7 = I7 * np.sin(2 * np.pi * 7 * f1 * t)
            armonico_9 = I9 * np.sin(2 * np.pi * 9 * f1 * t)

            # Señal total (suma de la fundamental y los armónicos)
            corriente_total = fundamental + armonico_3 + armonico_5 + armonico_7 + armonico_9

            # Gráfica
            plt.figure(figsize=(10, 6))
            plt.plot(t, corriente_total, label="Corriente total", color="blue")
            plt.plot(t, fundamental, label="Fundamental (50 Hz)", linestyle="--", color="red")
            plt.plot(t, armonico_3, label="3er armónico (150 Hz)", linestyle=":", color="green")
            plt.plot(t, armonico_9, label="9no armónico (450 Hz)", linestyle="-.", color="purple")

            # Configuración de la gráfica
            # plt.title("Señal de corriente con armónicos")
            plt.xlabel("Tiempo [s]")
            plt.ylabel("Corriente [A]")
            plt.grid(True)
            plt.legend(loc="upper right")
            plt.show()






            plt.savefig(temp_image_path2, format="png", dpi=300)

            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + """\\begin{figure}[H]
                """
                + f"""
                \\includegraphics[width=.985\\textwidth]{{{temp_image_path2}}}
                \\caption{{{titulo}}}
                """
                + """
                \\label{fig:dfassssdfsa}
                \\end{figure}
                """
            )

            return latex_code

        def f02figuraejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # 1. Crear un archivo temporal para la imagen
            titulo = "Flujo de Caja y Acumulado"
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image:
                temp_image_path1 = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal

            df = pd.DataFrame(np.random.rand(4, 4))
            # .......
            #
            fig, ax = plt.subplots(figsize=(6, 5))
            df.plot(ax=ax)
            #
            # .......

            ax.grid(axis="y", linestyle="--", alpha=0.7)
            plt.xticks(rotation=50)
            ax.set_ylabel("€")  # Etiqueta en el eje y

            plt.savefig(temp_image_path1, format="png", dpi=300)

            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + """\\begin{figure}[H]
                """
                + f"""
                \\includegraphics[width=.5\\textwidth]{{{temp_image_path1}}}
                \\caption{{{titulo}}}
                """
                + """
                \\label{fig:dfasdfsa}
                \\end{figure}
                """
            )

            return latex_code


class zz:
    class main:

        def parametros_latex(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Para cada hoja
            for key, values in ss.data.items():
                # Crear un DataFrame a partir de los datos
                df = pd.DataFrame(
                    values[1:], columns=values[0]
                )  # Crear DataFrame con la primera fila como encabezado
                df = df.set_index(df.columns[0])
                variable = f"hh {key}"
                variable = zz.texfile.normalizar_cadena(variable)
                setattr(ss, variable, df)  # Asignar como atributo de ss

                # Crear un DataFrame a partir de las columnas del DataFrame
                for col in df.columns:
                    variable = f"cc {key} {col}"
                    variable = zz.texfile.normalizar_cadena(variable)
                    setattr(ss, variable, df[[col]])

                # crear las variables con claves formadas por <indice><columna>
                for index, row in df.iterrows():
                    for col in df.columns:
                        variable = f"ii {key} {index} {col}"
                        variable = zz.texfile.normalizar_cadena(variable)
                        setattr(ss, variable, row[col])
                print(ss.__dict__)
            return ss

        def parametros_formulario(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # excel a ss sapacio de estado
            for param, value in ss.data[ss.nombre]:
                param = zz.texfile.normalizar_cadena(param)
                # Intentar convertir value a float, si no es posible, dejarlo igual
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass  # Si no se puede convertir, mantiene el valor original
                setattr(ss, param, value)
            return ss

        def main(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # segun si es un formulario o un documento nuevo
            pdf_path = Path(ss.nombre).with_suffix(".pdf")

            # if os.path.exists(f"{os.path.splitext(pdf_path)[0]}.tex"):
            if ss.archivo_nuevo:
                # print("El archivo .tex existe.")
                tex_path = f"{os.path.splitext(pdf_path)[0]}.tex"
                ss = zz.main.parametros_latex(ss)

                """calculos"""
                listacalculos = [
                    m
                    for m in dir(codigo_al_final_del_latex.calculos)
                    if callable(getattr(codigo_al_final_del_latex.calculos, m))
                    and not m.startswith("__")
                ]
                listacalculos = [s for s in listacalculos if not s.startswith("xx")]
                for calculo in sorted(listacalculos):
                    metodo = getattr(codigo_al_final_del_latex.calculos, calculo, None)
                    ss = metodo(ss)

                output_buffer = zz.texfile.replaceLatex(tex_path, ss)

            else:
                print("El archivo .tex no existe.")
                # Leer el archivo PDF desde el disco
                ss = zz.main.parametros_formulario(ss)

                """calculos"""
                listacalculos = [
                    m
                    for m in dir(codigo_al_final_del_latex.calculos)
                    if callable(getattr(codigo_al_final_del_latex.calculos, m))
                    and not m.startswith("__")
                ]
                listacalculos = [s for s in listacalculos if not s.startswith("xx")]
                for calculo in sorted(listacalculos):
                    metodo = getattr(codigo_al_final_del_latex.calculos, calculo, None)
                    ss = metodo(ss)

                with open(pdf_path, "rb") as file:
                    pdf_content = file.read()
                output_buffer = BytesIO(pdf_content)
                output_buffer.seek(0)
                output_buffer = zz.formulario.rellenoFormulario(output_buffer, ss)

            return output_buffer

            #######################

        def xlsxhojaconvalorespordefecto(df):
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # si es un nuevo archivo tex copia los parametros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            codigo, titulo = separar_y_mayusculizar(nombre)
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria
            # hasta que encuetre el problema de por que recarga el excel y no coje dela memoria

            #
            ruta_archivo_excel = f"./{codigo}.xlsx"
            if 1:
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja por defecto (llamada 'Sheet')
                    hoja_por_defecto = libro["Sheet"]
                    libro.remove(hoja_por_defecto)

                    hoja = libro.create_sheet(title=codigo)

                    # Seleccionar la hoja activa
                    # hoja = libro.active
                    # Guardar el nuevo libro de trabajo
                    libro.save(ruta_archivo_excel)

                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    hoja = libro[codigo]

                    # Borrar todo el contenido de la hoja
                    for fila in hoja.iter_rows():
                        for celda in fila:
                            celda.value = None

                    # Insertar los datos del DataFrame en la hoja de Excel
                    # Insertar la cabecera
                    hoja["A1"] = "Parámetro"
                    hoja["B1"] = "Valor"

                    # Insertar los datos del DataFrame en la hoja de Excel
                    for i, (parametro, valor) in enumerate(
                        zip(df["Descripción"], df["valor"]), start=2
                    ):
                        # hoja[f"A{i}"] = parametro.replace("_", "").replace(" ", "")
                        hoja[f"A{i}"] = parametro
                        hoja[f"B{i}"] = valor

                    # Guardar el archivo con los cambios
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    # print(ruta_archivo_excel)
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # Cargar el archivo Excel

            # Cargar todas las hojas del archivo Excel
            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )

            # Crear el diccionario data
            data = {
                sheet_name: [df.columns.tolist()]
                + df.values.tolist()  # Incluir cabecera y datos
                for sheet_name, df in sheets.items()
            }

            return data

    class formulario:
        def rellenoFormulario(pdf_buffer, ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            pdf_document = fitz.open(stream=pdf_buffer, filetype="pdf")
            output_buffer = BytesIO()
            variables = vars(ss)
            # print("CAMPOS DEL FORMULARIO")
            # Recorrer todas las páginas del PDF
            for page_num in range(len(pdf_document)):
                page = pdf_document.load_page(page_num)
                # Iterar sobre los widgets (campos de formulario)
                for widget in page.widgets():
                    field_name = widget.field_name  # Obtener el ss.nombre del campo
                    field_name = zz.texfile.normalizar_cadena(field_name)
                    # print(field_name)
                    if (
                        field_name in variables
                    ):  # Si el ss.nombre del campo está en el SimpleNamespace
                        widget.field_value = str(
                            variables[field_name]
                        )  # Asignar el valor
                        try:
                            widget.update()  # Actualizar el widget
                        except Exception as e:
                            """"""
                            print(f"Error al actualizar el campo {field_name}: {e}")
            pdf_document.save(output_buffer)
            pdf_document.close()
            output_buffer.seek(0)
            return output_buffer

    class texfile:

        def replaceLatex(ruta_tex, ss):
            """
            Compila un archivo .tex a PDF, reemplaza un texto y devuelve el PDF en memoria.
            :param ruta_tex: Ruta al archivo .tex original.
            :return: Objeto BytesIO con el contenido del PDF.
            """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            ruta_tex_absoluta = os.path.abspath(
                ruta_tex
            )  # Obtener la ruta absoluta del archivo original
            nombre_archivo = os.path.basename(ruta_tex)  # Nombre del archivo sin ruta

            # Crear un directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                ruta_tex_temp = os.path.join(temp_dir, nombre_archivo)

                contenido = file_tex()

                ##########################################

                if 1:
                    """figuras"""
                    listafiguras = [
                        m
                        for m in dir(codigo_al_final_del_latex.figuras)
                        if callable(getattr(codigo_al_final_del_latex.figuras, m))
                        and not m.startswith("__")
                    ]
                    listafiguras = [s for s in listafiguras if not s.startswith("xx")]
                    for figura in listafiguras:
                        metodo = getattr(
                            codigo_al_final_del_latex.figuras, figura, None
                        )
                        contenido = contenido.replace(f"ss.{figura}", metodo(ss))

                if 1:
                    """tablas"""
                    listatablas = [
                        m
                        for m in dir(codigo_al_final_del_latex.tablas)
                        if callable(getattr(codigo_al_final_del_latex.tablas, m))
                        and not m.startswith("__")
                    ]
                    listatablas = [s for s in listatablas if not s.startswith("xx")]
                    for tabla in listatablas:
                        # print(tabla)
                        metodo = getattr(
                            codigo_al_final_del_latex.tablas, tabla, None
                        )  # Obtener el método o None si no existe
                        contenido = contenido.replace(f"ss.{tabla}", metodo(ss))

                if 1:
                    """variables"""
                    # Reemplazo usando .replace() para cada variable
                    for variable, valor in vars(ss).items():
                        if (
                            valor is None
                            or (isinstance(valor, str) and valor.strip() == "")
                            or (isinstance(valor, pd.DataFrame) and valor.empty)
                            or (isinstance(valor, pd.Series) and valor.isna().all())
                        ):
                            valor = "%"

                        contenido = contenido.replace(f"ss.{variable}", str(valor))
                        print(f"ss.{variable}")

                # solo depuracion
                # solo depuracion
                # solo depuracion
                # solo depuracion
                with open(
                    "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.tex", "w", encoding="utf-8"
                ) as file:
                    file.write(contenido)
                # solo depuracion
                # solo depuracion
                # solo depuracion
                # solo depuracion

                ##########################################

                # Guardar los cambios en el archivo temporal
                with open(ruta_tex_temp, "w", encoding="utf-8") as file:
                    file.write(contenido)

                # Función para ejecutar pdflatex dos veces
                def ejecutar_pdflatex():
                    print(
                        f"\n{100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                    )
                    if 1:
                        subprocess.run(
                            ["pdflatex", "-output-directory", temp_dir, ruta_tex_temp],
                            check=True,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )

                    return True

                # # para que se cre el indice debe estar descomentado
                # Ejecutar pdflatex dos veces para generar la TOC correctamente
                if not (ejecutar_pdflatex() and ejecutar_pdflatex()):
                    return None  # Si la compilación falla, se detiene el proceso
                # ejecutar_pdflatex()

                # Ruta del PDF generado
                pdf_generado = os.path.join(
                    temp_dir, nombre_archivo.replace(".tex", ".pdf")
                )

                # Leer el PDF en memoria
                with open(pdf_generado, "rb") as file:
                    pdf_en_memoria = BytesIO(file.read())

            return pdf_en_memoria

        def normalizar_cadena(texto):
            if isinstance(texto, str):  # Asegurar que es un string
                texto = texto.lower()
                texto = "".join(
                    c
                    for c in unicodedata.normalize("NFD", texto)
                    if unicodedata.category(c) != "Mn"
                )
                # Capitalizar la primera letra después de un espacio
                texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
                # Eliminar espacios
                texto = re.sub(r"\s+", "", texto)
                # Mantener solo caracteres permitidos
                texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
            return texto


from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from PyPDF2 import PdfReader, PdfWriter
from io import BytesIO
import os
import matplotlib.pyplot as plt
import matplotlib

matplotlib.use("Agg")

import io
import base64
import pandas as pd
import fitz  # PyMuPDF
from types import SimpleNamespace
import subprocess
import shutil
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata

import numpy as np
import sys


router = APIRouter()


def separar_y_mayusculizar(texto):
    # Usamos una expresión regular para separar las mayúsculas
    partes = re.findall(r"[A-Z][^A-Z]*", texto)

    # Separar la primera parte (mantenerla como está)
    primera_parte = partes[0]

    # El resto de las partes lo convertimos a mayúsculas y las unimos con espacios
    resto = " ".join([parte.upper() for parte in partes[1:]])

    return primera_parte, resto


router = APIRouter()


@router.post("/Btx047")
def Btx047InstalaciónDeComunicaciónDeSaiLegrand(
    data: dict,
):

    # parra que en la primera carga, los campos que tgiene valores separados por ; para los selectores coja el primer valor por defecto, tambien 
    # debera hacerlo si se da a recalcular sin seleccionar un valor 
    for key, values in data.items():
        for i in range(1, len(values)):  # Saltamos la primera fila (encabezados)
            if isinstance(values[i][1], str):  # Verificar si es una cadena antes de dividir
                values[i][1] = values[i][1].split(';')[0]  # Tomar solo el primer valor
    # parra que en la primera carga, los campos que tgiene valores separados por ; para los selectores coja el primer valor por defecto, tambien 
    # debera hacerlo si se da a recalcular sin seleccionar un valor 
    for key, values in data.items():
        for i in range(1, len(values)):  # Saltamos la primera fila (encabezados)
            if isinstance(values[i][1], str):  # Verificar si es una cadena antes de dividir
                values[i][1] = values[i][1].split(';')[0]  # Tomar solo el primer valor
    print(data)

    ss = SimpleNamespace()

    # Ejemplo de uso
    ss.nombre = sys._getframe().f_code.co_name

    ss.codigo, ss.titulo = separar_y_mayusculizar(ss.nombre)

    ss.archivo_nuevo = True

    ss.data = data

    output_buffer = zz.main.main(ss)
    # solo depuracion
    # solo depuracion
    # solo depuracion
    with open(f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf", "wb") as f:
        f.write(output_buffer.getvalue())
    # solo depuracion
    # solo depuracion
    # solo depuracion
    # devuelve a fastapi el archivo
    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


if __name__ == "__main__":

    # Cambiar al directorio del script
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    df = codigo_al_final_del_latex.parametros.parametros()
    data = zz.main.xlsxhojaconvalorespordefecto(df)
    nombre_funcion = os.path.basename(__file__)[:-3]
    globals()[nombre_funcion](data)
