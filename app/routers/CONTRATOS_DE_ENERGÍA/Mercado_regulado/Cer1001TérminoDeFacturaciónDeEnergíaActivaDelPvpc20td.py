from fastapi import APIRouter

router = APIRouter()


@router.post(f"/Cer1001")
def Cer1001TÃ©rminoDeFacturaciÃ³nDeEnergÃ­aActivaDelPvpc20td(
    data: dict,
):
    data = comun.texfile.limpiar_valores(data)
    print(data)
    # parÃ¡metros
    pp = SimpleNamespace()
    # tablas y graficos en latex
    uu = SimpleNamespace()
    #
    vv = SimpleNamespace()
    xx = SimpleNamespace()
    yy = SimpleNamespace()
    # datos genericos
    zz = SimpleNamespace()

    zz.nombre = sys._getframe().f_code.co_name
    zz.ruta_script = os.path.dirname(
        os.path.abspath(__file__)
    )  # Obtener la carpeta sin el nombre del script

    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)
    zz.data = data

    output_buffer = comun.main.main(pp, uu, vv, xx, yy, zz)

    with open(f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf", "wb") as f:
        f.write(output_buffer.getvalue())

    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


def file_tex(pp, uu, vv, xx, yy, zz):

    return r"""
\newcommand{\xxMostrarVariablesAlFinal}{}

%\documentclass[a4paper,10pt]{article}
\documentclass[a4paper,10pt,twocolumn]{article}
\usepackage[top = 2cm, bottom = 2cm, left = 2cm, right = 2cm, asymmetric]{geometry} % Especificar los mÃ¡rgenes segÃºn la norma
\usepackage[spanish]{babel}    

\usepackage{graphicx}  % Para incluir imÃ¡genes si es necesario
\usepackage{amsmath, amssymb}  % Para expresiones matemÃ¡ticas
\usepackage{fancyhdr}  % Para personalizar encabezados
\usepackage{chngcntr}  % Para cambiar la numeraciÃ³n de apartados y subsecciones
\usepackage{tocbibind}  % Para incluir bibliografÃ­a en la tabla de contenidos
\usepackage{appendix}  % Para el formato de anexos
\usepackage{lipsum}  % Para generar texto de relleno (dummy text)
\usepackage{geometry}  % Para personalizar los mÃ¡rgenes
\usepackage{multicol}  % Para columnas
\usepackage{titlesec}  % Para personalizar los tÃ­tulos de las secciones
\usepackage[utf8]{inputenc}    
\usepackage[T1]{fontenc}       
\usepackage[pdftex,pdfencoding=auto, colorlinks=true, linkcolor=black, citecolor=black, filecolor=black, urlcolor=black]{hyperref}
\usepackage{tocloft}           
\usepackage{booktabs}          
\usepackage{float} % Paquete necesario para usar la opciÃ³n [H]
\usepackage{array}
\usepackage{longtable}
\usepackage{circuitikz}
\usepackage{tikz}
\usepackage{tikz-cd}
\usepackage{tikz-qtree}
\usepackage{smartdiagram}

\usepackage{qrcode}            % Paquete para generar QR
\usepackage{tabularx} % Agregar en el preÃ¡mbulo
\usepackage[absolute,overlay]{textpos} % Para posicionar objetos libremente
\usepackage{underscore}
\usepackage{pdfpages}

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{zz.titulo}  % Nombre del documento en el encabezado izquierdo
\fancyhead[C]{}  % Centro vacÃ­o
\fancyhead[R]{\thepage}  % NumeraciÃ³n de pÃ¡ginas en el encabezado derecho
% \renewcommand{\thepage}{\arabic{page}}  % Asegura la numeraciÃ³n de pÃ¡ginas en nÃºmeros arÃ¡bigos

% DefiniciÃ³n de los anexos
\newcommand{\annex}[1]{\section{Anexo #1} \addcontentsline{toc}{section}{Anexo #1}}
\addto\captionsspanish{%
  \renewcommand{\tablename}{Tabla}%
  \renewcommand{\listtablename}{Ãndice de tablas}%
}
\title{{\small Ref.:\uppercase{zz.codigo}}\\{\textbf{zz.titulo}}}
\author{
pp.ParticipantesP1NombreRaznSocial
}
\date{\today}





\begin{document}
\begin{Form}
%% --- PORTADA ---
%\begin{titlepage}
%    \centering
%    {\Large \textbf{TÃTULO DEL INFORME}}\[1.5cm]
%    %\includegraphics[width=0.4\textwidth]{logo.png}\[1cm] % Logo de la empresa/instituciÃ³n
%    
%    \textbf{Autor:} Nombre del Autor \[0.5cm]
%    \textbf{Fecha:} \today \[1cm]
%    \textbf{CÃ³digo del documento:} REF-XXXX-YYYY \[1cm]
%    \textbf{VersiÃ³n:} 1.0 \[1cm]
%    \textbf{Empresa / InstituciÃ³n:} Nombre \[2cm]
%    \vfill
%\end{titlepage}
	\maketitle
    %
	%\begin{abstract}

    %\end{abstract}
% --- ÃNDICE ---
	\tableofcontents  % Tabla de contenidos
	\listoffigures    % Lista de graficos
	\listoftables     % Lista de tablas
	\newpage



    

    













    




























































\begin{abstract}
Este informe analiza la evoluciÃ³n del Precio Voluntario para el PequeÃ±o Consumidor (PVPC) durante los Ãºltimos 12 meses, examinando tendencias anuales, variaciones semanales y diferencias entre periodos tarifarios. Los datos proceden oficialmente del sistema de informaciÃ³n de Red ElÃ©ctrica EspaÃ±ola (ESIOS).
\end{abstract}


\section{IntroducciÃ³n}

El Real Decreto 244/2019 regula el mecanismo de compensaciÃ³n simplificada para excedentes de autoconsumo. 

Este estudio utiliza datos oficiales publicados en el portal \href{https://www.esios.ree.es/es/pvpc}{ESIOS de REE} \cite{esios}, analizando:

\begin{itemize}
\item EvoluciÃ³n interanual
\item Patrones semanales
\item Diferencias entre periodos tarifarios (Punta, Llano, Valle)
\end{itemize}

\section{MetodologÃ­a}
\subsection{Fuente de datos}
Todos los datos proceden de la API pÃºblica de \textbf{ESIOS (Red ElÃ©ctrica EspaÃ±ola)} \cite{esios}, accediendo a:

\begin{itemize}
\item Series temporales horarias del PVPC
\item Metadatos de periodos tarifarios
\item Indicadores de mercado spot
\end{itemize}

\subsection{Procesamiento}
Los datos se procesaron mediante:

\begin{enumerate}
\item ExtracciÃ³n mediante API REST (Ãºltimos 365 dÃ­as)
\item ClasificaciÃ³n por periodos tarifarios (RD 216/2014)
\item CÃ¡lculo de medias mÃ³viles (7 dÃ­as)
\item AgregaciÃ³n temporal (hora â†’ dÃ­a â†’ semana â†’ mes)
\end{enumerate}

\section{Resultados}

\subsection{EvoluciÃ³n anual}
\subsubsection{PVPC}
uu.d04PrecioDeLaEnergÃ­aElÃ©ctricaPvpcAnual


\subsection{EvoluciÃ³n semanal}
uu.d14PrecioDeLaEnergÃ­aElÃ©ctricaPvpcSemanal




\subsection{AnÃ¡lisis por periodos}
uu.tablad21PvpcPorPeriodosTarifarios
uu.d21PvpcPorPeriodosTarifarios



La Tabla \ref{tab:periodos} confirma la diferencia significativa entre periodos.














\section{Precio en horario soleado}

\subsection{DefiniciÃ³n de Horarios}
Se clasifican las horas en dos categorÃ­as:

\begin{equation}
\text{H}(t) = 
\begin{cases} 
\text{Soleado} & 
\text{si } 
\\ (m_t \in [4,9] \cap h_t \in [8,20)) 
\\ 
\lor \\(m_t \notin [4,9] \cap h_t \in [10,17)) \\
\text{No Soleado} & \text{en otro caso}
\end{cases}
\end{equation}

donde:
\begin{itemize}
\item $m_t$: Mes (1=Enero, ..., 12=Diciembre)
\item $h_t$: Hora del dÃ­a (0, ..., 23)
\end{itemize}

\subsection{CÃ¡lculo de Precios Medios}
Para cada categorÃ­a $s \in \{\text{Soleado}, \text{No Soleado}\}$:

\begin{equation}
P_s = \frac{1}{|T_s|}\sum_{t \in T_s} \text{PVPC}(t)
\end{equation}

donde:
\begin{itemize}
\item $T_s$: Conjunto de periodos en el horario $s$
\item $|T_s|$: Cardinalidad (nÃºmero de horas)
\item PVPC$(t)$: Precio en el periodo $t$
\end{itemize}

\subsection{Resultados}

uu.tablad51PrecioExcedentesPorPeriodoSoleado


uu.d51PrecioPorPeriodoSoleado



























































    







    

















% --- 7. REFERENCIAS BIBLIOGRÃFICAS ---
\begin{thebibliography}{99}
    \bibitem{ccee} \href{https://www.boe.es/buscar/pdf/2021/BOE-A-2021-7120-consolidado.pdf}{ResoluciÃ³n de 28 de abril de 2021, de la DirecciÃ³n General de PolÃ­tica EnergÃ©tica y Minas, por la que se establece el contenido
mÃ­nimo y el modelo de factura de electricidad a utilizar por los comercializadores de referencia.}

\bibitem{cnmc}
\href{https://comparador.cnmc.gob.es/}{CNMC. Comparador de Ofertas de EnergÃ­a}


\bibitem{rd244}
\href{}{RD 244/2019  sobre autoconsumo}

\bibitem{esios} 
\href{https://www.esios.ree.es/es/pvpc}{ESIOS - Red ElÃ©ctrica de EspaÃ±a. PVPC y datos del sistema elÃ©ctrico
}


\bibitem{boe}
\href{}{Real Decreto 216/2014 por el que se establece la metodologÃ­a de cÃ¡lculo de los precios voluntarios para el pequeÃ±o consumidor.}


\end{thebibliography}





































\ifdefined\MostrarVariablesAlFinal
\newpage
\onecolumn
HojaDeDatos
\fi
\end{Form}
\end{document}

"""


class codigo_para_replace_en_el_latex:

    class datos:

        def d21PvpcPorPeriodosTarifarios(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            df = xx.d04PrecioDeLaEnergÃ­aElÃ©ctricaPvpcAnual.copy()

            # Suponiendo que ya tienes:
            df["datetime"] = pd.to_datetime(df.index, utc=True)
            df["hour"] = df["datetime"].dt.hour
            df["week"] = df["datetime"].dt.isocalendar().week
            df["year"] = df["datetime"].dt.isocalendar().year
            df["day_of_week"] = df["datetime"].dt.dayofweek  # 0 = lunes, 6 = domingo
            df["month"] = df["datetime"].dt.month

            # 1. Definir temporada (Alta = verano, Baja = invierno)
            df["temporada"] = "Baja"  # Por defecto: octubre a mayo
            df.loc[df["month"].isin([6, 7, 8, 9]), "temporada"] = (
                "Alta"  # Junio a septiembre
            )

            # 2. Asignar periodos para la tarifa 2.0TD (3 periodos)
            def asignar_periodo_20TD(row):
                if row["day_of_week"] >= 5:  # Fin de semana (siempre Valle)
                    return 3
                if row["temporada"] == "Alta":  # Verano
                    if (11 <= row["hour"] < 15) or (
                        18 <= row["hour"] < 22
                    ):  # Punta verano
                        return 1
                    elif (
                        (8 <= row["hour"] < 11)
                        or (15 <= row["hour"] < 18)
                        or (22 <= row["hour"] < 24)
                    ):  # Llano verano
                        return 2
                else:  # Invierno
                    if (10 <= row["hour"] < 14) or (
                        18 <= row["hour"] < 22
                    ):  # Punta invierno
                        return 1
                    elif (
                        (8 <= row["hour"] < 10)
                        or (14 <= row["hour"] < 18)
                        or (22 <= row["hour"] < 24)
                    ):  # Llano invierno
                        return 2
                return 3  # Valle (00:00-08:00)

            df["periodo_20TD"] = df.apply(asignar_periodo_20TD, axis=1)

            # 3. Asignar periodos para la tarifa 3.0TD o 6.0TD (6 periodos)
            def asignar_periodo_30TD(row):
                if row["day_of_week"] >= 5:  # Fin de semana (siempre P3 o P6)
                    return 6 if row["temporada"] == "Baja" else 3
                if row["temporada"] == "Alta":  # Verano (P1, P2, P3)
                    if (11 <= row["hour"] < 15) or (
                        18 <= row["hour"] < 22
                    ):  # P1 (Punta verano)
                        return 1
                    elif (
                        (8 <= row["hour"] < 11)
                        or (15 <= row["hour"] < 18)
                        or (22 <= row["hour"] < 24)
                    ):  # P2 (Llano verano)
                        return 2
                    else:  # P3 (Valle verano: 00:00-08:00)
                        return 3
                else:  # Invierno (P4, P5, P6)
                    if (10 <= row["hour"] < 14) or (
                        18 <= row["hour"] < 22
                    ):  # P4 (Punta invierno)
                        return 4
                    elif (
                        (8 <= row["hour"] < 10)
                        or (14 <= row["hour"] < 18)
                        or (22 <= row["hour"] < 24)
                    ):  # P5 (Llano invierno)
                        return 5
                    else:  # P6 (Valle invierno: 00:00-08:00)
                        return 6

            df["periodo_30TD"] = df.apply(asignar_periodo_30TD, axis=1)

            #
            #
            #
            #
            #
            #

            # Calcular media del PVPC por periodo_20TD
            media_por_periodo = (
                df.groupby("periodo_20TD")["PVPC"]
                .mean()
                .round(2)  # Redondear a 2 decimales
                .reset_index()
            )

            # Media global (todos los periodos)
            media_global = df["PVPC"].mean().round(2)

            # Crear JSON
            resultado_json = {
                "ultimos_12_meses": {
                    "media_precio_por_periodo": media_por_periodo.to_dict(
                        orient="records"
                    ),
                    "media_global_precio": media_global,
                }
            }

            # Crear DataFrame combinando ambos datos
            df = pd.DataFrame(
                {
                    "periodo": [1, 2, 3, "Global"],
                    "EUR/MWh medio": [
                        resultado_json["ultimos_12_meses"]["media_precio_por_periodo"][
                            0
                        ]["PVPC"],
                        resultado_json["ultimos_12_meses"]["media_precio_por_periodo"][
                            1
                        ]["PVPC"],
                        resultado_json["ultimos_12_meses"]["media_precio_por_periodo"][
                            2
                        ]["PVPC"],
                        resultado_json["ultimos_12_meses"]["media_global_precio"],
                    ],
                    "tipo": ["Punta", "Llano", "Valle", "Global"],
                }
            )

            df = df.set_index("periodo").drop(columns="tipo")
            df.index.name = "EUR/MWh_periodo"

            # Guardar en la instancia xx como atributo dinÃ¡mico
            setattr(yy, sys._getframe().f_code.co_name, df)  # Valores originales
            return pp, uu, vv, xx, yy, zz

        def d04PrecioDeLaEnergÃ­aElÃ©ctricaPvpcAnual(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            """Procesa los datos descargados y los convierte en DataFrames"""

            indicator_id = "1001"  # ID del indicador

            ruta = Path(f"{zz.ruta_script}/assets")  # Carpeta donde buscar

            # Buscar archivo con el patrÃ³n "indicador_1001_*.json"
            archivo_json = next(ruta.glob(f"indicador_{indicator_id}_*.json"), None)

            if archivo_json:
                with open(archivo_json, "r", encoding="utf-8") as f:
                    data = json.load(f)

            # Extraer valores y convertir a DataFrame
            valores = data["indicator"]["values"]
            df = pd.DataFrame(valores)

            # Filtrar solo geo_name "PenÃ­nsula" o "EspaÃ±a"
            df = df[df["geo_name"].isin(["PenÃ­nsula", "EspaÃ±a"])]

            # Convertir la columna datetime a formato datetime
            df["datetime"] = pd.to_datetime(df["datetime"])

            # Ordenar por fecha
            df = df.sort_values("datetime")

            # Mantener solo dos columnas: datetime (Ã­ndice) y Value
            df = (
                df[["datetime", "value"]]
                .rename(columns={"value": "PVPC"})
                .set_index("datetime")
            )
            df["PVPC medio semanal"] = df["PVPC"].rolling(window=24 * 7).mean()

            # Asignar nombre al Ã­ndice y a las columnas
            df.index.name = "EUR/MWh_Fecha"

            # Guardar en la instancia xx como atributo dinÃ¡mico
            setattr(xx, sys._getframe().f_code.co_name, df)  # Valores originales
            return pp, uu, vv, xx, yy, zz

        def d14PrecioDeLaEnergÃ­aElÃ©ctricaPvpcSemanal(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            """Procesa los datos descargados y los convierte en DataFrames"""

            df = xx.d04PrecioDeLaEnergÃ­aElÃ©ctricaPvpcAnual.copy()

            # Convertir a datetime y extraer componentes
            df["datetime"] = pd.to_datetime(df.index, utc=True)
            df["hour"] = df["datetime"].dt.hour
            df["week"] = df["datetime"].dt.isocalendar().week
            df["year"] = df["datetime"].dt.isocalendar().year
            df["day_of_week"] = df["datetime"].dt.dayofweek  # 0 = lunes, 6 = domingo

            # Calcular la media de 'PVPC' por hora y dÃ­a de semana
            media_pvpc = (
                df.groupby(["day_of_week", "hour"])["PVPC"].mean().reset_index()
            )

            # Calcular la hora semanal (0 a 167)
            media_pvpc["hora_semanal"] = (
                media_pvpc["day_of_week"] * 24 + media_pvpc["hour"]
            )

            # Crear el DataFrame final ordenado por hora semanal
            resultado = (
                media_pvpc[["hora_semanal", "PVPC"]]
                .sort_values("hora_semanal")
                .set_index("hora_semanal")
                .rename(columns={"PVPC": "PVPC medio"})
            )

            resultado.index.name = "EUR/MWh_hora semanal"

            # Guardar en la instancia xx como atributo dinÃ¡mico
            setattr(xx, sys._getframe().f_code.co_name, resultado)  # Valores originales
            print(xx)
            return pp, uu, vv, xx, yy, zz

        def d51PrecioPorPeriodoSoleado(pp, uu, vv, xx, yy, zz):
            print(f"\n{100*'+'}{__class__.__name__}.{sys._getframe().f_code.co_name}")
            df = xx.d04PrecioDeLaEnergÃ­aElÃ©ctricaPvpcAnual.copy()

            # Convertir a datetime y extraer componentes
            df["datetime"] = pd.to_datetime(df.index, utc=True)
            df["hour"] = df["datetime"].dt.hour
            df["month"] = df["datetime"].dt.month

            # Definir horas soleadas (funciÃ³n simplificada)
            def es_horario_soleado(row):
                if (row["month"] in [4, 5, 6, 7, 8, 9] and 8 <= row["hour"] < 20) or (
                    row["month"] in [1, 2, 3, 10, 11, 12] and 10 <= row["hour"] < 17
                ):
                    return True
                return False

            # Aplicar clasificaciÃ³n
            df["horario_soleado"] = df.apply(es_horario_soleado, axis=1)

            # Calcular medias simples
            precio_soleado = (
                df[df["horario_soleado"]]["PVPC"].mean().round(2)
            )
            precio_no_soleado = (
                df[~df["horario_soleado"]]["PVPC"].mean().round(2)
            )

            df_resultado = pd.DataFrame(
                {
                    "Horario": ["Soleado", "No Soleado"],
                    "Precio_Medio": [precio_soleado, precio_no_soleado],
                }
            ).set_index("Horario")

            df_resultado.index.name = "EUR/MWh_Periodo"

            # Guardar en la instancia xx como atributo dinÃ¡mico
            setattr(
                yy, sys._getframe().f_code.co_name, df_resultado
            )  # Valores originales
            return pp, uu, vv, xx, yy, zz

        def TodosLosExcelDeLaCarpetaXlsx(pp, uu, vv, xx, yy, zz):
            """para los datos que no hay ques esta descargancodon continuamente por ejm los pvpc..."""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpetaimagenes = Path(f"{zz.ruta_script}/assets")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.xlsx")]

            for imagen in imagenes:
                # _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                titulo = imagen[:-4]
                # Envolvemos el cÃ³digo generado con el entorno 'tabularx' para usar el ancho proporcional

                ruta_archivo_excel = Path(f"{zz.ruta_script}/assets/{imagen}")
                # libro = openpyxl.load_workbook(ruta_archivo_excel)

                # Cargar todas las hojas del archivo Excel
                sheets = pd.read_excel(
                    ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
                )
                for nombre_hoja, df in sheets.items():
                    nombre_hoja = comun.texfile.normalizar_cadena(nombre_hoja)
                    setattr(xx, f"{nombre_hoja}", df)  # Guardar en ss con nombre limpio

            return pp, uu, vv, xx, yy, zz

    class calculo:

        def C01CalculoEjemplo(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'} {__class__.__name__}.{sys._getframe().f_code.co_name}")
            # Crear un DataFrame con nÃºmeros aleatorios entre 0 y 100
            df = pd.DataFrame(
                np.random.randint(0, 100, size=(2, 3)), columns=["A", "B", "C"]
            )

            pp.TablaEjemplo = df.copy()
            return pp, uu, vv, xx, yy, zz

    class tabla:

        def T00df2tablas(pp, uu, vv, xx, yy, zz):
            # tabla de todos los df

            for var_name, value in pp.__dict__.items():
                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                if var_name == zz.codigo:
                    titulo = "ParÃ¡metros"

                if isinstance(value, pd.DataFrame):
                    # Redondear los valores del DataFrame a 1 decimal
                    value = value.round(1)

                    latex_code = value.to_latex(float_format="%.1f")

                    latex_code_with_resizebox = rf"""
                    \begin{{table}}[h!] \centering
                        % \resizebox{{0.4\textwidth}}{{!}}
                        {{
                        {latex_code}
                        }}
                        \caption{{{titulo}}}
                    \end{{table}}
                    """

                    # Asignar el LaTeX a la propiedad correspondiente de uu
                    setattr(uu, f"{var_name}", latex_code_with_resizebox)

            # Convierte todos los df de uu a tablas
            return pp, uu, vv, xx, yy, zz

        def T00df2tablayy(pp, uu, vv, xx, yy, zz):
            # tabla de todos los df

            for var_name, value in yy.__dict__.items():
                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                if var_name == zz.codigo:
                    titulo = "ParÃ¡metros"

                if isinstance(value, pd.DataFrame):
                    print(9999999999999999999999)
                    print(var_name)

                    # Redondear los valores del DataFrame a 1 decimal
                    # value = value.round(1)

                    latex_code = value.to_latex()

                    latex_code_with_resizebox = rf"""
                    \begin{{table}}[h!] \centering
                        % \resizebox{{0.4\textwidth}}{{!}}
                        {{
                        {latex_code}
                        }}
                        \caption{{{titulo}}}
                    \end{{table}}
                    """

                    # Asignar el LaTeX a la propiedad correspondiente de uu
                    setattr(uu, f"tabla{var_name}", latex_code_with_resizebox)

            # Convierte todos los df de uu a tablas
            return pp, uu, vv, xx, yy, zz

    class grafico:

        def g00df2graficos(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # tabla de todos los df

            for var_name, value in {**xx.__dict__, **yy.__dict__}.items():
                print(var_name)
                if hasattr(xx, var_name):
                    df = getattr(xx, var_name)
                else:
                    df = getattr(yy, var_name)

                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                # 1. Crear un archivo temporal para la imagen
                with tempfile.NamedTemporaryFile(
                    suffix=".png", delete=False
                ) as temp_image:
                    temp_image_path2 = (
                        temp_image.name
                    )  # Guardar la ruta del archivo temporal

                # ðŸ“Œ GrÃ¡fico del flujo de caja y acumulado
                if len(df) < 10:
                    fig, ax = plt.subplots(figsize=(5, 3))

                    df.plot(kind="bar", ax=ax)

                    plt.xticks(rotation=45)
                    plt.tight_layout()  # Ajustar para que no se corten las etiquetas

                    ax.set_ylabel(df.index.name.split("_")[0])
                    ax.set_xlabel(df.index.name.split("_")[1])
                else:

                    fig, ax = plt.subplots(figsize=(6, 5))

                    df.plot(ax=ax)

                    # Obtener el nombre de la columna automÃ¡ticamente
                    try:
                        ax.set_ylabel(df.index.name.split("_")[0])
                        ax.set_xlabel(df.index.name.split("_")[1])
                    except:
                        """"""

                plt.grid(axis="y", linestyle="--", alpha=0.7)
                plt.savefig(temp_image_path2, format="png", dpi=300)
                # Envolvemos el cÃ³digo generado con el entorno 'tabularx' para usar el ancho proporcional
                latex_code = (
                    ""
                    + r"""\begin{figure}[H] \centering
                    """
                    + rf"""
                    \includegraphics[width=.5\textwidth]{{{temp_image_path2}}}
                    \caption{{{titulo}}}
                    """
                    + """
                    \label{fig:dfassssdfsa}
                    \end{figure}
                    """
                )
                setattr(uu, f"{var_name}", latex_code)
            return pp, uu, vv, xx, yy, zz

        def f01figuraEjemplo(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            titulo = sys._getframe().f_code.co_name
            titulo = comun.texfile.separar_y_formatear(titulo)
            # 1. Crear un archivo temporal para la imagen
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image:
                temp_image_path1 = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal

            # ðŸ“Œ GrÃ¡fico del flujo de caja y acumulado
            fig, ax = plt.subplots(figsize=(6, 5))

            pp.TablaEjemplo.plot(ax=ax)

            plt.savefig(temp_image_path1, format="png", dpi=300)
            # Envolvemos el cÃ³digo generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + r"""\begin{figure}[H] \centering
                """
                + rf"""
                \includegraphics[width=.5\textwidth]{{{temp_image_path1}}}
                \caption{{{titulo}}}
                """
                + """
                \label{fig:dfassssdfsa}
                \end{figure}
                """
            )
            setattr(uu, f"{sys._getframe().f_code.co_name}", latex_code)
            return pp, uu, vv, xx, yy, zz

    class imagen:

        def TodasLasImagnesDeLaCarpetaImg(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpetaimagenes = Path(f"{zz.ruta_script}/{zz.codigo}/img")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.png")] + [
                f.name for f in carpetaimagenes.glob("*.jpg")
            ]

            for imagen in imagenes:
                # _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                titulo = imagen[:-4]
                # Envolvemos el cÃ³digo generado con el entorno 'tabularx' para usar el ancho proporcional

                latex_code = (
                    ""
                    + r"""\begin{figure}[H] \centering
                    """
                    + rf"""
                    \includegraphics[width=0.7\textwidth]{{{Path(f"{zz.ruta_script}/{zz.codigo}/img/{imagen}")}}} 
                    \caption{{{titulo}}}
                    """
                    + """
                    \label{fig:dfasdfsa}
                    \end{figure}
                    """
                )

                setattr(uu, f"imagen{imagen}", latex_code)
            return pp, uu, vv, xx, yy, zz


########################################################################################################
########################################################################################################


class comun:
    class main:

        def main(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # segun si es un formulario o un documento nuevo

            pp, uu, vv, xx, yy, zz = comun.main.data2parametros(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.datos(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.calulos(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.tablas(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.graficos(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.imagenes(pp, uu, vv, xx, yy, zz)

            output_buffer = comun.main.replaceLatex(pp, uu, vv, xx, yy, zz)

            # print(uu)
            return output_buffer

        def data2parametros(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Para cada hoja en el JSON
            # Convertir cada valor del diccionario en un DataFrame
            for key, value in zz.data.items():
                key_sanitized = comun.texfile.sanitize_name(
                    key
                )  # Limpiar el nombre de la clave
                df = pd.DataFrame(
                    value[1:], columns=value[0]
                )  # Primera fila como encabezado
                df.set_index(
                    df.columns[0], inplace=True
                )  # Usar la primera columna como Ã­ndice
                setattr(pp, f"df{key_sanitized}", df)  # Guardar en ss con nombre limpio
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == zz.codigo:
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar Ã­ndice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{idx_sanitized}",
                                val,
                            )
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == "Coordenadas" or key_sanitized == "Participantes":
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar Ã­ndice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{key_sanitized}{col_sanitized}{idx_sanitized}",
                                val,
                            )

            return pp, uu, vv, xx, yy, zz

        def datos(pp, uu, vv, xx, yy, zz):
            """datoes"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listadatoes = [
                m
                for m in dir(codigo_para_replace_en_el_latex.datos)
                if callable(getattr(codigo_para_replace_en_el_latex.datos, m))
                and not m.startswith("__")
            ]
            listadatoes = [s for s in listadatoes if not s.startswith("__")]

            for dato in listadatoes:
                metodo = getattr(codigo_para_replace_en_el_latex.datos, dato, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def calulos(pp, uu, vv, xx, yy, zz):
            """calculos, valores calculados en yy"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listacalculos = [
                m
                for m in dir(codigo_para_replace_en_el_latex.calculo)
                if callable(getattr(codigo_para_replace_en_el_latex.calculo, m))
                and not m.startswith("__")
            ]
            listacalculos = [s for s in listacalculos if not s.startswith("__")]
            for calculo in sorted(listacalculos):
                metodo = getattr(codigo_para_replace_en_el_latex.calculo, calculo, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)

            return pp, uu, vv, xx, yy, zz

        def graficos(pp, uu, vv, xx, yy, zz):
            """graficos"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listafiguras = [
                m
                for m in dir(codigo_para_replace_en_el_latex.grafico)
                if callable(getattr(codigo_para_replace_en_el_latex.grafico, m))
                and not m.startswith("__")
            ]
            listafiguras = [s for s in listafiguras if not s.startswith("__")]
            for grafico in listafiguras:
                metodo = getattr(codigo_para_replace_en_el_latex.grafico, grafico, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def imagenes(pp, uu, vv, xx, yy, zz):
            """imagenes"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listaimagenes = [
                m
                for m in dir(codigo_para_replace_en_el_latex.imagen)
                if callable(getattr(codigo_para_replace_en_el_latex.imagen, m))
                and not m.startswith("__")
            ]
            listaimagenes = [s for s in listaimagenes if not s.startswith("__")]

            for imagen in listaimagenes:
                metodo = getattr(codigo_para_replace_en_el_latex.imagen, imagen, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def tablas(pp, uu, vv, xx, yy, zz):
            """tablas"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listatablas = [
                m
                for m in dir(codigo_para_replace_en_el_latex.tabla)
                if callable(getattr(codigo_para_replace_en_el_latex.tabla, m))
                and not m.startswith("__")
            ]
            listatablas = [s for s in listatablas if not s.startswith("__")]
            for tabla in listatablas:
                # print(tabla)
                metodo = getattr(
                    codigo_para_replace_en_el_latex.tabla, tabla, None
                )  # Obtener el mÃ©todo o None si no existe
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)

            return pp, uu, vv, xx, yy, zz

        def replaceLatex(pp, uu, vv, xx, yy, zz):
            """
            Compila un archivo .tex a PDF, reemplaza un texto y devuelve el PDF en memoria.
            :param ruta_tex: Ruta al archivo .tex original.
            :return: Objeto BytesIO con el contenido del PDF.
            """

            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            pdf_path = Path(zz.nombre).with_suffix(".pdf")
            ruta_tex = f"{os.path.splitext(pdf_path)[0]}.tex"

            nombre_archivo = os.path.basename(ruta_tex)  # Nombre del archivo sin ruta

            # Crear un directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                ruta_tex_temp = os.path.join(temp_dir, nombre_archivo)

                contenido = file_tex(pp, uu, vv, xx, yy, zz)
                # tabla final con todas los parÃ¡metros  y variables
                HojaDeDatos = "\n"

                for var_name, value in pp.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\npp. {var_name}\n\npp.{var_name}"

                for var_name, value in uu.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nuu. {var_name}\n\nuu.{var_name}"

                for var_name, value in vv.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nvv. {var_name}\n\nvv.{var_name}"

                for var_name, value in xx.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nxx. {var_name}\n\nxx.{var_name}"

                for var_name, value in yy.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nyy. {var_name}\n\nyy.{var_name}"

                # for var_name, value in zz.__dict__.items():
                #     print(f"zz.{var_name}")
                #     HojaDeDatos = HojaDeDatos + f"\n\nzz. {var_name}\n\nzz.{var_name}"

                contenido = contenido.replace(f"HojaDeDatos", HojaDeDatos)

                if 1:
                    # Reemplazar directamente todas las variables en el texto
                    for var_name, value in pp.__dict__.items():
                        # print(f"pp.{var_name}")
                        contenido = contenido.replace(
                            f"pp.{var_name}",
                            f"{str(value)}",
                        )
                    for var_name, value in uu.__dict__.items():
                        # print(f"uu.{var_name}")
                        contenido = contenido.replace(f"uu.{var_name}", f"{str(value)}")

                    for var_name, value in vv.__dict__.items():
                        # print(f"vv.{var_name}")
                        contenido = contenido.replace(f"vv.{var_name}", f"{str(value)}")

                    for var_name, value in xx.__dict__.items():
                        # print(f"xx.{var_name}")
                        contenido = contenido.replace(f"xx.{var_name}", f"{str(value)}")

                    for var_name, value in yy.__dict__.items():
                        # print(f"yy.{var_name}")
                        contenido = contenido.replace(f"yy.{var_name}", f"{str(value)}")

                    for var_name, value in zz.__dict__.items():
                        # print(f"zz.{var_name}")
                        contenido = contenido.replace(f"zz.{var_name}", f"{str(value)}")

                with open(
                    "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.tex", "w", encoding="utf-8"
                ) as file:
                    file.write(contenido)

                # Guardar los cambios en el archivo temporal
                with open(ruta_tex_temp, "w", encoding="utf-8") as file:
                    file.write(contenido)

                # FunciÃ³n para ejecutar pdflatex dos veces
                def ejecutar_pdflatex():
                    print(
                        f"\n{100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                    )
                    if 1:
                        subprocess.run(
                            ["pdflatex", "-output-directory", temp_dir, ruta_tex_temp],
                            check=True,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )

                    return True

                # # para que se cre el indice debe estar descomentado
                # Ejecutar pdflatex dos veces para generar la TOC correctamente
                if not (ejecutar_pdflatex() and ejecutar_pdflatex()):
                    return None  # Si la compilaciÃ³n falla, se detiene el proceso
                # ejecutar_pdflatex()

                # Ruta del PDF generado
                pdf_generado = os.path.join(
                    temp_dir, nombre_archivo.replace(".tex", ".pdf")
                )

                # Leer el PDF en memoria
                with open(pdf_generado, "rb") as file:
                    pdf_en_memoria = BytesIO(file.read())

            return pdf_en_memoria

            #######################

    class texfile:

        def normalizar_cadena(texto):
            if isinstance(texto, str):  # Asegurar que es un string
                texto = texto.lower()
                texto = "".join(
                    c
                    for c in unicodedata.normalize("NFD", texto)
                    if unicodedata.category(c) != "Mn"
                )
                # Capitalizar la primera letra despuÃ©s de un espacio
                texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
                # Eliminar espacios
                texto = re.sub(r"\s+", "", texto)
                # Mantener solo caracteres permitidos
                texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
            return texto

        def limpiar_valores(data):
            for key, values in data.items():
                for i in range(
                    1, len(values)
                ):  # Saltamos la primera fila (encabezados)
                    for j in range(
                        1, len(values[i])
                    ):  # Recorremos todas las columnas (excepto encabezados)
                        if isinstance(
                            values[i][j], str
                        ):  # Verificar si es una cadena antes de dividir
                            valor_limpio = (
                                values[i][j].split(";")[0].strip()
                            )  # Tomar solo el primer valor y limpiar espacios

                            # Intentar convertir a flotante si es posible
                            try:
                                values[i][j] = float(valor_limpio)
                            except ValueError:
                                values[i][
                                    j
                                ] = valor_limpio  # Si no se puede convertir, dejar como estÃ¡
            return data

        def separar_y_mayusculizar(texto):
            # Usamos una expresiÃ³n regular para separar las mayÃºsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            # Separar la primera parte (mantenerla como estÃ¡)
            primera_parte = partes[0]

            # El resto de las partes lo convertimos a mayÃºsculas y las unimos con espacios
            resto = " ".join([parte.upper() for parte in partes[1:]])
            # resto = " ".join([parte for parte in partes[1:]])

            return primera_parte, resto

        def separar_y_formatear(texto):
            # Usamos una expresiÃ³n regular para separar las mayÃºsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            if not partes:
                return ""

            # Convertimos todo a minÃºsculas y capitalizamos solo la primera letra
            resultado = " ".join([parte.lower() for parte in partes])
            return resultado.capitalize()

        # FunciÃ³n para limpiar nombres de variables
        def sanitize_name(name):
            return re.sub(
                r"[^a-zA-Z0-9]", "", str(name)
            )  # Elimina cualquier carÃ¡cter no alfanumÃ©rico


class especifico:
    class previo:
        def p00parametros():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            nombre_funcion = os.path.basename(__file__)[:-3]
            codigo, _ = comun.texfile.separar_y_mayusculizar(nombre_funcion)
            nan = " "
            print(codigo)
            zz.data = {
                zz.codigo: [
                    ["Parametro", "Valor"],
                    ["InversiÃ³n inicial [â‚¬]", 0],
                    ["BonificaciÃ³n  inicial [â‚¬]", 10000],
                    ["DuraciÃ³n de la actuaciÃ³n [aÃ±os]", 15],
                    ["PrÃ©stamo recibido [â‚¬]", 100000],
                    ["Tasa de interÃ©s del prÃ©stamo [\%]", 5],
                    ["Plazo del prÃ©stamo [aÃ±os]", 15],
                    ["InflaciÃ³n annual [\%]", 3],
                    ["Ahorro annual [â‚¬]", 25000],
                    ["Costos fijos operativos [â‚¬]", 1000],
                    ["Tasa de descuento [\%]", 6],
                ],
                "Coordenadas": [
                    ["ParÃ¡metro", "Valor"],
                    ["lat", 36.6641681434112],
                    ["lng", -4.45863940137516],
                ],
                "Participantes": [
                    ["ParÃ¡metro", "P1"],
                    ["Rol", ""],
                    ["DNI/NIF/NIE", ""],
                    [
                        "Nombre/RazÃ³n Social",
                        "'EDITAR' para poner tu nombre y ajustar los parÃ¡metros.",
                    ],
                    ["TitulaciÃ³n", ""],
                ],
            }

        def p01carpetaauxiliar():
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpeta = Path(f"{zz.ruta_script}/assets")

            if not os.path.exists(carpeta):
                os.mkdir(carpeta)

        def p03apipvpcesios():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import requests
            import json
            import os
            from datetime import datetime, timedelta

            def obtener_fechas_descarga():
                """Calcula las fechas de inicio y fin para los Ãºltimos 12 meses completos"""
                hoy = datetime.today()
                primer_dia_mes_actual = datetime(hoy.year, hoy.month, 1)
                ultimo_dia_mes_anterior = primer_dia_mes_actual - timedelta(days=1)
                primer_dia_mes_anterior = datetime(
                    ultimo_dia_mes_anterior.year, ultimo_dia_mes_anterior.month, 1
                )

                primer_dia_12_meses_atras = primer_dia_mes_anterior - timedelta(
                    days=365
                )
                fecha_inicio = primer_dia_12_meses_atras.strftime("%Y-%m-%dT00:00:00")
                fecha_fin = ultimo_dia_mes_anterior.strftime("%Y-%m-%dT23:59:59")

                return fecha_inicio, fecha_fin

            def descargar_datos_api(indicator_ids, fecha_inicio, fecha_fin, api_key):
                """Descarga datos de la API y los guarda en archivos JSON"""
                headers = {
                    "Accept": "application/json; application/vnd.esios-api-v1+json",
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                }

                datos_descargados = {}

                for indicator_id in indicator_ids:

                    json_filename = f"{zz.ruta_script}/assets/indicador_{indicator_id}_{fecha_inicio[:7]}_a_{fecha_fin[:7]}.json"

                    if os.path.exists(json_filename):
                        print(
                            f"El archivo para el indicador {indicator_id} ya existe: {json_filename}. No se descargarÃ¡ nuevamente."
                        )
                        with open(json_filename, "r", encoding="utf-8") as json_file:
                            datos_descargados[indicator_id] = json.load(json_file)
                    else:
                        print(f"Descargando datos para el indicador {indicator_id}...")
                        url = f"https://api.esios.ree.es/indicators/{indicator_id}?start_date={fecha_inicio}&end_date={fecha_fin}"
                        response = requests.get(url, headers=headers)

                        if response.status_code == 200:
                            data = response.json()
                            with open(
                                json_filename, "w", encoding="utf-8"
                            ) as json_file:
                                json.dump(data, json_file, indent=4, ensure_ascii=False)
                            print(f"JSON guardado en {json_filename}")
                            datos_descargados[indicator_id] = data
                        else:
                            print(
                                f"Error al obtener el indicador {indicator_id}: {response.status_code}"
                            )
                            datos_descargados[indicator_id] = None

                return datos_descargados

            # ConfiguraciÃ³n y ejecuciÃ³n de la descarga
            API_KEY = "af850ff28e45f31c0345edbfadcea91ab243d7f45e1b875b4b7c67b3a9c41b2a"
            indicator_ids = [1001, 1739]
            fecha_inicio, fecha_fin = obtener_fechas_descarga()

            datos_api = descargar_datos_api(
                indicator_ids, fecha_inicio, fecha_fin, API_KEY
            )

        def p02xlsxhojaconvalorespordefecto():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # si es un nuevo archivo tex copia los parÃ¡metros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            codigo, titulo = comun.texfile.separar_y_mayusculizar(nombre)
            #
            ruta_archivo_excel = f"{zz.ruta_script}/{codigo}.xlsx"
            if 1:
                df = pd.DataFrame(zz.data, columns=["ParÃ¡metro", "Valor"])
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    # Crear un libro de trabajo de Excel
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja predeterminada creada por openpyxl
                    libro.remove(libro.active)

                    # Crear hojas con los nombres de los keys y escribir los datos
                    for sheet_name, sheet_data in zz.data.items():
                        hoja = libro.create_sheet(
                            title=sheet_name
                        )  # Crear la hoja con el nombre del key
                        for row in sheet_data:
                            hoja.append(row)  # Agregar cada fila a la hoja

                    # Guardar el archivo Excel
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    # print(ruta_archivo_excel)
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # Cargar todas las hojas del archivo Excel
            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )

            # Crear el diccionario data
            zz.data = {
                sheet_name: [df.columns.tolist()]
                + df.values.tolist()  # Incluir cabecera y datos
                for sheet_name, df in sheets.items()
            }


from fastapi.responses import StreamingResponse
from io import BytesIO
import os
import matplotlib.pyplot as plt
import pandas as pd
from types import SimpleNamespace
import subprocess
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata


import numpy as np
import sys

if __name__ == "__main__":
    zz = SimpleNamespace()

    # Copio la ruta del scripot para crear el excel
    zz.ruta_script = os.path.dirname(os.path.abspath(__file__))
    zz.nombre = os.path.basename(__file__)[:-3]
    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)

    #
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    os.chdir("../../../../")  # Cambia dos niveles arriba

    # clase especifico
    metodos = [
        m
        for m in dir(especifico.previo)
        if callable(getattr(especifico.previo, m)) and not m.startswith("__")
    ]
    metodos = [s for s in metodos if not s.startswith("__")]
    metodos = sorted(metodos)

    for metodo in metodos:
        metodoi = getattr(especifico.previo, metodo, None)
        metodoi()

    # la funcion de fastpai
    globals()[os.path.basename(__file__)[:-3]](zz.data)
