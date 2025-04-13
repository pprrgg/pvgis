from fastapi import APIRouter

router = APIRouter()


@router.post(f"/Cer1739")
def Cer1739PrecioDeLaEnergíaExcedentariaDelAutoconsumo(
    data: dict,
):
    data = comun.texfile.limpiar_valores(data)
    print(data)
    # parámetros
    pp = SimpleNamespace()
    # tablas y graficos en latex
    uu = SimpleNamespace()
    #
    vv = SimpleNamespace()
    xx = SimpleNamespace()
    yy = SimpleNamespace()
    # datos genericos
    zz = SimpleNamespace()

    zz.nombre = sys._getframe().f_code.co_name
    zz.ruta_script = os.path.dirname(
        os.path.abspath(__file__)
    )  # Obtener la carpeta sin el nombre del script

    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)
    zz.data = data

    output_buffer = comun.main.main(pp, uu, vv, xx, yy, zz)

    with open(f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf", "wb") as f:
        f.write(output_buffer.getvalue())

    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


def file_tex(pp, uu, vv, xx, yy, zz):

    return r"""

\newcommand{\xxMostrarVariablesAlFinal}{}
%\documentclass[a4paper,10pt]{article}
\documentclass[a4paper,10pt,twocolumn]{article}
\usepackage[top = 2cm, bottom = 2cm, left = 2cm, right = 2cm, asymmetric]{geometry} % Especificar los márgenes según la norma
\usepackage[spanish]{babel}    

\usepackage{graphicx}  % Para incluir imágenes si es necesario
\usepackage{amsmath, amssymb}  % Para expresiones matemáticas
\usepackage{fancyhdr}  % Para personalizar encabezados
\usepackage{chngcntr}  % Para cambiar la numeración de apartados y subsecciones
\usepackage{tocbibind}  % Para incluir bibliografía en la tabla de contenidos
\usepackage{appendix}  % Para el formato de anexos
\usepackage{lipsum}  % Para generar texto de relleno (dummy text)
\usepackage{geometry}  % Para personalizar los márgenes
\usepackage{multicol}  % Para columnas
\usepackage{titlesec}  % Para personalizar los títulos de las secciones
\usepackage[utf8]{inputenc}    
\usepackage[T1]{fontenc}       
\usepackage[pdftex,pdfencoding=auto, colorlinks=true, linkcolor=black, citecolor=black, filecolor=black, urlcolor=black]{hyperref}
\usepackage{tocloft}           
\usepackage{booktabs}          
\usepackage{float} % Paquete necesario para usar la opción [H]
\usepackage{array}
\usepackage{longtable}
\usepackage{circuitikz}
\usepackage{tikz}
\usepackage{tikz-cd}
\usepackage{tikz-qtree}
\usepackage{smartdiagram}

\usepackage{qrcode}            % Paquete para generar QR
\usepackage{tabularx} % Agregar en el preámbulo
\usepackage[absolute,overlay]{textpos} % Para posicionar objetos libremente
\usepackage{underscore}
\usepackage{pdfpages}

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{zz.titulo}  % Nombre del documento en el encabezado izquierdo
\fancyhead[C]{}  % Centro vacío
\fancyhead[R]{\thepage}  % Numeración de páginas en el encabezado derecho
% \renewcommand{\thepage}{\arabic{page}}  % Asegura la numeración de páginas en números arábigos

% Definición de los anexos
\newcommand{\annex}[1]{\section{Anexo #1} \addcontentsline{toc}{section}{Anexo #1}}
\addto\captionsspanish{%
  \renewcommand{\tablename}{Tabla}%
  \renewcommand{\listtablename}{Índice de tablas}%
}
\title{{\small Ref.:\uppercase{zz.codigo}}\\{\textbf{zz.titulo}}}
\author{
pp.ParticipantesP1NombreRaznSocial
}
\date{\today}







\begin{document}
\begin{Form}
%% --- PORTADA ---
%\begin{titlepage}
%    \centering
%    {\Large \textbf{TÍTULO DEL INFORME}}\[1.5cm]
%    %\includegraphics[width=0.4\textwidth]{logo.png}\[1cm] % Logo de la empresa/institución
%    
%    \textbf{Autor:} Nombre del Autor \[0.5cm]
%    \textbf{Fecha:} \today \[1cm]
%    \textbf{Código del documento:} REF-XXXX-YYYY \[1cm]
%    \textbf{Versión:} 1.0 \[1cm]
%    \textbf{Empresa / Institución:} Nombre \[2cm]
%    \vfill
%\end{titlepage}
	\maketitle
    %
	%\begin{abstract}

    %\end{abstract}
% --- ÍNDICE ---
	\tableofcontents  % Tabla de contenidos
	\listoffigures    % Lista de graficos
	\listoftables     % Lista de tablas
	\newpage



    

    













    




























































\begin{abstract}
Este informe analiza la evolución del Precio Voluntario para el Pequeño Consumidor (PVPC) durante los últimos 12 meses, examinando tendencias anuales, variaciones semanales y diferencias entre periodos tarifarios. Los datos proceden oficialmente del sistema de información de Red Eléctrica Española (ESIOS).
\end{abstract}


\section{Introducción}

El Real Decreto 244/2019 regula el mecanismo de compensación simplificada para excedentes de autoconsumo. Este informe compara:

\begin{itemize}
\item Precio PVPC horario
\item Valor de la energía excedentaria
\item Diferencias por periodos tarifarios
\end{itemize}

El PVPC es el mecanismo de fijación de precios para consumidores acogidos a la tarifa regulada. Este estudio utiliza datos oficiales publicados en el portal \href{https://www.esios.ree.es/es/pvpc}{ESIOS de REE} \cite{esios}, analizando:

\begin{itemize}
\item Evolución interanual
\item Patrones semanales
\item Diferencias entre periodos tarifarios (Punta, Llano, Valle)
\end{itemize}

\section{Metodología}
\subsection{Fuente de datos}
Todos los datos proceden de la API pública de \textbf{ESIOS (Red Eléctrica Española)} \cite{esios}, accediendo a:

\begin{itemize}
\item Series temporales horarias del PVPC
\item Metadatos de periodos tarifarios
\item Indicadores de mercado spot
\end{itemize}

\subsection{Procesamiento}
Los datos se procesaron mediante:

\begin{enumerate}
\item Extracción mediante API REST (últimos 365 días)
\item Clasificación por periodos tarifarios (RD 216/2014)
\item Cálculo de medias móviles (7 días)
\item Agregación temporal (hora → día → semana → mes)
\end{enumerate}

\section{Resultados}

\subsection{Evolución anual}
uu.d05PrecioDeLaEnergíaEléctricaExcedentariaAnual

\subsection{Evolución semanal}
uu.d15PrecioDeLaEnergíaEléctricaExcedentariaSemanal



\subsection{Análisis por periodos}
uu.tablad21PrecioEscedentesPorPeriodosTarifarios
uu.d21PrecioEscedentesPorPeriodosTarifarios

La Tabla \ref{tab:periodos} confirma la diferencia significativa entre periodos.


























\section{Precio Excedentes en horario soleado}

\subsection{Definición de Horarios}
Se clasifican las horas en dos categorías:

\begin{equation}
\text{H}(t) = 
\begin{cases} 
\text{Soleado} & 
\text{si } 
\\ (m_t \in [4,9] \cap h_t \in [8,20)) 
\\ 
\lor \\(m_t \notin [4,9] \cap h_t \in [10,17)) \\
\text{No Soleado} & \text{en otro caso}
\end{cases}
\end{equation}

donde:
\begin{itemize}
\item $m_t$: Mes (1=Enero, ..., 12=Diciembre)
\item $h_t$: Hora del día (0, ..., 23)
\end{itemize}

\subsection{Cálculo de Precios Medios}
Para cada categoría $s \in \{\text{Soleado}, \text{No Soleado}\}$:

\begin{equation}
P_s = \frac{1}{|T_s|}\sum_{t \in T_s} \text{PVPC}(t)
\end{equation}

donde:
\begin{itemize}
\item $T_s$: Conjunto de periodos en el horario $s$
\item $|T_s|$: Cardinalidad (número de horas)
\item PVPC$(t)$: Precio en el periodo $t$
\end{itemize}

\subsection{Resultados}

uu.tablad51PrecioExcedentesPorPeriodoSoleado


uu.d51PrecioExcedentesPorPeriodoSoleado



























































    







    

















% --- 7. REFERENCIAS BIBLIOGRÁFICAS ---
\begin{thebibliography}{99}
    \bibitem{ccee} \href{https://www.boe.es/buscar/pdf/2021/BOE-A-2021-7120-consolidado.pdf}{Resolución de 28 de abril de 2021, de la Dirección General de Política Energética y Minas, por la que se establece el contenido
mínimo y el modelo de factura de electricidad a utilizar por los comercializadores de referencia.}


\bibitem{rd244}
\href{}{RD 244/2019  sobre autoconsumo}

\bibitem{esios} 
\href{https://www.esios.ree.es/es/pvpc}{ESIOS - Red Eléctrica de España. PVPC y datos del sistema eléctrico
}


\bibitem{boe}
\href{}{Real Decreto 216/2014 por el que se establece la metodología de cálculo de los precios voluntarios para el pequeño consumidor.}


\end{thebibliography}





































\ifdefined\MostrarVariablesAlFinal
\newpage
\onecolumn
HojaDeDatos
\fi
\end{Form}
\end{document}

"""


class datos:

    class datos:

        def d00data2parametros(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Para cada hoja en el JSON
            # Convertir cada valor del diccionario en un DataFrame
            for key, value in zz.data.items():
                key_sanitized = comun.texfile.sanitize_name(
                    key
                )  # Limpiar el nombre de la clave
                df = pd.DataFrame(
                    value[1:], columns=value[0]
                )  # Primera fila como encabezado
                df.set_index(
                    df.columns[0], inplace=True
                )  # Usar la primera columna como índice

                setattr(xx, f"df{key_sanitized}", df)  # Guardar en ss con nombre limpio
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == zz.codigo:
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar índice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{idx_sanitized}",
                                val,
                            )
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == "Coordenadas" or key_sanitized == "Participantes":
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar índice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{key_sanitized}{col_sanitized}{idx_sanitized}",
                                val,
                            )

            return pp, uu, vv, xx, yy, zz

        def d00EjemplDeTabla(pp, uu, vv, xx, yy, zz):
            print(f"\n{100*'+'}{__class__.__name__}.{sys._getframe().f_code.co_name}")
            #
            df = pd.DataFrame(
                np.random.randint(0, 100, size=(2, 3)), columns=["A", "B", "C"]
            )
            # siempre tieen que ir el indice que es donde meto separados por _ indice_ejeX_ejeY
            df = df.reset_index(drop=True)  # drop=True elimina el índice anterior
            # df = df.set_index(df.columns[0])
            df.index.name = "EUR/MWh_hora semanal"
            #
            setattr(xx, sys._getframe().f_code.co_name, df)
            return pp, uu, vv, xx, yy, zz

        def d00TodosLosExcelDeLaCarpetaXlsx(pp, uu, vv, xx, yy, zz):
            """para los datos que no hay ques esta descargancodon continuamente por ejm los pvpc..."""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpetaimagenes = Path(f"{zz.ruta_script}/assets")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.xlsx")]

            for imagen in imagenes:
                # _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                titulo = imagen[:-4]
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional

                ruta_archivo_excel = Path(f"{zz.ruta_script}/assets/{imagen}")
                # libro = openpyxl.load_workbook(ruta_archivo_excel)

                # Cargar todas las hojas del archivo Excel
                sheets = pd.read_excel(
                    ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
                )
                for nombre_hoja, df in sheets.items():
                    nombre_hoja = comun.texfile.normalizar_cadena(nombre_hoja)
                    setattr(xx, f"{nombre_hoja}", df)  # Guardar en ss con nombre limpio

            return pp, uu, vv, xx, yy, zz

        def d51PrecioExcedentesPorPeriodoSoleado(pp, uu, vv, xx, yy, zz):
            print(f"\n{100*'+'}{__class__.__name__}.{sys._getframe().f_code.co_name}")
            df = xx.d05PrecioDeLaEnergíaEléctricaExcedentariaAnual.copy()

            # Convertir a datetime y extraer componentes
            df["datetime"] = pd.to_datetime(df.index, utc=True)
            df["hour"] = df["datetime"].dt.hour
            df["month"] = df["datetime"].dt.month

            # Definir horas soleadas (función simplificada)
            def es_horario_soleado(row):
                if (row["month"] in [4, 5, 6, 7, 8, 9] and 8 <= row["hour"] < 20) or (
                    row["month"] in [1, 2, 3, 10, 11, 12] and 10 <= row["hour"] < 17
                ):
                    return True
                return False

            # Aplicar clasificación
            df["horario_soleado"] = df.apply(es_horario_soleado, axis=1)

            # Calcular medias simples
            precio_soleado = (
                df[df["horario_soleado"]]["PVPC Excedentaria"].mean().round(2)
            )
            precio_no_soleado = (
                df[~df["horario_soleado"]]["PVPC Excedentaria"].mean().round(2)
            )

            df_resultado = pd.DataFrame(
                {
                    "Horario": ["Soleado", "No Soleado"],
                    "Precio_Medio": [precio_soleado, precio_no_soleado],
                }
            ).set_index("Horario")

            df_resultado.index.name = "EUR/MWh_Periodo"

            # Guardar en la instancia xx como atributo dinámico
            setattr(xx, sys._getframe().f_code.co_name, df_resultado)
            return pp, uu, vv, xx, yy, zz

        def d21PrecioEscedentesPorPeriodosTarifarios(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            df = xx.d05PrecioDeLaEnergíaEléctricaExcedentariaAnual.copy()

            # Suponiendo que ya tienes:
            df["datetime"] = pd.to_datetime(df.index, utc=True)
            df["hour"] = df["datetime"].dt.hour
            df["week"] = df["datetime"].dt.isocalendar().week
            df["year"] = df["datetime"].dt.isocalendar().year
            df["day_of_week"] = df["datetime"].dt.dayofweek  # 0 = lunes, 6 = domingo
            df["month"] = df["datetime"].dt.month

            # 1. Definir temporada (Alta = verano, Baja = invierno)
            df["temporada"] = "Baja"  # Por defecto: octubre a mayo
            df.loc[df["month"].isin([6, 7, 8, 9]), "temporada"] = (
                "Alta"  # Junio a septiembre
            )
            df["es_laborable"] = df["day_of_week"] < 5  # Lunes a Viernes


            # 2. Asignar periodos para la tarifa 2.0 TO (3 periodos)
            def asignar_periodo_20TD(row):
                if not row["es_laborable"]:  # Fin de semana o festivo (siempre Valle)
                    return 3
                
                # Día laborable (L-V)
                if 10 <= row["hour"] < 14 or 18 <= row["hour"] < 22:  # Punta (P1)
                    return 1
                elif (8 <= row["hour"] < 10 or 14 <= row["hour"] < 18 or 22 <= row["hour"] < 24):  # Llano (P2)
                    return 2
                else:  # Valle (P3) - 00:00-08:00
                    return 3

            df["periodo_20TD"] = df.apply(asignar_periodo_20TD, axis=1)





            # 2. Asignar periodos para la tarifa 2.0TD (3 periodos)
            def asignar_periodo_20TDantiguo(row):
                if row["day_of_week"] >= 5:  # Fin de semana (siempre Valle)
                    return 3
                if row["temporada"] == "Alta":  # Verano
                    if (11 <= row["hour"] < 15) or (
                        18 <= row["hour"] < 22
                    ):  # Punta verano
                        return 1
                    elif (
                        (8 <= row["hour"] < 11)
                        or (15 <= row["hour"] < 18)
                        or (22 <= row["hour"] < 24)
                    ):  # Llano verano
                        return 2
                else:  # Invierno
                    if (10 <= row["hour"] < 14) or (
                        18 <= row["hour"] < 22
                    ):  # Punta invierno
                        return 1
                    elif (
                        (8 <= row["hour"] < 10)
                        or (14 <= row["hour"] < 18)
                        or (22 <= row["hour"] < 24)
                    ):  # Llano invierno
                        return 2
                return 3  # Valle (00:00-08:00)

            df["periodo_20TD"] = df.apply(asignar_periodo_20TD, axis=1)

            # 3. Asignar periodos para la tarifa 3.0TD o 6.0TD (6 periodos)
            def asignar_periodo_30TD(row):
                if row["day_of_week"] >= 5:  # Fin de semana (siempre P3 o P6)
                    return 6 if row["temporada"] == "Baja" else 3
                if row["temporada"] == "Alta":  # Verano (P1, P2, P3)
                    if (11 <= row["hour"] < 15) or (
                        18 <= row["hour"] < 22
                    ):  # P1 (Punta verano)
                        return 1
                    elif (
                        (8 <= row["hour"] < 11)
                        or (15 <= row["hour"] < 18)
                        or (22 <= row["hour"] < 24)
                    ):  # P2 (Llano verano)
                        return 2
                    else:  # P3 (Valle verano: 00:00-08:00)
                        return 3
                else:  # Invierno (P4, P5, P6)
                    if (10 <= row["hour"] < 14) or (
                        18 <= row["hour"] < 22
                    ):  # P4 (Punta invierno)
                        return 4
                    elif (
                        (8 <= row["hour"] < 10)
                        or (14 <= row["hour"] < 18)
                        or (22 <= row["hour"] < 24)
                    ):  # P5 (Llano invierno)
                        return 5
                    else:  # P6 (Valle invierno: 00:00-08:00)
                        return 6

            df["periodo_30TD"] = df.apply(asignar_periodo_30TD, axis=1)

            #
            #
            #
            #
            #
            #

            # Calcular media del PVPC por periodo_20TD
            media_por_periodo = (
                df.groupby("periodo_20TD")["PVPC Excedentaria"]
                .mean()
                .round(2)  # Redondear a 2 decimales
                .reset_index()
            )

            # Media global (todos los periodos)
            media_global = df["PVPC Excedentaria"].mean().round(2)

            # Crear JSON
            resultado_json = {
                "ultimos_12_meses": {
                    "media_precio_por_periodo": media_por_periodo.to_dict(
                        orient="records"
                    ),
                    "media_global_precio": media_global,
                }
            }

            # Crear DataFrame combinando ambos datos
            df = pd.DataFrame(
                {
                    "periodo": [1, 2, 3, "Global"],
                    "Precio_medio": [
                        resultado_json["ultimos_12_meses"]["media_precio_por_periodo"][
                            0
                        ]["PVPC Excedentaria"],
                        resultado_json["ultimos_12_meses"]["media_precio_por_periodo"][
                            1
                        ]["PVPC Excedentaria"],
                        resultado_json["ultimos_12_meses"]["media_precio_por_periodo"][
                            2
                        ]["PVPC Excedentaria"],
                        resultado_json["ultimos_12_meses"]["media_global_precio"],
                    ],
                    "tipo": ["Punta", "Llano", "Valle", "Global"],
                }
            )

            df = df.set_index("periodo").drop(columns="tipo")
            df.index.name = "EUR/MWh_periodo"

            # Guardar en la instancia xx como atributo dinámico
            setattr(xx, sys._getframe().f_code.co_name, df)  # Valores originales
            return pp, uu, vv, xx, yy, zz

        def d05PrecioDeLaEnergíaEléctricaExcedentariaAnual(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{__class__.__name__}.{sys._getframe().f_code.co_name}")
            """Procesa los datos descargados y los convierte en DataFrames"""

            indicator_id = "1739"  # ID del indicador

            ruta = Path(f"{zz.ruta_script}/assets")  # Carpeta donde buscar

            # Buscar archivo con el patrón "indicador_1001_*.json"
            archivo_json = next(ruta.glob(f"indicador_{indicator_id}_*.json"), None)

            if archivo_json:
                with open(archivo_json, "r", encoding="utf-8") as f:
                    data = json.load(f)

            # Extraer valores y convertir a DataFrame
            valores = data["indicator"]["values"]
            df = pd.DataFrame(valores)

            # Filtrar solo geo_name "Península" o "España"
            df = df[df["geo_name"].isin(["Península", "España"])]

            # Convertir la columna datetime a formato datetime con utc=True
            df["datetime"] = pd.to_datetime(df["datetime"], utc=True)

            # Ordenar por fecha
            df = df.sort_values("datetime")

            # Mantener solo dos columnas: datetime (índice) y Value
            df = (
                df[["datetime", "value"]]
                .rename(columns={"value": "PVPC Excedentaria"})
                .set_index("datetime")
            )
            df["PVPC Excedentaria medio semanal"] = (
                df["PVPC Excedentaria"].rolling(window=24 * 7).mean()
            )

            # Asignar nombre al índice y a las columnas
            df.index.name = "EUR/MWh_Fecha"

            # Guardar en la instancia xx como atributo dinámico
            setattr(xx, sys._getframe().f_code.co_name, df)  # Valores originales
            return pp, uu, vv, xx, yy, zz

        def d15PrecioDeLaEnergíaEléctricaExcedentariaSemanal(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            """Procesa los datos descargados y los convierte en DataFrames"""

            df = xx.d05PrecioDeLaEnergíaEléctricaExcedentariaAnual.copy()

            # Convertir a datetime y extraer componentes
            df["datetime"] = pd.to_datetime(df.index, utc=True)
            df["hour"] = df["datetime"].dt.hour
            df["week"] = df["datetime"].dt.isocalendar().week
            df["year"] = df["datetime"].dt.isocalendar().year
            df["day_of_week"] = df["datetime"].dt.dayofweek  # 0 = lunes, 6 = domingo

            # Calcular la media de 'PVPC Excedentaria' por hora y día de semana
            media_pvpc = (
                df.groupby(["day_of_week", "hour"])["PVPC Excedentaria"]
                .mean()
                .reset_index()
            )

            # Calcular la hora semanal (0 a 167)
            media_pvpc["hora_semanal"] = (
                media_pvpc["day_of_week"] * 24 + media_pvpc["hour"]
            )

            # Crear el DataFrame final ordenado por hora semanal
            resultado = (
                media_pvpc[["hora_semanal", "PVPC Excedentaria"]]
                .sort_values("hora_semanal")
                .set_index("hora_semanal")
                .rename(columns={"PVPC Excedentaria": "PVPC Excedentaria"})
            )

            resultado.index.name = "EUR/MWh_hora semanal"

            # Guardar en la instancia xx como atributo dinámico
            setattr(xx, sys._getframe().f_code.co_name, resultado)  # Valores originales
            return pp, uu, vv, xx, yy, zz


########################################################################################################
########################################################################################################


class comun:
    class main:

        def main(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # segun si es un formulario o un documento nuevo

            pp, uu, vv, xx, yy, zz = comun.main.datos(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.datos2latex(pp, uu, vv, xx, yy, zz)

            output_buffer = comun.main.replaceLatex(pp, uu, vv, xx, yy, zz)

            # print(uu)
            return output_buffer

        def datos(pp, uu, vv, xx, yy, zz):
            """datoes"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listadatoes = [
                m
                for m in dir(datos.datos)
                if callable(getattr(datos.datos, m)) and not m.startswith("__")
            ]
            listadatoes = [s for s in listadatoes if not s.startswith("__")]
            listadatoes = sorted(listadatoes)

            for dato in listadatoes:
                print(dato)
                metodo = getattr(datos.datos, dato, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def datos2latex(pp, uu, vv, xx, yy, zz):
            """tablas"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listatablas = [
                m
                for m in dir(comun.datos2latex)
                if callable(getattr(comun.datos2latex, m)) and not m.startswith("__")
            ]
            listatablas = [s for s in listatablas if not s.startswith("__")]
            for tabla in listatablas:
                # print(tabla)
                metodo = getattr(
                    comun.datos2latex, tabla, None
                )  # Obtener el método o None si no existe
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)

            return pp, uu, vv, xx, yy, zz

        def replaceLatex(pp, uu, vv, xx, yy, zz):
            """
            Compila un archivo .tex a PDF, reemplaza un texto y devuelve el PDF en memoria.
            :param ruta_tex: Ruta al archivo .tex original.
            :return: Objeto BytesIO con el contenido del PDF.
            """

            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            pdf_path = Path(zz.nombre).with_suffix(".pdf")
            ruta_tex = f"{os.path.splitext(pdf_path)[0]}.tex"

            nombre_archivo = os.path.basename(ruta_tex)  # Nombre del archivo sin ruta

            # Crear un directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                ruta_tex_temp = os.path.join(temp_dir, nombre_archivo)

                contenido = file_tex(pp, uu, vv, xx, yy, zz)
                # tabla final con todas los parámetros  y variables
                HojaDeDatos = "\n"

                for var_name, value in pp.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\npp. {var_name}\n\npp.{var_name}"

                for var_name, value in uu.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nuu. {var_name}\n\nuu.{var_name}"

                for var_name, value in vv.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nvv. {var_name}\n\nvv.{var_name}"

                for var_name, value in xx.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nxx. {var_name}\n\nxx.{var_name}"

                for var_name, value in yy.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nyy. {var_name}\n\nyy.{var_name}"

                # noooooooo,!!!!! que el json del data da error en latex
                # for var_name, value in zz.__dict__.items():
                #     print(f"zz.{var_name}")
                #     HojaDeDatos = HojaDeDatos + f"\n\nzz. {var_name}\n\nzz.{var_name}"

                contenido = contenido.replace(f"HojaDeDatos", HojaDeDatos)

                if 1:
                    # Reemplazar directamente todas las variables en el texto
                    for var_name, value in pp.__dict__.items():
                        # print(f"pp.{var_name}")
                        contenido = contenido.replace(
                            f"pp.{var_name}",
                            f"{str(value)}",
                        )
                    for var_name, value in uu.__dict__.items():
                        # print(f"uu.{var_name}")
                        contenido = contenido.replace(f"uu.{var_name}", f"{str(value)}")

                    for var_name, value in vv.__dict__.items():
                        # print(f"vv.{var_name}")
                        contenido = contenido.replace(f"vv.{var_name}", f"{str(value)}")

                    for var_name, value in xx.__dict__.items():
                        # print(f"xx.{var_name}")
                        contenido = contenido.replace(f"xx.{var_name}", f"{str(value)}")

                    for var_name, value in yy.__dict__.items():
                        # print(f"yy.{var_name}")
                        contenido = contenido.replace(f"yy.{var_name}", f"{str(value)}")

                    for var_name, value in zz.__dict__.items():
                        # print(f"zz.{var_name}")
                        contenido = contenido.replace(f"zz.{var_name}", f"{str(value)}")

                with open(
                    "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.tex", "w", encoding="utf-8"
                ) as file:
                    file.write(contenido)

                # Guardar los cambios en el archivo temporal
                with open(ruta_tex_temp, "w", encoding="utf-8") as file:
                    file.write(contenido)

                # Función para ejecutar pdflatex dos veces
                def ejecutar_pdflatex():
                    print(
                        f"\n{100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                    )
                    if 1:
                        subprocess.run(
                            ["pdflatex", "-output-directory", temp_dir, ruta_tex_temp],
                            check=True,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )

                    return True

                # # para que se cre el indice debe estar descomentado
                # Ejecutar pdflatex dos veces para generar la TOC correctamente
                if not (ejecutar_pdflatex() and ejecutar_pdflatex()):
                    return None  # Si la compilación falla, se detiene el proceso
                # ejecutar_pdflatex()

                # Ruta del PDF generado
                pdf_generado = os.path.join(
                    temp_dir, nombre_archivo.replace(".tex", ".pdf")
                )

                # Leer el PDF en memoria
                with open(pdf_generado, "rb") as file:
                    pdf_en_memoria = BytesIO(file.read())

            return pdf_en_memoria

            #######################

    class datos2latex:

        def tabla(pp, uu, vv, xx, yy, zz):
            # tabla de todos los df

            for var_name, value in {**xx.__dict__, **yy.__dict__}.items():
                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                if var_name == zz.codigo:
                    titulo = "Parámetros"

                if isinstance(value, pd.DataFrame):
                    # para no hacer tablas de series muy
                    if len(value) > 10:
                        continue
                    # Redondear los valores del DataFrame a 1 decimal
                    value = value.round(1)

                    latex_code = value.to_latex(float_format="%.1f")

                    latex_code_with_resizebox = rf"""
                    \begin{{table}}[H] \centering
                        % \resizebox{{0.4\textwidth}}{{!}}
                        {{
                        {latex_code}
                        }}
                        \caption{{{titulo}}}
                    \end{{table}}
                    """

                    # Asignar el LaTeX a la propiedad correspondiente de uu
                    setattr(uu, f"tabla{var_name}", latex_code_with_resizebox)

            # Convierte todos los df de uu a tablas
            return pp, uu, vv, xx, yy, zz

        def grafico(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # tabla de todos los df

            for var_name, value in {**xx.__dict__, **yy.__dict__}.items():
                print(var_name)
                if hasattr(xx, var_name):
                    df = getattr(xx, var_name)
                else:
                    df = getattr(yy, var_name)

                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                # 1. Crear un archivo temporal para la imagen
                with tempfile.NamedTemporaryFile(
                    suffix=".png", delete=False
                ) as temp_image:
                    temp_image_path2 = (
                        temp_image.name
                    )  # Guardar la ruta del archivo temporal

                # 📌 Gráfico del flujo de caja y acumulado
                if len(df) < 24:
                    fig, ax = plt.subplots(figsize=(5, 3))
                    try:
                        df.plot(kind="bar", ax=ax)

                        plt.xticks(rotation=45)
                        plt.tight_layout()  # Ajustar para que no se corten las etiquetas
                    except:
                        """no valores numericos para plt"""

                    try:
                        ax.set_ylabel(df.index.name.split("_")[0])
                        ax.set_xlabel(df.index.name.split("_")[1])
                    except:
                        """"""
                else:

                    fig, ax = plt.subplots(figsize=(6, 5))

                    df.plot(ax=ax)

                    # Obtener el nombre de la columna automáticamente
                    try:
                        ax.set_ylabel(df.index.name.split("_")[0])
                        ax.set_xlabel(df.index.name.split("_")[1])
                    except:
                        """"""

                plt.grid(axis="y", linestyle="--", alpha=0.7)
                plt.savefig(temp_image_path2, format="png", dpi=300)
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
                latex_code = (
                    ""
                    + r"""\begin{figure}[H] \centering
                    """
                    + rf"""
                    \includegraphics[width=.5\textwidth]{{{temp_image_path2}}}
                    \caption{{{titulo}}}
                    """
                    + """
                    \label{fig:dfassssdfsa}
                    \end{figure}
                    """
                )
                setattr(uu, f"{var_name}", latex_code)
            return pp, uu, vv, xx, yy, zz

        def imagen(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpetaimagenes = Path(f"{zz.ruta_script}/{zz.codigo}/img")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.png")] + [
                f.name for f in carpetaimagenes.glob("*.jpg")
            ]

            for imagen in imagenes:
                # _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                titulo = imagen[:-4]
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional

                latex_code = (
                    ""
                    + r"""\begin{figure}[H] \centering
                    """
                    + rf"""
                    \includegraphics[width=0.7\textwidth]{{{Path(f"{zz.ruta_script}/{zz.codigo}/img/{imagen}")}}} 
                    \caption{{{titulo}}}
                    """
                    + """
                    \label{fig:dfasdfsa}
                    \end{figure}
                    """
                )

                setattr(uu, f"imagen{imagen}", latex_code)
            return pp, uu, vv, xx, yy, zz

    class texfile:

        def normalizar_cadena(texto):
            if isinstance(texto, str):  # Asegurar que es un string
                texto = texto.lower()
                texto = "".join(
                    c
                    for c in unicodedata.normalize("NFD", texto)
                    if unicodedata.category(c) != "Mn"
                )
                # Capitalizar la primera letra después de un espacio
                texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
                # Eliminar espacios
                texto = re.sub(r"\s+", "", texto)
                # Mantener solo caracteres permitidos
                texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
            return texto

        def limpiar_valores(data):
            for key, values in data.items():
                for i in range(
                    1, len(values)
                ):  # Saltamos la primera fila (encabezados)
                    for j in range(
                        1, len(values[i])
                    ):  # Recorremos todas las columnas (excepto encabezados)
                        if isinstance(
                            values[i][j], str
                        ):  # Verificar si es una cadena antes de dividir
                            valor_limpio = (
                                values[i][j].split(";")[0].strip()
                            )  # Tomar solo el primer valor y limpiar espacios

                            # Intentar convertir a flotante si es posible
                            try:
                                values[i][j] = float(valor_limpio)
                            except ValueError:
                                values[i][
                                    j
                                ] = valor_limpio  # Si no se puede convertir, dejar como está
            return data

        def separar_y_mayusculizar(texto):
            # Usamos una expresión regular para separar las mayúsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            # Separar la primera parte (mantenerla como está)
            primera_parte = partes[0]

            # El resto de las partes lo convertimos a mayúsculas y las unimos con espacios
            resto = " ".join([parte.upper() for parte in partes[1:]])
            # resto = " ".join([parte for parte in partes[1:]])

            return primera_parte, resto

        def separar_y_formatear(texto):
            # Usamos una expresión regular para separar las mayúsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            if not partes:
                return ""

            # Convertimos todo a minúsculas y capitalizamos solo la primera letra
            resultado = " ".join([parte.lower() for parte in partes])
            return resultado.capitalize()

        # Función para limpiar nombres de variables
        def sanitize_name(name):
            return re.sub(
                r"[^a-zA-Z0-9]", "", str(name)
            )  # Elimina cualquier carácter no alfanumérico


class especifico:
    class previo:
        def p00parametros():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            nombre_funcion = os.path.basename(__file__)[:-3]
            codigo, _ = comun.texfile.separar_y_mayusculizar(nombre_funcion)
            nan = " "
            print(codigo)
            zz.data = {
                zz.codigo: [
                    ["Parametro", "Valor"],
                    ["Inversión inicial [€]", 0],
                    ["Bonificación  inicial [€]", 10000],
                    ["Duración de la actuación [años]", 15],
                    ["Préstamo recibido [€]", 100000],
                    ["Tasa de interés del préstamo [\%]", 5],
                    ["Plazo del préstamo [años]", 15],
                    ["Inflación annual [\%]", 3],
                    ["Ahorro annual [€]", 25000],
                    ["Costos fijos operativos [€]", 1000],
                    ["Tasa de descuento [\%]", 6],
                ],
                "Coordenadas": [
                    ["Parámetro", "Valor"],
                    ["lat", 36.6641681434112],
                    ["lng", -4.45863940137516],
                ],
                "Participantes": [
                    ["Parámetro", "P1"],
                    ["Rol", ""],
                    ["DNI/NIF/NIE", ""],
                    [
                        "Nombre/Razón Social",
                        "'EDITAR' para poner tu nombre y ajustar los parámetros.",
                    ],
                    ["Titulación", ""],
                ],
            }

        def p01carpetaauxiliar():
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpeta = Path(f"{zz.ruta_script}/assets")

            if not os.path.exists(carpeta):
                os.mkdir(carpeta)

        def p03apipvpcesios():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import requests
            import json
            import os
            from datetime import datetime, timedelta

            def obtener_fechas_descarga():
                """Calcula las fechas de inicio y fin para los últimos 12 meses completos"""
                hoy = datetime.today()
                primer_dia_mes_actual = datetime(hoy.year, hoy.month, 1)
                ultimo_dia_mes_anterior = primer_dia_mes_actual - timedelta(days=1)
                primer_dia_mes_anterior = datetime(
                    ultimo_dia_mes_anterior.year, ultimo_dia_mes_anterior.month, 1
                )

                primer_dia_12_meses_atras = primer_dia_mes_anterior - timedelta(
                    days=365
                )
                fecha_inicio = primer_dia_12_meses_atras.strftime("%Y-%m-%dT00:00:00")
                fecha_fin = ultimo_dia_mes_anterior.strftime("%Y-%m-%dT23:59:59")

                return fecha_inicio, fecha_fin

            def descargar_datos_api(indicator_ids, fecha_inicio, fecha_fin, api_key):
                """Descarga datos de la API y los guarda en archivos JSON"""
                headers = {
                    "Accept": "application/json; application/vnd.esios-api-v1+json",
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                }

                datos_descargados = {}

                for indicator_id in indicator_ids:

                    json_filename = f"{zz.ruta_script}/assets/indicador_{indicator_id}_{fecha_inicio[:7]}_a_{fecha_fin[:7]}.json"

                    if os.path.exists(json_filename):
                        print(
                            f"El archivo para el indicador {indicator_id} ya existe: {json_filename}. No se descargará nuevamente."
                        )
                        with open(json_filename, "r", encoding="utf-8") as json_file:
                            datos_descargados[indicator_id] = json.load(json_file)
                    else:
                        print(f"Descargando datos para el indicador {indicator_id}...")
                        url = f"https://api.esios.ree.es/indicators/{indicator_id}?start_date={fecha_inicio}&end_date={fecha_fin}"
                        response = requests.get(url, headers=headers)

                        if response.status_code == 200:
                            data = response.json()
                            with open(
                                json_filename, "w", encoding="utf-8"
                            ) as json_file:
                                json.dump(data, json_file, indent=4, ensure_ascii=False)
                            print(f"JSON guardado en {json_filename}")
                            datos_descargados[indicator_id] = data
                        else:
                            print(
                                f"Error al obtener el indicador {indicator_id}: {response.status_code}"
                            )
                            datos_descargados[indicator_id] = None

                return datos_descargados

            # Configuración y ejecución de la descarga
            API_KEY = "af850ff28e45f31c0345edbfadcea91ab243d7f45e1b875b4b7c67b3a9c41b2a"
            indicator_ids = [1001, 1739]
            fecha_inicio, fecha_fin = obtener_fechas_descarga()

            datos_api = descargar_datos_api(
                indicator_ids, fecha_inicio, fecha_fin, API_KEY
            )

        def p02xlsxhojaconvalorespordefecto():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # si es un nuevo archivo tex copia los parámetros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            codigo, titulo = comun.texfile.separar_y_mayusculizar(nombre)
            #
            ruta_archivo_excel = f"{zz.ruta_script}/{codigo}.xlsx"
            if 1:
                df = pd.DataFrame(zz.data, columns=["Parámetro", "Valor"])
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    # Crear un libro de trabajo de Excel
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja predeterminada creada por openpyxl
                    libro.remove(libro.active)

                    # Crear hojas con los nombres de los keys y escribir los datos
                    for sheet_name, sheet_data in zz.data.items():
                        hoja = libro.create_sheet(
                            title=sheet_name
                        )  # Crear la hoja con el nombre del key
                        for row in sheet_data:
                            hoja.append(row)  # Agregar cada fila a la hoja

                    # Guardar el archivo Excel
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    # print(ruta_archivo_excel)
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # Cargar todas las hojas del archivo Excel
            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )

            # Crear el diccionario data
            zz.data = {
                sheet_name: [df.columns.tolist()]
                + df.values.tolist()  # Incluir cabecera y datos
                for sheet_name, df in sheets.items()
            }


from fastapi.responses import StreamingResponse
from io import BytesIO
import os
import matplotlib.pyplot as plt
import pandas as pd
from types import SimpleNamespace
import subprocess
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata


import numpy as np
import sys

if __name__ == "__main__":
    zz = SimpleNamespace()

    # Copio la ruta del scripot para crear el excel
    zz.ruta_script = os.path.dirname(os.path.abspath(__file__))
    zz.nombre = os.path.basename(__file__)[:-3]
    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)

    #
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    os.chdir("../../../../")  # Cambia dos niveles arriba

    # clase especifico
    metodos = [
        m
        for m in dir(especifico.previo)
        if callable(getattr(especifico.previo, m)) and not m.startswith("__")
    ]
    metodos = [s for s in metodos if not s.startswith("__")]
    metodos = sorted(metodos)

    for metodo in metodos:
        metodoi = getattr(especifico.previo, metodo, None)
        metodoi()

    # la funcion de fastpai
    globals()[os.path.basename(__file__)[:-3]](zz.data)
