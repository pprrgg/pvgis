
#!/usr/bin/env python3


class calculos:

    class calculos:
        def c01parametros(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            for param, value in ss.data[ss.codigo]:
                param = normalizar_cadena(param)
                print(f"ss.{param} = {value}")

                # Intentar convertir value a float, si no es posible, dejarlo igual
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass  # Si no se puede convertir, mantiene el valor original

                setattr(ss, param, value)

            for param, value in ss.data[ss.codigo][
                1:
            ]:  # Empezamos desde el índice 1 para omitir los encabezados

                # Intentamos convertir a float si es posible
                try:
                    float_value = float(value)
                    setattr(ss, param, float_value)
                    globals()[param] = getattr(ss, param)

                except (ValueError, TypeError):
                    # Si no es convertible a float, lo dejamos como está
                    pass
            return ss

        def c02calculoejemplo(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Crear un DataFrame con números aleatorios
            np.random.seed(42)  # Para reproducibilidad
            df = pd.DataFrame(np.random.rand(4, 4))

            # Crear la fila de total con "-" excepto en la última columna
            total_row = ["TOTAL"] + ["-"] * (df.shape[1] - 2) + [df.iloc[:, -1].sum()]

            # Agregar la fila de total
            df.loc["gsdfg"] = total_row

            df = df.copy()

            return ss


class zz:
    class main:

        def main(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            if 1:
                # Leer el archivo PDF desde el disco
                ss = zz.main.parametros_formulario(ss)

                '''calculos'''
                listacalculos = [
                    m
                    for m in dir(calculos.calculos)
                    if callable(getattr(calculos.calculos, m))
                    and not m.startswith("__")
                ]
                listacalculos = [s for s in listacalculos if not s.startswith("xx")]
                for calculo in sorted(listacalculos):
                    metodo = getattr(calculos.calculos, calculo, None)
                    ss = metodo(ss)
                ""
                try:
                    "dependiendo si lo ejecuto en local o en produccion"
                    with open(ss.path, "rb") as file:
                        pdf_content = file.read()
                except:
                    with open(f"{ss.nombre}.pdf", "rb") as file:
                        pdf_content = file.read()

                output_buffer = BytesIO(pdf_content)
                output_buffer.seek(0)

            return output_buffer

            #######################

        def parametros_formulario(ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # excel a ss sapacio de estado
            for param, value in ss.data[ss.codigo]:
                param = normalizar_cadena(param)
                # Intentar convertir value a float, si no es posible, dejarlo igual
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass  # Si no se puede convertir, mantiene el valor original
                setattr(ss, param, value)
            return ss

        def xlsxhojaconvalorespordefecto(ss):
            ''' '''
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # si es un nuevo archivo tex copia los parametros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            codigo, titulo = separar_y_mayusculizar(nombre)
            ruta_archivo_excel = f"./{codigo}.xlsx"
            if 1:
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    print("!" * 88 + f"creando el nuevo {ruta_archivo_excel}")
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja por defecto (llamada 'Sheet')
                    hoja_por_defecto = libro["Sheet"]
                    libro.remove(hoja_por_defecto)

                    hoja = libro.create_sheet(title=codigo)

                    # Seleccionar la hoja activa
                    # hoja = libro.active
                    # Guardar el nuevo libro de trabajo
                    libro.save(ruta_archivo_excel)

                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    # guarda los parametros en la hoja de calculo
                    hoja = libro[codigo]

                    # Borrar todo el contenido de la hoja
                    for fila in hoja.iter_rows():
                        for celda in fila:
                            celda.value = None

                    # Insertar los datos del DataFrame en la hoja de Excel
                    # Insertar la cabecera
                    hoja["A1"] = "Parámetro"
                    hoja["B1"] = "Valor"

                    # Insertar los datos del DataFrame en la hoja de Excel
                    for i, (parametro, valor) in enumerate(
                        zip(df["Descripción"], df["valor"]), start=2
                    ):
                        # hoja[f"A{i}"] = parametro.replace("_", "").replace(" ", "")
                        hoja[f"A{i}"] = parametro
                        hoja[f"B{i}"] = valor

                    # Guardar el archivo con los cambios
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # leerlo y devolverlo como json, igual que luego lo enfiara react
            # Cargar el archivo Excel

            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )  # Carga todas las hojas

            # Función para transformar el formato

            # Convertir cada hoja al formato requerido
            data = {
                # sheet_name: [["Parámetro", "Valor"]] + df.values.tolist()
                sheet_name: df.values.tolist()
                for sheet_name, df in sheets.items()
            }
            return data


from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from PyPDF2 import PdfReader, PdfWriter
from io import BytesIO
import os
import matplotlib.pyplot as plt
import matplotlib

matplotlib.use("Agg")

import io
import base64
import pandas as pd
import fitz  # PyMuPDF
from types import SimpleNamespace
import subprocess
import shutil
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata

import numpy as np
import sys




class formulario:
    class formulario:
        def rellenoFormulario(pdf_buffer, ss):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            pdf_document = fitz.open(stream=pdf_buffer, filetype="pdf")
            output_buffer = BytesIO()
            variables = vars(ss)
            # print("CAMPOS DEL FORMULARIO")
            # Recorrer todas las páginas del PDF
            for page_num in range(len(pdf_document)):
                page = pdf_document.load_page(page_num)
                # Iterar sobre los widgets (campos de formulario)
                for widget in page.widgets():
                    field_name = widget.field_name  # Obtener el nombre del campo
                    field_name = normalizar_cadena(field_name)
                    # print(field_name)
                    if (
                        field_name in variables
                    ):  # Si el nombre del campo está en el SimpleNamespace
                        # variables[field_name]=round(variables[field_name],1)
                        widget.field_value = str(
                            variables[field_name]
                        )  # Asignar el valor
                        try:
                            widget.update()  # Actualizar el widget
                        except Exception as e:
                            ''''''
                            print(f"Error al actualizar el campo {field_name}: {e}")
            pdf_document.save(output_buffer)
            pdf_document.close()
            output_buffer.seek(0)
            return output_buffer



        def leer_los_campos_del_pdf_formulario():
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            import os
            import pandas as pd
            import fitz  # PyMuPDF

            def extract_pdf_fields(pdf_path):
                try:
                    doc = fitz.open(pdf_path)
                    fields = set()
                    for page in doc:
                        for widget in page.widgets():
                            if widget.field_name:
                                fields.add(widget.field_name)
                    return list(fields)
                except Exception as e:
                    print(f"Error al leer {pdf_path}: {e}")
                    return []

            def get_pdf_fields_as_df(pdf_path):
                print(
                    f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                )
                # Extrae los campos del PDF
                fields = extract_pdf_fields(pdf_path)

                # Si hay campos extraídos, crea el DataFrame con 'Parametro' y 'Valor'
                if fields:
                    df = pd.DataFrame(
                        {"Descripción": fields, "valor": [0] * len(fields)}
                    )
                else:
                    df = pd.DataFrame(columns=["Descripción", "valor"])

                return df

            # Ejemplo de uso:
            # Obtener el nombre del script actual
            script_name = os.path.basename(__file__)

            # Cambiar la extensión a .pdf
            pdf_path = os.path.splitext(script_name)[0] + ".pdf"

            df_fields = get_pdf_fields_as_df(pdf_path)
            return df_fields


def separar_y_mayusculizar(texto):
    # Usamos una expresión regular para separar las mayúsculas
    partes = re.findall(r"[A-Z][^A-Z]*", texto)

    # Separar la primera parte (mantenerla como está)
    primera_parte = partes[0]

    # El resto de las partes lo convertimos a mayúsculas y las unimos con espacios
    resto = " ".join([parte.upper() for parte in partes[1:]])

    return primera_parte, resto


def normalizar_cadena(texto):
    if isinstance(texto, str):  # Asegurar que es un string
        texto = texto.lower()
        texto = "".join(
            c
            for c in unicodedata.normalize("NFD", texto)
            if unicodedata.category(c) != "Mn"
        )
        # Capitalizar la primera letra después de un espacio
        texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
        # Eliminar espacios
        texto = re.sub(r"\s+", "", texto)
        # Mantener solo caracteres permitidos
        texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
    return texto




router = APIRouter()



@router.post('/Ind290')
def Ind290RecuperaciónDeCalorEnCircuitoFrigorífico(data: dict):



    # parra que en la primera carga, los campos que tgiene valores separados por ; para los selectores coja el primer valor por defecto, tambien 
    # debera hacerlo si se da a recalcular sin seleccionar un valor 
    for key, values in data.items():
        for i in range(1, len(values)):  # Saltamos la primera fila (encabezados)
            if isinstance(values[i][1], str):  # Verificar si es una cadena antes de dividir
                values[i][1] = values[i][1].split(';')[0]  # Tomar solo el primer valor
    # parra que en la primera carga, los campos que tgiene valores separados por ; para los selectores coja el primer valor por defecto, tambien 
    # debera hacerlo si se da a recalcular sin seleccionar un valor 
    for key, values in data.items():
        for i in range(1, len(values)):  # Saltamos la primera fila (encabezados)
            if isinstance(values[i][1], str):  # Verificar si es una cadena antes de dividir
                values[i][1] = values[i][1].split(';')[0]  # Tomar solo el primer valor
    print(data)
    ss = SimpleNamespace()
    # Ejemplo de uso

    ss.nombre = "Ind290RecuperaciónDeCalorEnCircuitoFrigorífico"
    ss.path = f"app/routers/CERTIFICADOS_DE_AHORRO_ENERGÉTICO_CAE/Industrial/Ind290RecuperaciónDeCalorEnCircuitoFrigorífico.pdf"


    ss.codigo, ss.titulo = separar_y_mayusculizar(ss.nombre)
    ss.data = data
    output_buffer = zz.main.main(ss)
    output_buffer = formulario.formulario.rellenoFormulario(output_buffer, ss)


    # solo para depuracion en produccion quitarlo
    # solo para depuracion en produccion quitarlo
    # solo para depuracion en produccion quitarlo
    # solo para depuracion en produccion quitarlo
    # solo para depuracion en produccion quitarlo
    # with open(
    #     f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf",
    #     "wb",
    # ) as f:
    #     f.write(output_buffer.getvalue())
    # solo para depuracion en produccion quitarlo
    # solo para depuracion en produccion quitarlo
    # solo para depuracion en produccion quitarlo
    # solo para depuracion en produccion quitarlo
    # solo para depuracion en produccion quitarlo

    # devuelve a fastapi el archivo
    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


if __name__ == "__main__":

    # Cambiar al directorio del script
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    df = formulario.formulario.leer_los_campos_del_pdf_formulario()
    data = zz.main.xlsxhojaconvalorespordefecto(df)

    nombre_funcion = os.path.basename(__file__)[:-3]
    globals()[nombre_funcion](data)

