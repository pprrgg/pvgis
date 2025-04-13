


from fastapi import APIRouter

router = APIRouter()


@router.post(f"/Codigo")
def CodigoNombreSeparadoPorMayusculasSinEspacios(
    data: dict,
):
    data = comun.texfile.limpiar_valores(data)
    print(data)
    # parámetros
    pp = SimpleNamespace()
    # tablas y graficos en latex
    uu = SimpleNamespace()
    #
    vv = SimpleNamespace()
    xx = SimpleNamespace()
    yy = SimpleNamespace()
    # datos genericos
    zz = SimpleNamespace()

    zz.nombre = sys._getframe().f_code.co_name
    zz.ruta_script = os.path.dirname(
        os.path.abspath(__file__)
    )  # Obtener la carpeta sin el nombre del script

    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)
    zz.data = data

    output_buffer = comun.main.main(pp, uu, vv, xx, yy, zz)

    with open(f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf", "wb") as f:
        f.write(output_buffer.getvalue())

    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


def file_tex(pp, uu, vv, xx, yy, zz):

    return r"""

\newcommand{\MostrarVariablesAlFinal}{}
%\documentclass[a4paper,10pt]{article}
\documentclass[a4paper,10pt,twocolumn]{article}
\usepackage[top = 2cm, bottom = 2cm, left = 2cm, right = 2cm, asymmetric]{geometry} % Especificar los márgenes según la norma
\usepackage[spanish]{babel}    

\usepackage{graphicx}  % Para incluir imágenes si es necesario
\usepackage{amsmath, amssymb}  % Para expresiones matemáticas
\usepackage{fancyhdr}  % Para personalizar encabezados
\usepackage{chngcntr}  % Para cambiar la numeración de apartados y subsecciones
\usepackage{tocbibind}  % Para incluir bibliografía en la tabla de contenidos
\usepackage{appendix}  % Para el formato de anexos
\usepackage{lipsum}  % Para generar texto de relleno (dummy text)
\usepackage{geometry}  % Para personalizar los márgenes
\usepackage{multicol}  % Para columnas
\usepackage{titlesec}  % Para personalizar los títulos de las secciones
\usepackage[utf8]{inputenc}    
\usepackage[T1]{fontenc}       
\usepackage[pdftex,pdfencoding=auto, colorlinks=true, linkcolor=black, citecolor=black, filecolor=black, urlcolor=black]{hyperref}
\usepackage{tocloft}           
\usepackage{booktabs}          
\usepackage{float} % Paquete necesario para usar la opción [H]
\usepackage{array}
\usepackage{longtable}
\usepackage{circuitikz}
\usepackage{tikz}
\usepackage{tikz-cd}
\usepackage{tikz-qtree}
\usepackage{smartdiagram}

\usepackage{qrcode}            % Paquete para generar QR
\usepackage{tabularx} % Agregar en el preámbulo
\usepackage[absolute,overlay]{textpos} % Para posicionar objetos libremente
\usepackage{underscore}
\usepackage{pdfpages}

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{zz.titulo}  % Nombre del documento en el encabezado izquierdo
\fancyhead[C]{}  % Centro vacío
\fancyhead[R]{\thepage}  % Numeración de páginas en el encabezado derecho
% \renewcommand{\thepage}{\arabic{page}}  % Asegura la numeración de páginas en números arábigos

% Definición de los anexos
\newcommand{\annex}[1]{\section{Anexo #1} \addcontentsline{toc}{section}{Anexo #1}}
\addto\captionsspanish{%
  \renewcommand{\tablename}{Tabla}%
  \renewcommand{\listtablename}{Índice de tablas}%
}
\title{{ \qrcode[height=1.3cm]{https://doctec.blog/} \\ \small Ref.:\uppercase{zz.codigo}}\\{\textbf{zz.titulo}}}
\author{
pp.ParticipantesP1NombreRaznSocial
}
\date{\today}










\begin{document}


\begin{Form}
	\maketitle

	\tableofcontents  % Tabla de contenidos
	\listoffigures    % Lista de graficos
	\listoftables     % Lista de tablas
	\newpage



    












\section{PARTE INICIAL}

% 6.1 Cubierta (páginas 1 y 2 de la cubierta)
\subsection{Cubierta (páginas 1 y 2 de la cubierta)}
Además de proporcionar, en parte, protección física al informe, la cubierta exterior (página 1 de la cubierta) sirve para dar la presentación inicial del informe al usuario. Por ello, debería ser clara, distintiva e informativa. La disposición de la información que lleva la cubierta exterior, se especifica en 10.1.1. En 12.1 y 12.2 se dan detalles relativos a la reproducción de las cubiertas.

uu.mapaUbicacion


uu.tabladfCodigo

Por razones económicas, la portada tal como se describe en 6.2 puede servir como cubierta exterior (página 1 de la cubierta). Se recomienda especialmente esta solución. Entonces, dicha página debería llevar los identificadores del informe en las tres posiciones descritas en 10.1.1 a).
El interior de la cubierta (página 2 de la cubierta) puede utilizarse para incluir alguna de las indicaciones especiales menos importantes descritas en 10.2.9. En algunos casos, se puede utilizar para el prefacio [véase 6.6].

uu.dfCodigo


% 6.2 Portada
\subsection{Portada}
La portada de cualquier documento es la fuente preferente de información bibliográfica para el tratamiento y recuperación eficientes del documento [véase 10.1.3]. Por ello, es esencial que cada informe incluya una portada. Mientras la portada no sustituya a la primera página de la cubierta (página de la cubierta 1), no es necesario que la portada ocupe una página completa. Por razones de economía, puede ocupar un espacio de cabecera antes del resumen [véase 6.3] o antes del resumen y el índice [véase 6.4]. Para una mayor economía, la hoja de datos del documento [véase 9.1 y 10.1.4] puede reemplazar a la portada y al resumen. Esto se recomienda cuando la portada está impresa en la primera página de la cubierta.

Cuando un informe se publica en dos o más partes [véase 5.2.2], cada parte debe contener una portada en la que se indique el número de la parte.



Para uso en red interna se requiere de autorización previa de AENOR.



% 6.3 Resumen
\subsection{Resumen}
Cada informe debe contener un resumen que debe presentarse inmediatamente después de la información de la portada [véase 6.2] y/o incluido en la hoja de datos del documento [véase 10.1.4]. Cuando el resumen figura en una página separada, debe ir precedido o seguido por la identificación bibliográfica completa (biblid) del informe: autor(es), título, identificador del informe, organización responsable y fecha de publicación, tal como figura en la Norma UNE 50127.

El texto del resumen debe estar de acuerdo con la Norma UNE 50103. En síntesis, debe ser tan informativo como lo permita la naturaleza del documento, para que los lectores puedan decidir si es necesario leer el documento completo. Debe definir el objetivo, métodos, resultados y conclusiones presentadas en el documento original, bien en ese orden, o destacando inicialmente los resultados y conclusiones.

Debe ser conciso sin ser oscuro, reteniendo la información básica y el carácter del documento original. Los resúmenes de la mayoría de los informes deben tener menos de 250 palabras y en ningún caso más de 500. Deben estar escritos en un solo párrafo; emplear normalmente frases completas, verbos en forma activa y con tercera persona. No se deben utilizar figuras y símbolos, tales como tablas cortas y fórmulas, más que cuando no haya ninguna alternativa aceptable.



% 6.4 Índice
\subsection{Índice}
El índice es un elemento esencial de todos los informes, salvo los muy cortos. Debe figurar inmediatamente después del resumen. Debe constar de los títulos de las principales subdivisiones del informe y los anexos, junto con el número de las páginas en las que aparecen. También debería de incluirse una lista de ilustraciones y de tablas.

Cuando un informe se publica en dos o más partes [véase 5.2.2], en cada parte debe aparecer el índice completo.

Cuando varios informes se publican en volúmenes separados de un conjunto [véase 5.2.1], cada informe debe contener un índice del volumen particular en que se encuentra y puede contener también un índice común para el conjunto completo.



% 6.5 Glosario de signos, símbolos, unidades, abreviaturas, acrónimos o términos
\subsection{Glosario de signos, símbolos, unidades, abreviaturas, acrónimos o términos}
Cuando el informe contenga signos, símbolos, unidades, abreviaturas, acrónimos o términos que puedan no ser comprendidos rápidamente por los posibles lectores, deberían definirse en una o varias listas después del índice. La existencia de esas listas no justifica la omisión de una explicación sobre esos símbolos cuando aparezcan por primera vez en el texto.



% 6.6 Prefacio
\subsection{Prefacio}
Un prefacio puede considerarse como una nota de presentación para definir el estudio, destacar algún aspecto en particular, mostrar su relación con trabajos relacionados o trazar las circunstancias históricas que condujeron a su iniciación. No siempre es necesario.

Si se requiere prefacio, debería aparecer al final de la parte inicial, inmediatamente antes del cuerpo del informe. Sin embargo, por determinados motivos, puede ser conveniente llamar la atención sobre la información contenida en el prefacio, colocándolo en página 2 de la cubierta.

Cuando un informe se publica en dos o más partes [véase 5.2.2], el prefacio debe aparecer sólo en la primera parte. Cuando varios informes se publican en volúmenes separados de un conjunto [véase 5.2.1], cada informe debe contener su propio prefacio.




\section{Cuerpo del Informe}

\subsection{Introducción}
Cada informe debe comenzar con una introducción que establezca brevemente el campo de aplicación y los objetivos del trabajo descrito, su relación con otros trabajos y el enfoque general. No debe repetir o parafrasear el resumen, ni dar detalles de la teoría experimental, método o resultados, ni anticipar las conclusiones o recomendaciones. Si no hay un prefacio separado, la información que podría figurar en el prefacio puede incluirse en la introducción.

\subsection{Núcleo del Informe, con Ilustraciones y Tablas}
El núcleo del informe debería dividirse en capítulos numerados que cubran aspectos tales como teoría, método, resultados y discusión.

\subsubsection{Organización del Núcleo}
Es conveniente dividir los capítulos en apartados y estos en subapartados, cada uno con su propio encabezamiento. La numeración de los capítulos, apartados y subapartados debe seguir las normativas establecidas.

\subsubsection{Contenido del Núcleo}
La información en el núcleo del informe no debe ser demasiado detallada. Las descripciones de la teoría, métodos y resultados deberían ser suficientes para que un especialista en la materia pueda reproducir las etapas de la investigación. Si se requieren pruebas matemáticas o detalles completos, deben presentarse en los anexos.

\subsubsection{Ilustraciones y Tablas}
Todas las ilustraciones y tablas esenciales para la comprensión del texto principal deberían incluirse en el núcleo del informe y colocarse como se indica en las normas correspondientes.

\subsubsection{Discusión}
Se puede añadir un capítulo independiente para la discusión de los nuevos aspectos del trabajo descrito y la interpretación o comentarios de los resultados, fundamentando las conclusiones y recomendaciones del informe.

\subsection{Conclusiones y Recomendaciones}
Las conclusiones deben ser el reflejo claro y ordenado de las deducciones hechas como consecuencia del trabajo descrito. Pueden incluir datos cuantitativos pero no deberían incluir detalles extensos de los argumentos o resultados. Las recomendaciones deben ser concisas y justificadas por el trabajo descrito. Si existe un gran número de recomendaciones, estas pueden constituir un capítulo independiente.























% --- 7. REFERENCIAS BIBLIOGRÁFICAS ---
\begin{thebibliography}{99}
    \bibitem{ccee} \href{https://www.boe.es/buscar/pdf/2021/BOE-A-2021-7120-consolidado.pdf}{Resolución de 28 de abril de 2021, de la Dirección General de Política Energética y Minas, por la que se establece el contenido
mínimo y el modelo de factura de electricidad a utilizar por los comercializadores de referencia.}


\bibitem{rd244}
\href{}{RD 244/2019  sobre autoconsumo}

\bibitem{esios} 
\href{https://www.esios.ree.es/es/pvpc}{ESIOS - Red Eléctrica de España. PVPC y datos del sistema eléctrico
}


\bibitem{boe}
\href{}{Real Decreto 216/2014 por el que se establece la metodología de cálculo de los precios voluntarios para el pequeño consumidor.}


\bibitem{boe}
\href{https://www.cnmc.es/sites/default/files/3416756_135.pdf}{ACUERDO POR EL QUE SE CONTESTAN CONSULTAS RELATIVAS A LA
APLICACIÓN DE LA CIRCULAR 3/2020, DE 15 DE ENERO, POR LA QUE SE
ESTABLECE LA METODOLOGÍA PARA EL CÁLCULO DE LOS PEAJES DE
TRANSPORTE Y DISTRIBUCIÓN DE ENERGÍA ELÉCTRICA.}




\end{thebibliography}








\ifdefined\MostrarVariablesAlFinal
\newpage
\onecolumn
HojaDeDatos
\fi
\end{Form}
\end{document}

"""


class datos:

    class datos:

        def d00data2parametros(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Para cada hoja en el JSON
            # Convertir cada valor del diccionario en un DataFrame
            for key, value in zz.data.items():
                key_sanitized = comun.texfile.sanitize_name(
                    key
                )  # Limpiar el nombre de la clave
                df = pd.DataFrame(
                    value[1:], columns=value[0]
                )  # Primera fila como encabezado
                df.set_index(
                    df.columns[0], inplace=True
                )  # Usar la primera columna como índice

                setattr(xx, f"df{key_sanitized}", df)  # Guardar en ss con nombre limpio
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == zz.codigo:
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar índice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{idx_sanitized}",
                                val,
                            )
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == "Coordenadas" or key_sanitized == "Participantes":
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar índice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{key_sanitized}{col_sanitized}{idx_sanitized}",
                                val,
                            )

            return pp, uu, vv, xx, yy, zz

        def d00EjemplDeTabla(pp, uu, vv, xx, yy, zz):
            print(f"\n{100*'+'}{__class__.__name__}.{sys._getframe().f_code.co_name}")
            #
            df = pd.DataFrame(
                np.random.randint(0, 100, size=(2, 3)), columns=["A", "B", "C"]
            )
            # siempre tieen que ir el indice que es donde meto separados por _ indice_ejeX_ejeY
            df = df.reset_index(drop=True)  # drop=True elimina el índice anterior
            # df = df.set_index(df.columns[0])
            df.index.name = "EUR/MWh_hora semanal"
            #
            setattr(xx, sys._getframe().f_code.co_name, df)
            return pp, uu, vv, xx, yy, zz

        def d00TodosLosExcelDeLaCarpetaXlsx(pp, uu, vv, xx, yy, zz):
            """para los datos que no hay ques esta descargancodon continuamente por ejm los pvpc..."""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpetaimagenes = Path(f"{zz.ruta_script}/assets")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.xlsx")]

            for imagen in imagenes:
                # _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                titulo = imagen[:-4]
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional

                ruta_archivo_excel = Path(f"{zz.ruta_script}/assets/{imagen}")
                # libro = openpyxl.load_workbook(ruta_archivo_excel)

                # Cargar todas las hojas del archivo Excel
                sheets = pd.read_excel(
                    ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
                )
                for nombre_hoja, df in sheets.items():
                    nombre_hoja = comun.texfile.normalizar_cadena(nombre_hoja)
                    setattr(xx, f"{nombre_hoja}", df)  # Guardar en ss con nombre limpio

            return pp, uu, vv, xx, yy, zz



########################################################################################################
########################################################################################################


class comun:
    class main:

        def main(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # segun si es un formulario o un documento nuevo

            pp, uu, vv, xx, yy, zz = comun.main.datos(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.datos2latex(pp, uu, vv, xx, yy, zz)

            output_buffer = comun.main.replaceLatex(pp, uu, vv, xx, yy, zz)

            # print(uu)
            return output_buffer

        def datos(pp, uu, vv, xx, yy, zz):
            """datoes"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listadatoes = [
                m
                for m in dir(datos.datos)
                if callable(getattr(datos.datos, m)) and not m.startswith("__")
            ]
            listadatoes = [s for s in listadatoes if not s.startswith("__")]
            listadatoes = sorted(listadatoes)

            for dato in listadatoes:
                print(dato)
                metodo = getattr(datos.datos, dato, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def datos2latex(pp, uu, vv, xx, yy, zz):
            """tablas"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listatablas = [
                m
                for m in dir(comun.datos2latex)
                if callable(getattr(comun.datos2latex, m)) and not m.startswith("__")
            ]
            listatablas = [s for s in listatablas if not s.startswith("__")]
            for tabla in listatablas:
                # print(tabla)
                metodo = getattr(
                    comun.datos2latex, tabla, None
                )  # Obtener el método o None si no existe
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)

            return pp, uu, vv, xx, yy, zz

        def replaceLatex(pp, uu, vv, xx, yy, zz):
            """
            Compila un archivo .tex a PDF, reemplaza un texto y devuelve el PDF en memoria.
            :param ruta_tex: Ruta al archivo .tex original.
            :return: Objeto BytesIO con el contenido del PDF.
            """

            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            pdf_path = Path(zz.nombre).with_suffix(".pdf")
            ruta_tex = f"{os.path.splitext(pdf_path)[0]}.tex"

            nombre_archivo = os.path.basename(ruta_tex)  # Nombre del archivo sin ruta

            # Crear un directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                ruta_tex_temp = os.path.join(temp_dir, nombre_archivo)

                contenido = file_tex(pp, uu, vv, xx, yy, zz)
                # tabla final con todas los parámetros  y variables
                HojaDeDatos = "\n"

                for var_name, value in pp.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\npp. {var_name}\n\npp.{var_name}"

                for var_name, value in uu.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nuu. {var_name}\n\nuu.{var_name}"

                for var_name, value in vv.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nvv. {var_name}\n\nvv.{var_name}"

                for var_name, value in xx.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nxx. {var_name}\n\nxx.{var_name}"

                for var_name, value in yy.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nyy. {var_name}\n\nyy.{var_name}"

                # noooooooo,!!!!! que el json del data da error en latex
                # for var_name, value in zz.__dict__.items():
                #     print(f"zz.{var_name}")
                #     HojaDeDatos = HojaDeDatos + f"\n\nzz. {var_name}\n\nzz.{var_name}"

                contenido = contenido.replace(f"HojaDeDatos", HojaDeDatos)

                if 1:
                    # Reemplazar directamente todas las variables en el texto
                    for var_name, value in pp.__dict__.items():
                        # print(f"pp.{var_name}")
                        contenido = contenido.replace(
                            f"pp.{var_name}",
                            f"{str(value)}",
                        )
                    for var_name, value in uu.__dict__.items():
                        # print(f"uu.{var_name}")
                        contenido = contenido.replace(f"uu.{var_name}", f"{str(value)}")

                    for var_name, value in vv.__dict__.items():
                        # print(f"vv.{var_name}")
                        contenido = contenido.replace(f"vv.{var_name}", f"{str(value)}")

                    for var_name, value in xx.__dict__.items():
                        # print(f"xx.{var_name}")
                        contenido = contenido.replace(f"xx.{var_name}", f"{str(value)}")

                    for var_name, value in yy.__dict__.items():
                        # print(f"yy.{var_name}")
                        contenido = contenido.replace(f"yy.{var_name}", f"{str(value)}")

                    for var_name, value in zz.__dict__.items():
                        # print(f"zz.{var_name}")
                        contenido = contenido.replace(f"zz.{var_name}", f"{str(value)}")

                with open(
                    "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.tex", "w", encoding="utf-8"
                ) as file:
                    file.write(contenido)

                # Guardar los cambios en el archivo temporal
                with open(ruta_tex_temp, "w", encoding="utf-8") as file:
                    file.write(contenido)

                # Función para ejecutar pdflatex dos veces
                def ejecutar_pdflatex():
                    print(
                        f"\n{100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                    )
                    if 1:
                        subprocess.run(
                            ["pdflatex", "-output-directory", temp_dir, ruta_tex_temp],
                            check=True,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )

                    return True

                # # para que se cre el indice debe estar descomentado
                # Ejecutar pdflatex dos veces para generar la TOC correctamente
                if not (ejecutar_pdflatex() and ejecutar_pdflatex()):
                    return None  # Si la compilación falla, se detiene el proceso
                # ejecutar_pdflatex()

                # Ruta del PDF generado
                pdf_generado = os.path.join(
                    temp_dir, nombre_archivo.replace(".tex", ".pdf")
                )

                # Leer el PDF en memoria
                with open(pdf_generado, "rb") as file:
                    pdf_en_memoria = BytesIO(file.read())

            return pdf_en_memoria

            #######################

    class datos2latex:

        def tabla(pp, uu, vv, xx, yy, zz):
            # tabla de todos los df

            for var_name, value in {**xx.__dict__, **yy.__dict__}.items():
                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                if var_name == zz.codigo:
                    titulo = "Parámetros"

                if isinstance(value, pd.DataFrame):
                    # para no hacer tablas de series muy
                    if len(value) > 10:
                        continue
                    # Redondear los valores del DataFrame a 1 decimal
                    value = value.round(1)

                    latex_code = value.to_latex(float_format="%.1f")

                    latex_code_with_resizebox = rf"""
                    \begin{{table}}[H] \centering
                        % \resizebox{{0.4\textwidth}}{{!}}
                        {{
                        {latex_code}
                        }}
                        \caption{{{titulo}}}
                    \end{{table}}
                    """

                    # Asignar el LaTeX a la propiedad correspondiente de uu
                    setattr(uu, f"tabla{var_name}", latex_code_with_resizebox)

            # Convierte todos los df de uu a tablas
            return pp, uu, vv, xx, yy, zz

        def grafico(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # tabla de todos los df

            for var_name, value in {**xx.__dict__, **yy.__dict__}.items():
                print(var_name)
                if hasattr(xx, var_name):
                    df = getattr(xx, var_name)
                else:
                    df = getattr(yy, var_name)

                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                # 1. Crear un archivo temporal para la imagen
                with tempfile.NamedTemporaryFile(
                    suffix=".png", delete=False
                ) as temp_image:
                    temp_image_path2 = (
                        temp_image.name
                    )  # Guardar la ruta del archivo temporal

                # 📌 Gráfico del flujo de caja y acumulado
                if len(df) < 24:
                    fig, ax = plt.subplots(figsize=(5, 3))
                    try:
                        df.plot(kind="bar", ax=ax)

                        plt.xticks(rotation=45)
                        plt.tight_layout()  # Ajustar para que no se corten las etiquetas
                    except:
                        """no valores numericos para plt"""

                    try:
                        ax.set_ylabel(df.index.name.split("_")[0])
                        ax.set_xlabel(df.index.name.split("_")[1])
                    except:
                        """"""
                else:

                    fig, ax = plt.subplots(figsize=(6, 5))

                    df.plot(ax=ax)

                    # Obtener el nombre de la columna automáticamente
                    try:
                        ax.set_ylabel(df.index.name.split("_")[0])
                        ax.set_xlabel(df.index.name.split("_")[1])
                    except:
                        """"""

                plt.grid(axis="y", linestyle="--", alpha=0.7)
                plt.savefig(temp_image_path2, format="png", dpi=300)
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
                latex_code = (
                    ""
                    + r"""\begin{figure}[H] \centering
                    """
                    + rf"""
                    \includegraphics[width=.5\textwidth]{{{temp_image_path2}}}
                    \caption{{{titulo}}}
                    """
                    + """
                    \label{fig:dfassssdfsa}
                    \end{figure}
                    """
                )
                setattr(uu, f"{var_name}", latex_code)
            return pp, uu, vv, xx, yy, zz

        def imagen(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpetaimagenes = Path(f"{zz.ruta_script}/{zz.codigo}/img")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.png")] + [
                f.name for f in carpetaimagenes.glob("*.jpg")
            ]

            for imagen in imagenes:
                # _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                titulo = imagen[:-4]
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional

                latex_code = (
                    ""
                    + r"""\begin{figure}[H] \centering
                    """
                    + rf"""
                    \includegraphics[width=0.7\textwidth]{{{Path(f"{zz.ruta_script}/{zz.codigo}/img/{imagen}")}}} 
                    \caption{{{titulo}}}
                    """
                    + """
                    \label{fig:dfasdfsa}
                    \end{figure}
                    """
                )

                setattr(uu, f"imagen{imagen}", latex_code)
            return pp, uu, vv, xx, yy, zz






        def mapaUbicacionbien(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            with tempfile.NamedTemporaryFile(
                suffix=".png", delete=False
            ) as temp_image:
                temp_image_path = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal


            import geopandas as gpd
            import matplotlib.pyplot as plt
            from shapely.geometry import Point
            from geopy.geocoders import Nominatim
            from geopy.extra.rate_limiter import RateLimiter
            # Cargar el GeoJSON
            geojson_path = f"{zz.ruta_script}/assets/spain-communities.geojson"
            spain_communities = gpd.read_file(geojson_path)
            spain_communities = spain_communities[spain_communities["name"] != "Canarias"]
            # Configurar geocodificador
            geolocator = Nominatim(user_agent="mapa_espana_detallado")
            geocode = RateLimiter(geolocator.reverse, min_delay_seconds=1)
            # Coordenadas del punto
            # lat, lng = 37.3883251804357, -4.508585281106158
            lat, lng = pp.CoordenadasValorlat, pp.CoordenadasValorlng

            point = Point(lng, lat)
            # Obtener dirección detallada
            try:
                location = geocode(f"{lat}, {lng}", language='es', exactly_one=True)
                if location:
                    address = location.raw['address']
                    calle = address.get('road', '')
                    numero = address.get('house_number', '')
                    ciudad = address.get('city', address.get('town', address.get('village', '')))
                    provincia = address.get('state', '')
                    codigo_postal = address.get('postcode', '')  # <- Aquí obtenemos el CP

                    direccion_completa = f"{calle} {numero} , {ciudad} {codigo_postal}, {provincia}"
                else:
                    direccion_completa = "Dirección no disponible"
            except Exception as e:
                direccion_completa = f"Error al obtener dirección: {str(e)}"
            # Convertir a proyección adecuada para España (UTM zona 30N)
            spain_communities = spain_communities.to_crs(epsg=25830)
            point_gdf = gpd.GeoSeries([point], crs="EPSG:4326").to_crs(epsg=25830)
            # Crear figura
            fig, ax = plt.subplots(figsize=(7, 7))
            # Título con dirección completa
            ax.set_title(f"{direccion_completa}", fontsize=14, pad=20)
            # Dibujar comunidades
            spain_communities.plot(ax=ax, color="lightgrey", edgecolor="black", linewidth=0.5)
            # Dibujar punto de ubicación
            point_gdf.plot(
                ax=ax,
                color="red",
                markersize=100,
                marker="o",
                edgecolor="black",
                label="Ubicación"
            )
            # Añadir coordenadas como texto (proyectadas de nuevo para mostrar en grados)
            ax.annotate(
                f"Coordenadas:\n{lat:.6f}° N\n{lng:.6f}° W",
                xy=point_gdf.geometry[0].coords[0],
                xytext=(15, 15),
                textcoords="offset points",
                bbox=dict(boxstyle="round", fc="white", ec="black", alpha=0.9),
                fontsize=10
            )
            # Ajustar aspecto del mapa
            xmin, ymin, xmax, ymax = spain_communities.total_bounds
            padding = 80000  # padding en metros porque estamos en UTM
            ax.set_xlim(xmin - padding, xmax + padding)
            ax.set_ylim(ymin - padding, ymax + padding)
            # Añadir leyenda y grid
            ax.legend(loc='upper right')
            ax.grid(True, linestyle='--', alpha=0.5)
            plt.tight_layout()
            # plt.show()



            plt.savefig(temp_image_path, format="png", dpi=300)
            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + r"""\begin{figure}[H] \centering
                """
                + rf"""
                \includegraphics[width=.5\textwidth]{{{temp_image_path}}}
                \caption{{{'titulo'}}}
                """
                + """
                \label{fig:dfassssdfsa}
                \end{figure}
                """
            )
            setattr(uu, f"mapaUbicacion", latex_code)
            return pp, uu, vv, xx, yy, zz




    class texfile:

        def normalizar_cadena(texto):
            if isinstance(texto, str):  # Asegurar que es un string
                texto = texto.lower()
                texto = "".join(
                    c
                    for c in unicodedata.normalize("NFD", texto)
                    if unicodedata.category(c) != "Mn"
                )
                # Capitalizar la primera letra después de un espacio
                texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
                # Eliminar espacios
                texto = re.sub(r"\s+", "", texto)
                # Mantener solo caracteres permitidos
                texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
            return texto

        def limpiar_valores(data):
            for key, values in data.items():
                for i in range(
                    1, len(values)
                ):  # Saltamos la primera fila (encabezados)
                    for j in range(
                        1, len(values[i])
                    ):  # Recorremos todas las columnas (excepto encabezados)
                        if isinstance(
                            values[i][j], str
                        ):  # Verificar si es una cadena antes de dividir
                            valor_limpio = (
                                values[i][j].split(";")[0].strip()
                            )  # Tomar solo el primer valor y limpiar espacios

                            # Intentar convertir a flotante si es posible
                            try:
                                values[i][j] = float(valor_limpio)
                            except ValueError:
                                values[i][
                                    j
                                ] = valor_limpio  # Si no se puede convertir, dejar como está
            return data

        def separar_y_mayusculizar(texto):
            # Usamos una expresión regular para separar las mayúsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            # Separar la primera parte (mantenerla como está)
            primera_parte = partes[0]

            # El resto de las partes lo convertimos a mayúsculas y las unimos con espacios
            resto = " ".join([parte.upper() for parte in partes[1:]])
            # resto = " ".join([parte for parte in partes[1:]])

            return primera_parte, resto

        def separar_y_formatear(texto):
            # Usamos una expresión regular para separar las mayúsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            if not partes:
                return ""

            # Convertimos todo a minúsculas y capitalizamos solo la primera letra
            resultado = " ".join([parte.lower() for parte in partes])
            return resultado.capitalize()

        # Función para limpiar nombres de variables
        def sanitize_name(name):
            return re.sub(
                r"[^a-zA-Z0-9]", "", str(name)
            )  # Elimina cualquier carácter no alfanumérico


class especifico:
    class previo:
        def p00parametros():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            nombre_funcion = os.path.basename(__file__)[:-3]
            codigo, _ = comun.texfile.separar_y_mayusculizar(nombre_funcion)
            nan = " "
            print(codigo)
            zz.data = {
                zz.codigo: [
                    ["Parametro", "Valor"],
                    ["Inversión inicial [€]", 0],
                    ["Bonificación  inicial [€]", 10000],
                    ["Duración de la actuación [años]", 15],
                    ["Préstamo recibido [€]", 100000],
                    ["Tasa de interés del préstamo [\%]", 5],
                    ["Plazo del préstamo [años]", 15],
                    ["Inflación annual [\%]", 3],
                    ["Ahorro annual [€]", 25000],
                    ["Costos fijos operativos [€]", 1000],
                    ["Tasa de descuento [\%]", 6],
                ],
                "Coordenadas": [
                    ["Parámetro", "Valor"],
                    ["lat", 36.6641681434112],
                    ["lng", -4.45863940137516],
                ],
                "Participantes": [
                    ["Parámetro", "P1"],
                    ["Rol", ""],
                    ["DNI/NIF/NIE", ""],
                    [
                        "Nombre/Razón Social",
                        "'EDITAR' para poner tu nombre y ajustar los parámetros.",
                    ],
                    ["Titulación", ""],
                ],
            }

        def p01carpetaauxiliar():
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpeta = Path(f"{zz.ruta_script}/assets")

            if not os.path.exists(carpeta):
                os.mkdir(carpeta)

        def p02xlsxhojaconvalorespordefecto():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # si es un nuevo archivo tex copia los parámetros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            codigo, titulo = comun.texfile.separar_y_mayusculizar(nombre)
            #
            ruta_archivo_excel = f"{zz.ruta_script}/{codigo}.xlsx"
            if 1:
                df = pd.DataFrame(zz.data, columns=["Parámetro", "Valor"])
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    # Crear un libro de trabajo de Excel
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja predeterminada creada por openpyxl
                    libro.remove(libro.active)

                    # Crear hojas con los nombres de los keys y escribir los datos
                    for sheet_name, sheet_data in zz.data.items():
                        hoja = libro.create_sheet(
                            title=sheet_name
                        )  # Crear la hoja con el nombre del key
                        for row in sheet_data:
                            hoja.append(row)  # Agregar cada fila a la hoja

                    # Guardar el archivo Excel
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    # print(ruta_archivo_excel)
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # Cargar todas las hojas del archivo Excel
            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )

            # Crear el diccionario data
            zz.data = {
                sheet_name: [df.columns.tolist()]
                + df.values.tolist()  # Incluir cabecera y datos
                for sheet_name, df in sheets.items()
            }


from fastapi.responses import StreamingResponse
from io import BytesIO
import os
import matplotlib.pyplot as plt
import pandas as pd
from types import SimpleNamespace
import subprocess
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata


import numpy as np
import sys

if __name__ == "__main__":
    zz = SimpleNamespace()

    # Copio la ruta del scripot para crear el excel
    zz.ruta_script = os.path.dirname(os.path.abspath(__file__))
    zz.nombre = os.path.basename(__file__)[:-3]
    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)

    #
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    os.chdir("../../../../")  # Cambia dos niveles arriba

    # clase especifico
    metodos = [
        m
        for m in dir(especifico.previo)
        if callable(getattr(especifico.previo, m)) and not m.startswith("__")
    ]
    metodos = [s for s in metodos if not s.startswith("__")]
    metodos = sorted(metodos)

    for metodo in metodos:
        metodoi = getattr(especifico.previo, metodo, None)
        metodoi()

    # la funcion de fastpai
    globals()[os.path.basename(__file__)[:-3]](zz.data)
















































