from fastapi import APIRouter

router = APIRouter()


@router.post(f"/Aif030")
def Aif030AnálisisDeOfertaDeAutoconsumo(
    data: dict,
):
    data = comun.texfile.limpiar_valores(data)
    print(data)
    # parámetros
    pp = SimpleNamespace()
    # tablas y graficos en latex
    uu = SimpleNamespace()
    #
    vv = SimpleNamespace()
    xx = SimpleNamespace()
    yy = SimpleNamespace()
    # datos genericos
    zz = SimpleNamespace()

    zz.nombre = sys._getframe().f_code.co_name
    zz.ruta_script = os.path.dirname(
        os.path.abspath(__file__)
    )  # Obtener la carpeta sin el nombre del script

    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)
    zz.data = data

    output_buffer = comun.main.main(pp, uu, vv, xx, yy, zz)

    with open(f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf", "wb") as f:
        f.write(output_buffer.getvalue())

    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


def file_tex(pp, uu, vv, xx, yy, zz):

    return r"""

\newcommand{\MostrarVariablesAlFinal}{}
%\documentclass[a4paper,10pt]{article}
\documentclass[a4paper,10pt,twocolumn]{article}
\usepackage[top = 2cm, bottom = 2cm, left = 2cm, right = 2cm, asymmetric]{geometry} % Especificar los márgenes según la norma
\usepackage[spanish]{babel}    

\usepackage{graphicx}  % Para incluir imágenes si es necesario
\usepackage{amsmath, amssymb}  % Para expresiones matemáticas
\usepackage{fancyhdr}  % Para personalizar encabezados
\usepackage{chngcntr}  % Para cambiar la numeración de apartados y subsecciones
\usepackage{tocbibind}  % Para incluir bibliografía en la tabla de contenidos
\usepackage{appendix}  % Para el formato de anexos
\usepackage{lipsum}  % Para generar texto de relleno (dummy text)
\usepackage{geometry}  % Para personalizar los márgenes
\usepackage{multicol}  % Para columnas
\usepackage{titlesec}  % Para personalizar los títulos de las secciones
\usepackage[utf8]{inputenc}    
\usepackage[T1]{fontenc}       
\usepackage[pdftex,pdfencoding=auto, colorlinks=true, linkcolor=black, citecolor=black, filecolor=black, urlcolor=black]{hyperref}
\usepackage{tocloft}           
\usepackage{booktabs}          
\usepackage{float} % Paquete necesario para usar la opción [H]
\usepackage{array}
\usepackage{longtable}
\usepackage{circuitikz}
\usepackage{tikz}
\usepackage{tikz-cd}
\usepackage{tikz-qtree}
\usepackage{smartdiagram}
\usepackage{siunitx}

\usepackage{qrcode}            % Paquete para generar QR
\usepackage{tabularx} % Agregar en el preámbulo
\usepackage[absolute,overlay]{textpos} % Para posicionar objetos libremente
\usepackage{underscore}
\usepackage{pdfpages}

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{zz.titulo}  % Nombre del documento en el encabezado izquierdo
\fancyhead[C]{}  % Centro vacío
\fancyhead[R]{\thepage}  % Numeración de páginas en el encabezado derecho
% \renewcommand{\thepage}{\arabic{page}}  % Asegura la numeración de páginas en números arábigos

% Definición de los anexos
\newcommand{\annex}[1]{\section{Anexo #1} \addcontentsline{toc}{section}{Anexo #1}}
\addto\captionsspanish{%
  \renewcommand{\tablename}{Tabla}%
  \renewcommand{\listtablename}{Índice de tablas}%
}
\title{{\small Ref.:\uppercase{zz.codigo}}\\{\textbf{zz.titulo}}}
\author{
pp.ParticipantesP1NombreRaznSocial
}
\date{\today}







\begin{document}
\begin{Form}
%% --- PORTADA ---
%\begin{titlepage}
%    \centering
%    {\Large \textbf{TÍTULO DEL INFORME}}\[1.5cm]
%    %\includegraphics[width=0.4\textwidth]{logo.png}\[1cm] % Logo de la empresa/institución
%    
%    \textbf{Autor:} Nombre del Autor \[0.5cm]
%    \textbf{Fecha:} \today \[1cm]
%    \textbf{Código del documento:} REF-XXXX-YYYY \[1cm]
%    \textbf{Versión:} 1.0 \[1cm]
%    \textbf{Empresa / Institución:} Nombre \[2cm]
%    \vfill
%\end{titlepage}
	\maketitle
    %
	%\begin{abstract}

    %\end{abstract}
% --- ÍNDICE ---
	\tableofcontents  % Tabla de contenidos
	\listoffigures    % Lista de graficos
	\listoftables     % Lista de tablas
	\newpage



    

    













    

































































uu.mapaUbicacion


uu.tabladfCodigo


uu.d03fechadfConsumo

uu.d50barConsumoMesual

uu.d50barConsumoSemanalMedio

uu.d50barConsumoDiaLaboralMedio

uu.tablad50barConsumoDiaLaboralMedio




uu.d00t20tdheatTarifa


uu.d00tdh6heatTarifa


uu.d10heatfmtConsumoMensualPorPeriodo





















































    







    

















% --- 7. REFERENCIAS BIBLIOGRÁFICAS ---
\begin{thebibliography}{99}
    \bibitem{ccee} \href{https://www.boe.es/buscar/pdf/2021/BOE-A-2021-7120-consolidado.pdf}{Resolución de 28 de abril de 2021, de la Dirección General de Política Energética y Minas, por la que se establece el contenido
mínimo y el modelo de factura de electricidad a utilizar por los comercializadores de referencia.}


\bibitem{rd244}
\href{}{RD 244/2019  sobre autoconsumo}

\bibitem{esios} 
\href{https://www.esios.ree.es/es/pvpc}{ESIOS - Red Eléctrica de España. PVPC y datos del sistema eléctrico
}


\bibitem{boe}
\href{}{Real Decreto 216/2014 por el que se establece la metodología de cálculo de los precios voluntarios para el pequeño consumidor.}


\bibitem{boe}
\href{https://www.cnmc.es/sites/default/files/3416756_135.pdf}{CNMC. ACUERDO POR EL QUE SE CONTESTAN CONSULTAS RELATIVAS A LA
APLICACIÓN DE LA CIRCULAR 3/2020, DE 15 DE ENERO, POR LA QUE SE
ESTABLECE LA METODOLOGÍA PARA EL CÁLCULO DE LOS PEAJES DE
TRANSPORTE Y DISTRIBUCIÓN DE ENERGÍA ELÉCTRICA.}



\bibitem{boe}
\href{https://www.boe.es/buscar/act.php?id=BOE-A-2020-1066}{Circular 3/2020, de 15 de enero, de la Comisión Nacional de los Mercados y la Competencia, por la que se establece la metodología para el cálculo de los peajes de transporte y distribución de electricidad.}



\end{thebibliography}





































\ifdefined\MostrarVariablesAlFinal
\newpage
\onecolumn
HojaDeDatos
\fi
\end{Form}
\end{document}

"""


class datos:

    class datos:

        def d00TodosLosExcelDeLaCarpetaXlsx(pp, uu, vv, xx, yy, zz):
            """para los datos que no hay ques esta descargancodon continuamente por ejm los pvpc..."""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpetaimagenes = Path(f"{zz.ruta_script}/assets")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.xlsx")]

            for imagen in imagenes:
                # _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                titulo = imagen[:-4]
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional

                ruta_archivo_excel = Path(f"{zz.ruta_script}/assets/{imagen}")
                # libro = openpyxl.load_workbook(ruta_archivo_excel)

                # Cargar todas las hojas del archivo Excel
                sheets = pd.read_excel(
                    ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
                )
                for nombre_hoja, df in sheets.items():
                    nombre_hoja = comun.texfile.normalizar_cadena(nombre_hoja)
                    setattr(xx, f"{nombre_hoja}", df)  # Guardar en ss con nombre limpio

            return pp, uu, vv, xx, yy, zz

        def d00data2parametros(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Para cada hoja en el JSON
            # Convertir cada valor del diccionario en un DataFrame
            for key, value in zz.data.items():
                key_sanitized = comun.texfile.sanitize_name(
                    key
                )  # Limpiar el nombre de la clave
                df = pd.DataFrame(
                    value[1:], columns=value[0]
                )  # Primera fila como encabezado
                df.set_index(
                    df.columns[0], inplace=True
                )  # Usar la primera columna como índice

                setattr(xx, f"df{key_sanitized}", df)  # Guardar en ss con nombre limpio
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == zz.codigo:
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar índice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{idx_sanitized}",
                                val,
                            )
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == "Coordenadas" or key_sanitized == "Participantes":
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar índice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{key_sanitized}{col_sanitized}{idx_sanitized}",
                                val,
                            )

            return pp, uu, vv, xx, yy, zz

        def d00t20tdheatTarifa(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # 1. Definir los mapeos de periodos

            # Definimos los periodos hora a hora para cada mes (abreviados a 3 letras)
            meses = {
                "Ene": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Feb": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Mar": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Abr": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "May": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Jun": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Jul": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Ago": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Sep": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Oct": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Nov": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Dic": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "SDF": [
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                    3,
                ],
            }

            # Creamos el DataFrame
            horas = [f"{h:02d}:00" for h in range(24)]  # Columnas: 00:00 a 23:00
            df = pd.DataFrame.from_dict(meses, orient="index", columns=horas)
            df = df.T

            df = df.reset_index(drop=True)  # drop=True elimina el índice anterior
            df.index.name = " _2.0 TD"

            # df_tarifa.index.name = "Hora"

            #
            setattr(xx, sys._getframe().f_code.co_name, df)
            return pp, uu, vv, xx, yy, zz

        def d00tdh6heatTarifa(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # Definimos los periodos hora a hora para cada mes (abreviados a 3 letras)
            meses = {
                "Ene": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    2,
                    1,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Feb": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    2,
                    1,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Mar": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    3,
                    2,
                    2,
                    2,
                    2,
                    2,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    2,
                    2,
                    3,
                    3,
                ],
                "Abr": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    5,
                    4,
                    4,
                    4,
                    4,
                    4,
                    5,
                    5,
                    5,
                    5,
                    4,
                    4,
                    4,
                    4,
                    5,
                    5,
                ],
                "May": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    5,
                    4,
                    4,
                    4,
                    4,
                    4,
                    5,
                    5,
                    5,
                    5,
                    4,
                    4,
                    4,
                    4,
                    5,
                    5,
                ],
                "Jun": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    4,
                    3,
                    3,
                    3,
                    3,
                    3,
                    4,
                    4,
                    4,
                    4,
                    3,
                    3,
                    3,
                    3,
                    4,
                    4,
                ],
                "Jul": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    2,
                    1,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "Ago": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    4,
                    3,
                    3,
                    3,
                    3,
                    3,
                    4,
                    4,
                    4,
                    4,
                    3,
                    3,
                    3,
                    3,
                    4,
                    4,
                ],
                "Sep": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    4,
                    3,
                    3,
                    3,
                    3,
                    3,
                    4,
                    4,
                    4,
                    4,
                    3,
                    3,
                    3,
                    3,
                    4,
                    4,
                ],
                "Oct": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    5,
                    4,
                    4,
                    4,
                    4,
                    4,
                    5,
                    5,
                    5,
                    5,
                    4,
                    4,
                    4,
                    4,
                    5,
                    5,
                ],
                "Nov": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    3,
                    2,
                    2,
                    2,
                    2,
                    2,
                    3,
                    3,
                    3,
                    3,
                    2,
                    2,
                    2,
                    2,
                    3,
                    3,
                ],
                "Dic": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    2,
                    1,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    2,
                    2,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                ],
                "SDF": [
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                    6,
                ],
            }

            # Creamos el DataFrame
            horas = [f"{h:02d}:00" for h in range(24)]  # Columnas: 00:00 a 23:00
            df = pd.DataFrame.from_dict(meses, orient="index", columns=horas)
            df = df.T
            df = df.reset_index(drop=True)  # drop=True elimina el índice anterior
            df.index.name = " _2.0 TD"

            setattr(xx, sys._getframe().f_code.co_name, df)
            return pp, uu, vv, xx, yy, zz

        def d03fechadfConsumo(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            df = xx.dfconsumo

            # Extraer componentes de la fecha
            anio = int(df.iloc[0, 0])
            mes = int(df.iloc[1, 0])
            dia = int(df.iloc[2, 0])

            # Crear fecha inicial
            fecha_inicio = pd.Timestamp(year=anio, month=mes, day=dia)

            # Filtrar datos (eliminar primeras 3 filas)
            df = df.iloc[3:].copy()
            # Generar índice horario
            df.index = pd.date_range(start=fecha_inicio, periods=len(df), freq="H")

            df.index.name = "MWh_hora semanal"

            # Guardar en la instancia xx como atributo dinámico
            setattr(xx, sys._getframe().f_code.co_name, df)  # Valores originales

            # # Convertir a datetime y extraer componentes
            # df["datetime"] = pd.to_datetime(df.index, utc=True)
            # df["hour"] = df["datetime"].dt.hour
            # df["week"] = df["datetime"].dt.isocalendar().week
            # df["year"] = df["datetime"].dt.isocalendar().year
            # df["day_of_week"] = df["datetime"].dt.dayofweek  # 0 = lunes, 6 = domingo

            return pp, uu, vv, xx, yy, zz

        def d30AsignarAgrupacionesDeTiempo(pp, uu, vv, xx, yy, zz):
            """asignar columnas de semana mes hora, tarifas..."""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            df = xx.d03fechadfConsumo.copy()

            # Primero asegurémonos de que tenemos las columnas necesarias
            df["datetime"] = pd.to_datetime(df.index, utc=True)
            df["hour"] = df["datetime"].dt.hour
            df["month"] = (
                df["datetime"].dt.month_name().str[:3]
            )  # Obtener abreviación del mes (Ene, Feb, etc.)
            df["month_num"] = df["datetime"].dt.month - 1  # Convertir mes a índice 0-11
            df["day_of_week"] = df["datetime"].dt.dayofweek  # 0=lunes, 6=domingo

            # Definir si es fin de semana (Sábado=5, Domingo=6) o laborable
            df["es_fin_de_semana"] = df["day_of_week"].isin([5, 6])
            print(df.head())

            # Función para obtener la tarifa correspondiente
            df_tarifa = xx.d00t20tdheatTarifa

            def obtener_tarifa(fila):
                # Para fines de semana, usar la columna 12 (SDF)
                mes = 12 if fila["es_fin_de_semana"] else fila["month_num"]
                hora = fila["hour"]  # Ya tenemos el número 0-23
                try:
                    return df_tarifa.iloc[hora][mes]
                except (KeyError, IndexError):
                    return None  # o algún valor por defecto

            # Aplicar la función para crear la nueva columna
            df["t20td"] = df.apply(obtener_tarifa, axis=1)

            # Función para obtener la tarifa correspondiente
            df_tarifa = xx.d00tdh6heatTarifa

            def obtener_tarifa(fila):
                # Para fines de semana, usar la columna 12 (SDF)
                mes = 12 if fila["es_fin_de_semana"] else fila["month_num"]
                hora = fila["hour"]  # Ya tenemos el número 0-23
                try:
                    return df_tarifa.iloc[hora][mes]
                except (KeyError, IndexError):
                    return None  # o algún valor por defecto

            # Aplicar la función para crear la nueva columna
            df["tdh6"] = df.apply(obtener_tarifa, axis=1)

            # df = add_tarifa_column(df, df_tarifa)
            setattr(xx, sys._getframe().f_code.co_name, df)  # Valores originales

            return pp, uu, vv, xx, yy, zz

        def d50barConsumoMensualTotal(pp, uu, vv, xx, yy, zz):
            """asignar columnas de semana mes hora, tarifas..."""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            df = xx.d30AsignarAgrupacionesDeTiempo.copy()
            #
            # Agrupamos por month_num (o month si prefieres el nombre) y tdh6, y sumamos los valores
            df_resultado = df.groupby(["month_num", "tdh6"])["c1"].sum().unstack()
            # Si quieres resetear el índice para tener month_num como columna en lugar de índice
            df_resultado = df_resultado.fillna(0)

            df_resultado.index.name = "kW/h_mes"
            df_resultado.index = df_resultado.index + 1

            #
            # df = add_tarifa_column(df, df_tarifa)
            setattr(
                xx, sys._getframe().f_code.co_name, df_resultado
            )  # Valores originales
            return pp, uu, vv, xx, yy, zz

        def d50barConsumoHorarioMedioDeLunesAViernes(pp, uu, vv, xx, yy, zz):
            """asignar columnas de semana mes hora, tarifas..."""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            df = xx.d30AsignarAgrupacionesDeTiempo.copy()
            df = df[df["day_of_week"] < 5]  # Lunes=0, Viernes=4

            df_resultado = (
                df.groupby(["hour", "tdh6"])["c1"].sum().unstack() / len(df) * 24
            )
            df_resultado = df_resultado.fillna(0)
            #
            # Si quieres resetear el índice para tener month_num como columna en lugar de índice
            df_resultado.index.name = "kW/h_hora"
            #
            # df = add_tarifa_column(df, df_tarifa)
            setattr(
                xx, sys._getframe().f_code.co_name, df_resultado
            )  # Valores originales
            return pp, uu, vv, xx, yy, zz

        def d50barConsumoHorarioMedioEnFinDeSemana(pp, uu, vv, xx, yy, zz):
            """asignar columnas de semana mes hora, tarifas..."""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            df = xx.d30AsignarAgrupacionesDeTiempo.copy()
            df = df[df["day_of_week"] > 4]  # Lunes=0, Viernes=4

            df_resultado = (
                df.groupby(["hour", "tdh6"])["c1"].sum().unstack() / len(df) * 24
            )
            df_resultado = df_resultado.fillna(0)
            #
            # Si quieres resetear el índice para tener month_num como columna en lugar de índice
            df_resultado.index.name = "kW/h_hora"
            #
            # df = add_tarifa_column(df, df_tarifa)
            setattr(
                xx, sys._getframe().f_code.co_name, df_resultado
            )  # Valores originales
            return pp, uu, vv, xx, yy, zz


########################################################################################################
########################################################################################################


class comun:
    class main:

        def main(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            # segun si es un formulario o un documento nuevo

            pp, uu, vv, xx, yy, zz = comun.main.datos(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.datos2latex(pp, uu, vv, xx, yy, zz)

            output_buffer = comun.main.replaceLatex(pp, uu, vv, xx, yy, zz)

            # print(uu)
            return output_buffer

        def datos(pp, uu, vv, xx, yy, zz):
            """datoes"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listadatoes = [
                m
                for m in dir(datos.datos)
                if callable(getattr(datos.datos, m)) and not m.startswith("__")
            ]
            listadatoes = [s for s in listadatoes if not s.startswith("__")]
            listadatoes = sorted(listadatoes)

            for dato in listadatoes:
                print(dato)
                metodo = getattr(datos.datos, dato, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def datos2latex(pp, uu, vv, xx, yy, zz):
            """tablas"""
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            listatablas = [
                m
                for m in dir(comun.datos2latex)
                if callable(getattr(comun.datos2latex, m)) and not m.startswith("__")
            ]
            listatablas = [s for s in listatablas if not s.startswith("__")]
            for tabla in listatablas:
                # print(tabla)
                metodo = getattr(
                    comun.datos2latex, tabla, None
                )  # Obtener el método o None si no existe
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)

            return pp, uu, vv, xx, yy, zz

        def replaceLatex(pp, uu, vv, xx, yy, zz):
            """
            Compila un archivo .tex a PDF, reemplaza un texto y devuelve el PDF en memoria.
            :param ruta_tex: Ruta al archivo .tex original.
            :return: Objeto BytesIO con el contenido del PDF.
            """

            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            pdf_path = Path(zz.nombre).with_suffix(".pdf")
            ruta_tex = f"{os.path.splitext(pdf_path)[0]}.tex"

            nombre_archivo = os.path.basename(ruta_tex)  # Nombre del archivo sin ruta

            # Crear un directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                ruta_tex_temp = os.path.join(temp_dir, nombre_archivo)

                contenido = file_tex(pp, uu, vv, xx, yy, zz)
                # tabla final con todas los parámetros  y variables
                HojaDeDatos = "\n"

                for var_name, value in pp.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\npp. {var_name}\n\npp.{var_name}"

                for var_name, value in uu.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nuu. {var_name}\n\nuu.{var_name}"

                for var_name, value in vv.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nvv. {var_name}\n\nvv.{var_name}"

                for var_name, value in xx.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nxx. {var_name}\n\nxx.{var_name}"

                for var_name, value in yy.__dict__.items():
                    HojaDeDatos = HojaDeDatos + f"\n\nyy. {var_name}\n\nyy.{var_name}"

                # noooooooo,!!!!! que el json del data da error en latex
                # for var_name, value in zz.__dict__.items():
                #     print(f"zz.{var_name}")
                #     HojaDeDatos = HojaDeDatos + f"\n\nzz. {var_name}\n\nzz.{var_name}"

                contenido = contenido.replace(f"HojaDeDatos", HojaDeDatos)

                if 1:
                    # Reemplazar directamente todas las variables en el texto
                    for var_name, value in pp.__dict__.items():
                        # print(f"pp.{var_name}")
                        contenido = contenido.replace(
                            f"pp.{var_name}",
                            f"{str(value)}",
                        )
                    for var_name, value in uu.__dict__.items():
                        # print(f"uu.{var_name}")
                        contenido = contenido.replace(f"uu.{var_name}", f"{str(value)}")

                    for var_name, value in vv.__dict__.items():
                        # print(f"vv.{var_name}")
                        contenido = contenido.replace(f"vv.{var_name}", f"{str(value)}")

                    for var_name, value in xx.__dict__.items():
                        # print(f"xx.{var_name}")
                        contenido = contenido.replace(f"xx.{var_name}", f"{str(value)}")

                    for var_name, value in yy.__dict__.items():
                        # print(f"yy.{var_name}")
                        contenido = contenido.replace(f"yy.{var_name}", f"{str(value)}")

                    for var_name, value in zz.__dict__.items():
                        # print(f"zz.{var_name}")
                        contenido = contenido.replace(f"zz.{var_name}", f"{str(value)}")

                with open(
                    "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.tex", "w", encoding="utf-8"
                ) as file:
                    file.write(contenido)

                # Guardar los cambios en el archivo temporal
                with open(ruta_tex_temp, "w", encoding="utf-8") as file:
                    file.write(contenido)

                # Función para ejecutar pdflatex dos veces
                def ejecutar_pdflatex():
                    print(
                        f"\n{100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                    )
                    if 1:
                        subprocess.run(
                            ["pdflatex", "-output-directory", temp_dir, ruta_tex_temp],
                            check=True,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )

                    return True

                # # para que se cre el indice debe estar descomentado
                # Ejecutar pdflatex dos veces para generar la TOC correctamente
                if not (ejecutar_pdflatex() and ejecutar_pdflatex()):
                    return None  # Si la compilación falla, se detiene el proceso
                # ejecutar_pdflatex()

                # Ruta del PDF generado
                pdf_generado = os.path.join(
                    temp_dir, nombre_archivo.replace(".tex", ".pdf")
                )

                # Leer el PDF en memoria
                with open(pdf_generado, "rb") as file:
                    pdf_en_memoria = BytesIO(file.read())

            return pdf_en_memoria

            #######################

    class datos2latex:

        def tabla(pp, uu, vv, xx, yy, zz):
            # tabla de todos los df

            for var_name, value in {**xx.__dict__, **yy.__dict__}.items():
                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                if var_name == zz.codigo:
                    titulo = "Parámetros"

                if isinstance(value, pd.DataFrame):
                    # para no hacer tablas de series muy
                    if len(value) > 24:
                        continue
                    # Redondear los valores del DataFrame a 1 decimal
                    value = value.round(1)
                    value.index.name = ""
                    latex_code = value.to_latex(float_format="%.1f")

                    latex_code_with_resizebox = rf"""
                    \begin{{table}}[H] \centering
                        % \resizebox{{0.4\textwidth}}{{!}}
                        {{
                        {latex_code}
                        }}
                        \caption{{{titulo}}}
                    \end{{table}}
                    """

                    # Asignar el LaTeX a la propiedad correspondiente de uu
                    setattr(uu, f"tabla{var_name}", latex_code_with_resizebox)

            # Convierte todos los df de uu a tablas
            return pp, uu, vv, xx, yy, zz

        def grafico(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            from matplotlib.colors import ListedColormap

            # tabla de todos los df

            # 4. Paleta de colores según la imagen (asumiendo colores estándar)
            colores_periodos = {
                1: "red",
                2: "#f79e05",
                3: "#ffd97a",
                4: "#fae7b9",
                5: "#b1c993",
                6: "#afc460",
            }
            # colores = [colores_periodos[p] for p in sorted(media_por_periodo.columns)]

            for var_name, value in {**xx.__dict__, **yy.__dict__}.items():
                print(var_name)
                if hasattr(xx, var_name):
                    df = getattr(xx, var_name)
                else:
                    df = getattr(yy, var_name)

                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                # 1. Crear un archivo temporal para la imagen
                with tempfile.NamedTemporaryFile(
                    suffix=".png", delete=False
                ) as temp_image:
                    temp_image_path2 = (
                        temp_image.name
                    )  # Guardar la ruta del archivo temporal

                # 📌 Gráfico del flujo de caja y acumulado
                if "heat" in var_name:
                    """"""
                    if "fmt" in var_name:
                        fmt = ".1f"
                    else:
                        fmt = "d"
                    fig, ax = plt.subplots(figsize=(5, len(df) * 0.3))
                    cmap = ListedColormap(
                        [colores_periodos[p] for p in sorted(colores_periodos.keys())]
                    )
                    try:
                        # 3. Generar heatmap
                        sns.heatmap(
                            df,
                            annot=True,
                            fmt=fmt,  # Formato de 1 decimal (cambia de "d" a ".1f")
                            # cmap="YlOrBr",  # Paleta amarillo-naranja-marrón
                            cmap=cmap,  # Paleta en escala de grises (negro=0, blanco=1)
                            cbar=False,
                            linewidths=0.3,
                            linecolor="lightgray",
                            annot_kws={
                                "size": 10,
                                "rotation": 0,
                            },  # Texto más pequeño y rotado 45°
                            ax=ax,
                        )  # Usar el axes creado

                        # Mover ticks del eje X a la parte superior
                        ax.xaxis.tick_top()  # <--- Esta es la línea clave
                        ax.xaxis.set_label_position(
                            "top"
                        )  # Opcional: si quieres la etiqueta arriba también

                        plt.xticks(rotation=0)
                        plt.yticks(rotation=0)
                        plt.tight_layout()  # Ajustar para que no se corten las etiquetas
                        # plt.subplots_adjust(left=0.051,bottom=0.15)  # Puedes ajustar estos valores

                    except:
                        """no valores numericos para plt"""

                    try:
                        ax.set_ylabel(df.index.name.split("_")[0])
                        ax.set_xlabel(df.index.name.split("_")[1])
                    except:
                        """"""

                elif "bar" in var_name:
                    fig, ax = plt.subplots(figsize=(5, 3))
                    try:
                        df.plot(kind="bar", stacked=True, color=colores_periodos, ax=ax)

                        plt.xticks(rotation=45)
                        plt.tight_layout()  # Ajustar para que no se corten las etiquetas
                    except:
                        """no valores numericos para plt"""

                    try:
                        ax.set_ylabel(df.index.name.split("_")[0])
                        ax.set_xlabel(df.index.name.split("_")[1])
                    except:
                        """"""
                    plt.tight_layout()
                    plt.grid(axis="y", linestyle="--", alpha=0.7)

                else:
                    fig, ax = plt.subplots(figsize=(6, 5))
                    try:
                        df.plot(ax=ax)
                    except:
                        """"""

                    # Obtener el nombre de la columna automáticamente
                    try:
                        ax.set_ylabel(df.index.name.split("_")[0])
                        ax.set_xlabel(df.index.name.split("_")[1])
                    except:
                        """"""

                    plt.grid(axis="y", linestyle="--", alpha=0.7)
                plt.savefig(temp_image_path2, format="png", dpi=300)
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
                latex_code = (
                    ""
                    + r"""\begin{figure}[H] \centering
                    """
                    + rf"""
                    \includegraphics[width=.5\textwidth]{{{temp_image_path2}}}
                    \caption{{{titulo}}}
                    """
                    + """
                    \label{fig:dfassssdfsa}
                    \end{figure}
                    """
                )
                setattr(uu, f"{var_name}", latex_code)
            return pp, uu, vv, xx, yy, zz

        def imagen(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpetaimagenes = Path(f"{zz.ruta_script}/{zz.codigo}/img")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.png")] + [
                f.name for f in carpetaimagenes.glob("*.jpg")
            ]

            for imagen in imagenes:
                # _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                titulo = imagen[:-4]
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional

                latex_code = (
                    ""
                    + r"""\begin{figure}[H] \centering
                    """
                    + rf"""
                    \includegraphics[width=0.7\textwidth]{{{Path(f"{zz.ruta_script}/{zz.codigo}/img/{imagen}")}}} 
                    \caption{{{titulo}}}
                    """
                    + """
                    \label{fig:dfasdfsa}
                    \end{figure}
                    """
                )

                setattr(uu, f"imagen{imagen}", latex_code)
            return pp, uu, vv, xx, yy, zz




        def mapaUbicacion(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            with tempfile.NamedTemporaryFile(
                suffix=".png", delete=False
            ) as temp_image:
                temp_image_path = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal


            import geopandas as gpd
            import matplotlib.pyplot as plt
            from shapely.geometry import Point
            from geopy.geocoders import Nominatim
            from geopy.extra.rate_limiter import RateLimiter
            # Cargar el GeoJSON
            geojson_path = f"{zz.ruta_script}/assets/spain-communities.geojson"
            spain_communities = gpd.read_file(geojson_path)
            spain_communities = spain_communities[spain_communities["name"] != "Canarias"]
            # Configurar geocodificador
            geolocator = Nominatim(user_agent="mapa_espana_detallado")
            geocode = RateLimiter(geolocator.reverse, min_delay_seconds=1)
            # Coordenadas del punto
            # lat, lng = 37.3883251804357, -4.508585281106158
            lat, lng = pp.CoordenadasValorlat, pp.CoordenadasValorlng

            point = Point(lng, lat)
            # Obtener dirección detallada
            try:
                location = geocode(f"{lat}, {lng}", language='es', exactly_one=True)
                if location:
                    address = location.raw['address']
                    calle = address.get('road', '')
                    numero = address.get('house_number', '')
                    ciudad = address.get('city', address.get('town', address.get('village', '')))
                    provincia = address.get('state', '')
                    codigo_postal = address.get('postcode', '')  # <- Aquí obtenemos el CP

                    direccion_completa = f"{calle} {numero} , {ciudad} {codigo_postal}, {provincia}"
                else:
                    direccion_completa = "Dirección no disponible"
            except Exception as e:
                direccion_completa = f"Error al obtener dirección: {str(e)}"
            # Convertir a proyección adecuada para España (UTM zona 30N)
            spain_communities = spain_communities.to_crs(epsg=25830)
            point_gdf = gpd.GeoSeries([point], crs="EPSG:4326").to_crs(epsg=25830)
            # Crear figura
            fig, ax = plt.subplots(figsize=(7, 7))
            # Título con dirección completa
            ax.set_title(f"{direccion_completa}", fontsize=14, pad=20)
            # Dibujar comunidades
            spain_communities.plot(ax=ax, color="lightgrey", edgecolor="black", linewidth=0.5)
            # Dibujar punto de ubicación
            point_gdf.plot(
                ax=ax,
                color="red",
                markersize=100,
                marker="o",
                edgecolor="black",
                label="Ubicación"
            )
            # Añadir coordenadas como texto (proyectadas de nuevo para mostrar en grados)
            ax.annotate(
                f"Coordenadas:\n{lat:.6f}° N\n{lng:.6f}° W",
                xy=point_gdf.geometry[0].coords[0],
                xytext=(15, 15),
                textcoords="offset points",
                bbox=dict(boxstyle="round", fc="white", ec="black", alpha=0.9),
                fontsize=10
            )

            # ax.legend(loc='upper right')
            # ax.grid(True, linestyle='--', alpha=0.5)
            plt.tight_layout()
            # plt.show()
            ax.set_axis_off()  # <- Esta línea elimina los ejes



            plt.savefig(temp_image_path, format="png", dpi=300)
            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + r"""\begin{figure}[H] \centering
                """
                + rf"""
                \includegraphics[width=.5\textwidth]{{{temp_image_path}}}
                \caption{{{'Ubicación'}}}
                """
                + """
                \label{fig:dfassssdfsa}
                \end{figure}
                """
            )
            setattr(uu, sys._getframe().f_code.co_name, latex_code)
            return pp, uu, vv, xx, yy, zz




    class texfile:

        def normalizar_cadena(texto):
            if isinstance(texto, str):  # Asegurar que es un string
                texto = texto.lower()
                texto = "".join(
                    c
                    for c in unicodedata.normalize("NFD", texto)
                    if unicodedata.category(c) != "Mn"
                )
                # Capitalizar la primera letra después de un espacio
                texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
                # Eliminar espacios
                texto = re.sub(r"\s+", "", texto)
                # Mantener solo caracteres permitidos
                texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
            return texto

        def limpiar_valores(data):
            for key, values in data.items():
                for i in range(
                    1, len(values)
                ):  # Saltamos la primera fila (encabezados)
                    for j in range(
                        1, len(values[i])
                    ):  # Recorremos todas las columnas (excepto encabezados)
                        if isinstance(
                            values[i][j], str
                        ):  # Verificar si es una cadena antes de dividir
                            valor_limpio = (
                                values[i][j].split(";")[0].strip()
                            )  # Tomar solo el primer valor y limpiar espacios

                            # Intentar convertir a flotante si es posible
                            try:
                                values[i][j] = float(valor_limpio)
                            except ValueError:
                                values[i][
                                    j
                                ] = valor_limpio  # Si no se puede convertir, dejar como está
            return data

        def separar_y_mayusculizar(texto):
            # Usamos una expresión regular para separar las mayúsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            # Separar la primera parte (mantenerla como está)
            primera_parte = partes[0]

            # El resto de las partes lo convertimos a mayúsculas y las unimos con espacios
            resto = " ".join([parte.upper() for parte in partes[1:]])
            # resto = " ".join([parte for parte in partes[1:]])

            return primera_parte, resto

        def separar_y_formatear(texto):
            # Usamos una expresión regular para separar las mayúsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            if not partes:
                return ""

            # Convertimos todo a minúsculas y capitalizamos solo la primera letra
            resultado = " ".join([parte.lower() for parte in partes])
            return resultado.capitalize()

        # Función para limpiar nombres de variables
        def sanitize_name(name):
            return re.sub(
                r"[^a-zA-Z0-9]", "", str(name)
            )  # Elimina cualquier carácter no alfanumérico


class especifico:
    class previo:
        def p00parametros():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            nombre_funcion = os.path.basename(__file__)[:-3]
            codigo, _ = comun.texfile.separar_y_mayusculizar(nombre_funcion)
            nan = " "
            print(codigo)
            zz.data = {
                "Codigo": [
                    ["Parametro", "Valor"],
                    ["Inversión inicial [€]", 0],
                    ["Bonificación  inicial [€]", 10000],
                    ["Duración de la actuación [años]", 15],
                    ["Préstamo recibido [€]", 100000],
                    ["Tasa de interés del préstamo [\\%]", 5],
                    ["Plazo del préstamo [años]", 15],
                    ["Inflación annual [\\%]", 3],
                    ["Ahorro annual [€]", 25000],
                    ["Costos fijos operativos [€]", 1000],
                    ["Tasa de descuento [\\%]", 6],
                ],
                "Coordenadas": [
                    ["Parámetro", "Valor"],
                    ["lat", 37.3883251804357],
                    ["lng", -4.508585281106158],
                ],
                "Participantes": [
                    ["Parámetro", "P1"],
                    ["Rol", nan],
                    ["DNI/NIF/NIE", nan],
                    [
                        "Nombre/Razón Social",
                        "'EDITAR' para poner tu nombre y ajustar los parámetros.",
                    ],
                    ["Titulación", nan],
                ],
                "consumo": [
                    ["Unnamed: 0", "c1"],
                    ["Año", 2017],
                    ["Mes", 1],
                    ["Día", 1],
                    [nan, 42],
                    [nan, 39],
                    [nan, 42],
                    [nan, 42],
                    [nan, 39],
                    [nan, 44],
                    [nan, 44],
                    [nan, 39],
                    [nan, 29],
                    [nan, 19],
                    [nan, 17],
                    [nan, 17],
                    [nan, 17],
                    [nan, 17],
                    [nan, 17],
                    [nan, 17],
                    [nan, 17],
                    [nan, 17],
                    [nan, 34],
                    [nan, 42],
                    [nan, 42],
                    [nan, 42],
                    [nan, 39],
                    [nan, 39],
                    [nan, 47],
                    [nan, 47],
                    [nan, 49],
                    [nan, 49],
                    [nan, 49],
                    [nan, 52],
                    [nan, 52],
                    [nan, 44],
                    [nan, 34],
                    [nan, 24],
                    [nan, 22],
                    [nan, 27],
                    [nan, 24],
                    [nan, 27],
                    [nan, 24],
                    [nan, 19],
                    [nan, 19],
                    [nan, 22],
                    [nan, 37],
                    [nan, 42],
                    [nan, 42],
                    [nan, 39],
                    [nan, 44],
                    [nan, 42],
                    [nan, 49],
                    [nan, 52],
                    [nan, 47],
                    [nan, 44],
                    [nan, 47],
                    [nan, 44],
                    [nan, 42],
                    [nan, 44],
                    [nan, 34],
                    [nan, 22],
                    [nan, 17],
                    [nan, 19],
                    [nan, 19],
                    [nan, 17],
                    [nan, 19],
                    [nan, 22],
                    [nan, 22],
                    [nan, 19],
                    [nan, 34],
                    [nan, 39],
                ],
            }

        def p01carpetaauxiliar():
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            carpeta = Path(f"{zz.ruta_script}/assets")

            if not os.path.exists(carpeta):
                os.mkdir(carpeta)

        def p02xlsxhojaconvalorespordefecto():
            """ """
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")
            # si es un nuevo archivo tex copia los parámetros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            codigo, titulo = comun.texfile.separar_y_mayusculizar(nombre)
            #
            ruta_archivo_excel = f"{zz.ruta_script}/{codigo}.xlsx"
            if 1:
                df = pd.DataFrame(zz.data, columns=["Parámetro", "Valor"])
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    # Crear un libro de trabajo de Excel
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja predeterminada creada por openpyxl
                    libro.remove(libro.active)

                    # Crear hojas con los nombres de los keys y escribir los datos
                    for sheet_name, sheet_data in zz.data.items():
                        hoja = libro.create_sheet(
                            title=sheet_name
                        )  # Crear la hoja con el nombre del key
                        for row in sheet_data:
                            hoja.append(row)  # Agregar cada fila a la hoja

                    # Guardar el archivo Excel
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    # print(ruta_archivo_excel)
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # Cargar todas las hojas del archivo Excel
            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )

            # Crear el diccionario data
            zz.data = {
                sheet_name: [df.columns.tolist()]
                + df.values.tolist()  # Incluir cabecera y datos
                for sheet_name, df in sheets.items()
            }


from fastapi.responses import StreamingResponse
from io import BytesIO
import os
import matplotlib.pyplot as plt
import pandas as pd
from types import SimpleNamespace
import subprocess
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata
import seaborn as sns


import numpy as np
import sys

if __name__ == "__main__":
    zz = SimpleNamespace()

    # Copio la ruta del scripot para crear el excel
    zz.ruta_script = os.path.dirname(os.path.abspath(__file__))
    zz.nombre = os.path.basename(__file__)[:-3]
    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)

    #
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    os.chdir("../../../../")  # Cambia dos niveles arriba

    # clase especifico
    metodos = [
        m
        for m in dir(especifico.previo)
        if callable(getattr(especifico.previo, m)) and not m.startswith("__")
    ]
    metodos = [s for s in metodos if not s.startswith("__")]
    metodos = sorted(metodos)

    for metodo in metodos:
        metodoi = getattr(especifico.previo, metodo, None)
        metodoi()

    # la funcion de fastpai
    globals()[os.path.basename(__file__)[:-3]](zz.data)
