#!/usr/bin/env python3

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from io import BytesIO
import os
import matplotlib.pyplot as plt
import pandas as pd
from types import SimpleNamespace
import subprocess
import tempfile
import re
import numpy_financial as npf
import json
from pathlib import Path
import openpyxl
import unicodedata


import numpy as np
import sys


router = APIRouter()


@router.post(f"/Acaf000")
def Acaf000AnálisisFinanciero(
    data: dict,
):
    data = comun.texfile.limpiar_valores(data)
    print(data)
    # parámetros
    pp = SimpleNamespace()
    # tablas y graficos en latex
    uu = SimpleNamespace()
    #
    vv = SimpleNamespace()
    xx = SimpleNamespace()
    yy = SimpleNamespace()
    # datos genericos
    zz = SimpleNamespace()

    zz.nombre = sys._getframe().f_code.co_name
    zz.ruta_script = os.path.dirname(
        os.path.abspath(__file__)
    )  # Obtener la carpeta sin el nombre del script

    zz.codigo, zz.titulo = comun.texfile.separar_y_mayusculizar(zz.nombre)
    zz.data = data

    output_buffer = comun.main.main(pp, uu, vv, xx, yy, zz)

    with open(f"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.pdf", "wb") as f:
        f.write(output_buffer.getvalue())

    return StreamingResponse(
        output_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=CAE_Agrario_Completo.pdf"
        },
    )


def file_tex(pp, uu, vv, xx, yy, zz):

    return r"""

\documentclass[a4paper,10pt]{article}
%\documentclass[a4paper,10pt,twocolumn]{article}
\usepackage[top = 2cm, bottom = 2cm, left = 2cm, right = 2cm, asymmetric]{geometry} % Especificar los márgenes según la norma
\usepackage[spanish]{babel}    

\usepackage{graphicx}  % Para incluir imágenes si es necesario
\usepackage{amsmath, amssymb}  % Para expresiones matemáticas
\usepackage{fancyhdr}  % Para personalizar encabezados
\usepackage{chngcntr}  % Para cambiar la numeración de apartados y subsecciones
\usepackage{tocbibind}  % Para incluir bibliografía en la tabla de contenidos
\usepackage{appendix}  % Para el formato de anexos
\usepackage{lipsum}  % Para generar texto de relleno (dummy text)
\usepackage{geometry}  % Para personalizar los márgenes
\usepackage{multicol}  % Para columnas
\usepackage{titlesec}  % Para personalizar los títulos de las secciones
\usepackage[utf8]{inputenc}    
\usepackage[T1]{fontenc}       
\usepackage[pdftex,pdfencoding=auto, colorlinks=true, linkcolor=black, citecolor=black, filecolor=black, urlcolor=black]{hyperref}
\usepackage{tocloft}           
\usepackage{booktabs}          
\usepackage{float} % Paquete necesario para usar la opción [H]
\usepackage{array}
\usepackage{longtable}
\usepackage{circuitikz}
\usepackage{tikz}
\usepackage{tikz-cd}

\usepackage{tikz-qtree}
\usepackage{smartdiagram}

\usepackage{qrcode}            % Paquete para generar QR
\usepackage{tabularx} % Agregar en el preámbulo
\usepackage[absolute,overlay]{textpos} % Para posicionar objetos libremente
\usepackage{underscore}

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{zz.titulo}  % Nombre del documento en el encabezado izquierdo
\fancyhead[C]{}  % Centro vacío
\fancyhead[R]{\thepage}  % Numeración de páginas en el encabezado derecho
% \renewcommand{\thepage}{\arabic{page}}  % Asegura la numeración de páginas en números arábigos

% Definición de los anexos
\newcommand{\annex}[1]{\section{Anexo #1} \addcontentsline{toc}{section}{Anexo #1}}
\addto\captionsspanish{%
  \renewcommand{\tablename}{Tabla}%
  \renewcommand{\listtablename}{Índice de tablas}%
}
\title{{\small Ref.:\uppercase{zz.codigo}}\\{\textbf{zz.titulo}}}
\author{
pp.ParticipantesP1NombreRaznSocial
}
\date{\today}





\newcommand{\xxMostrarVariablesAlFinal}{}  

\begin{document}
\begin{Form}
%% --- PORTADA ---
%\begin{titlepage}
%    \centering
%    {\Large \textbf{TÍTULO DEL INFORME}}\\[1.5cm]
%    %\includegraphics[width=0.4\textwidth]{logo.png}\\[1cm] % Logo de la empresa/institución
%    
%    \textbf{Autor:} Nombre del Autor \\[0.5cm]
%    \textbf{Fecha:} \today \\[1cm]
%    \textbf{Código del documento:} REF-XXXX-YYYY \\[1cm]
%    \textbf{Versión:} 1.0 \\[1cm]
%    \textbf{Empresa / Institución:} Nombre \\[2cm]
%    \vfill
%\end{titlepage}
	\maketitle
    %
	%\begin{abstract}

    %\end{abstract}
% --- ÍNDICE ---
	\tableofcontents  % Tabla de contenidos
	\listoffigures    % Lista de graficos
	\listoftables     % Lista de tablas
	\newpage



    











	\section{Introducción}
	Este informe presenta un análisis financiero detallado,  con el objetivo de evaluar la viabilidad y rentabilidad del proyecto.

	\section{Objetivos del Análisis}
	\begin{itemize}
		\item Evaluar la viabilidad económica del proyecto.
		\item Determinar los principales indicadores financieros.
		\item Identificar riesgos y oportunidades.
	\end{itemize}


	\section{Metodología}
	Se emplean técnicas de análisis financiero conforme al cálculo del VAN (Valor Actual Neto), TIR (Tasa Interna de Retorno) y PR (Período de Retorno).

	\section{Parámetros}

   uu.dfAcaf000


	\section{Resultados}

    uu.TablaResultados

	\section{Flujo de Caja}

    uu.f01figuraFlujoDeCaja
    
	uu.TablaFlujoDeCaja


	\newpage
	\section{Indicadores Financieros}

	\subsection{Período de Recuperación de la Inversión (PR)}

	El PR es el tiempo necesario para recuperar la inversión inicial:

	\[
		PR = \frac{I_0}{B_t}
	\]

	Donde:

	\begin{center}
		\begin{tabular*}{0.4\textwidth}{@{\extracolsep{\fill}} p{0.025\textwidth} p{0.35\textwidth} p{0.025\textwidth} @{}}
			\( I_0 \) & Es la inversión inicial & € \\
			\( B_t \) & Es el beneficio neto en el período \( t \) & € \\
		\end{tabular*}
	\end{center}

	\subsection{Valor Actual Neto (VAN)}

	El VAN calcula el valor presente de los flujos de caja futuros descontados, menos la inversión inicial. Representa el valor que la inversión agrega en términos monetarios.

	\[
		VAN = \sum_{t=0}^{n} \frac{B_t}{(1+r)^t} - I_0
	\]

	Donde:

	\begin{center}
		\begin{tabular*}{0.4\textwidth}{@{\extracolsep{\fill}} p{0.025\textwidth} p{0.35\textwidth} p{0.025\textwidth} @{}}
			\( B_t \) & Beneficio neto en el período \( t \) &€ \\
			\( r \) & Tasa de descuento & \% \\
			\( n \) & Número total de períodos & - \\
			\( I_0 \) & Inversión inicial & € \\
		\end{tabular*}
	\end{center}



	\subsection{Tasa Interna de Retorno (TIR)}

	La Tasa Interna de Retorno (TIR) es la tasa de descuento que hace que el VAN sea igual a cero. Se calcula resolviendo la siguiente ecuación:

	\[
		\sum_{t=0}^{n} \frac{B_t}{(1+TIR)^t} = I_0
	\]

	Donde:
	\begin{center}
		\begin{tabular*}{0.4\textwidth}{@{\extracolsep{\fill}} p{0.025\textwidth} p{0.35\textwidth} p{0.025\textwidth} @{}}
			\( B_t \) & Es el beneficio neto en el período \( t \) & € \\
			\( TIR \) & Es la tasa interna de retorno & - \\
			\( n \) & Es el número total de períodos & - \\
			\( I_0 \) & Es la inversión inicial & € \\
		\end{tabular*}
	\end{center}



	\section{Análisis de Indicadores }


	\begin{enumerate}

		\item \textbf{Período de Recuperación (PR o Payback Period)}:
		      \begin{itemize}
			      \item Debe ser menor que la vida útil esperada del proyecto.
			      \item Mientras más corto sea el PR, más rápida será la recuperación del capital invertido, lo que reduce el riesgo.
			            Lo ideal es que el PR sea menos de la mitad de duración indicativa de la actuación.
		      \end{itemize}
		\item \textbf{Valor Actual Neto (VAN)}:
		      \begin{itemize}
			      \item Debe ser mayor a 0, es decir, \( VAN > 0 \).
			      \item Un VAN positivo indica que la inversión generará un valor adicional después de descontar los costos. Si es negativo, significa que no se recupera la inversión con la rentabilidad esperada.
		      \end{itemize}

		\item \textbf{Tasa Interna de Retorno (TIR)}:
		      \begin{itemize}
			      \item Debe ser mayor que la tasa de descuento o costo de oportunidad del capital, es decir, \( TIR > r \), donde \( r \) es la tasa de descuento.
			      \item Si la TIR es superior a la tasa mínima de rendimiento aceptable, la inversión es viable.
		      \end{itemize}


	\end{enumerate}

	En resumen, un proyecto es rentable si cumple las siguientes condiciones:


	\begin{table}[h]
		\centering
		\begin{tabular}{|ccc|}
			\hline
			VAN & > & 0    \\
			\hline
			TIR & > & r    \\
			\hline
			PR  & < & Di/2 \\
			\hline
		\end{tabular}
		\label{tab:ejemplo}
	\end{table}





    



















    













% --- 7. REFERENCIAS BIBLIOGRÁFICAS ---
\begin{thebibliography}{99}
    \bibitem{ccee} \href{file:///home/pk/Downloads/5ValoraruPproxectodeInvestimento_cas.pdf}{Cómo valorar un proyecto de inversión.C.E.E.I GALICIA, S.A. (BIC GALICIA)}
\end{thebibliography}





































































































\ifdefined\MostrarVariablesAlFinal
\newpage
\onecolumn
HojaDeDatos
\fi
\end{Form}
\end{document}

"""


def parámetros():

    nombre_funcion = os.path.basename(__file__)[:-3]
    codigo, _ = comun.texfile.separar_y_mayusculizar(nombre_funcion)
    nan = " "
    print(codigo)
    return {
        "Acaf000": [
            ["Parametro", "Valor"],
            ["Inversión inicial [€]", 0],
            ["Bonificación  inicial [€]", 10000],
            ["Duración de la actuación [años]", 15],
            ["Préstamo recibido [€]", 100000],
            ["Tasa de interés del préstamo [\\%]", 5],
            ["Plazo del préstamo [años]", 15],
            ["Inflación annual [\\%]", 3],
            ["Ahorro annual [€]", 25000],
            ["Costos fijos operativos [€]", 1000],
            ["Tasa de descuento [\\%]", 6],
        ],
        "Coordenadas": [
            ["Parámetro", "Valor"],
            ["lat", 36.6641681434112],
            ["lng", -4.45863940137516],
        ],
        "Participantes": [
            ["Parámetro", "P1", "P2", "P3"],
            ["Rol", nan, nan, nan],
            ["DNI/NIF/NIE", nan, nan, nan],
            [
                "Nombre/Razón Social",
                "'EDITAR' para poner tu nombre y ajustar los parámetros.",
                nan,
                nan,
            ],
            ["Titulación", nan, nan, nan],
            ["Dirección", nan, nan, nan],
            ["Población", nan, nan, nan],
            ["Teléfono", nan, nan, nan],
            ["Correo Electrónico", nan, nan, nan],
        ],
    }


class codigo_para_replace_en_el_latex:

    class calculo:

        def C01CalculoFinanciero(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'} {__class__.__name__}.{sys._getframe().f_code.co_name}")

            # # Datos iniciales (manteniendo nombres originales)
            # Inversininicial = 0
            # Bonificacininicial = 10000
            # Duracindelaactuacinaos = 25
            # Prstamorecibido = 100000
            # Tasadeintersdelprstamo = 5  # 5%
            # Plazodelprstamoaos = 10
            # Inflacinannual = 3  # 3%
            # Ahorroannual = 25000
            # Costosfijosoperativos = 1000
            # Tasadedescuento = 6  # 6%

            # Datos iniciales (manteniendo nombres originales)
            Inversininicial = pp.Inversininicial
            Bonificacininicial = pp.Bonificacininicial
            Duracindelaactuacinaos = int(pp.Duracindelaactuacinaos)
            Prstamorecibido = pp.Prstamorecibido
            Tasadeintersdelprstamo = pp.Tasadeintersdelprstamo
            Plazodelprstamoaos = pp.Plazodelprstamoaos
            Inflacinannual = pp.Inflacinannual
            Ahorroannual = pp.Ahorroannual
            Costosfijosoperativos = pp.Costosfijosoperativos
            Tasadedescuento = pp.Tasadedescuento

            # Cálculos iniciales
            Tasadeintersdecimal = Tasadeintersdelprstamo / 100
            Tasadedescuentodecimal = Tasadedescuento / 100
            Inflacindecimal = Inflacinannual / 100

            # Cálculo de la cuota anual del préstamo (sistema francés)
            factor = (1 + Tasadeintersdecimal) ** Plazodelprstamoaos
            Cuotaanualprestamo = (
                Prstamorecibido * (Tasadeintersdecimal * factor) / (factor - 1)
            )

            # Inicialización de listas para el DataFrame
            años = list(range(0, Duracindelaactuacinaos + 1))
            ingresos = [0]  # Año 0 no tiene ingresos
            gastos_fijos = [0]  # Año 0 no tiene gastos fijos
            pagos_prestamo = [0]  # Año 0 no tiene pagos de préstamo
            amortizaciones = [0]  # Año 0 no tiene amortización
            intereses_pagos = [0]  # Año 0 no tiene intereses
            bonificaciones = [Bonificacininicial] + [
                0
            ] * Duracindelaactuacinaos  # Bonificación solo en año 0
            flujos_caja = []
            flujo_acumulado = 0

            # Variables para cálculos
            saldoprestamo = Prstamorecibido
            PR_alcanzado = False
            PR_ano = None

            # Año 0 (inversión inicial con préstamo como salida de caja)
            flujo_ano0 = -Inversininicial - Prstamorecibido + Bonificacininicial
            flujos_caja.append(flujo_ano0)
            flujo_acumulado += flujo_ano0

            # Años 1 al Duracindelaactuacinaos
            for ano in range(1, Duracindelaactuacinaos + 1):
                # Cálculo de intereses del préstamo
                intereses = (
                    saldoprestamo * Tasadeintersdecimal
                    if ano <= Plazodelprstamoaos
                    else 0
                )

                # Amortización del préstamo
                amortizacion = (
                    Cuotaanualprestamo - intereses if ano <= Plazodelprstamoaos else 0
                )
                saldoprestamo -= amortizacion

                # Cálculo de ahorro ajustado por inflación (ingresos)
                ahorroajustado = Ahorroannual * (1 + Inflacindecimal) ** (ano - 1)

                # Flujo de caja anual
                flujo_ano = ahorroajustado - Costosfijosoperativos - Cuotaanualprestamo

                # Almacenar datos para el DataFrame
                ingresos.append(ahorroajustado)
                gastos_fijos.append(Costosfijosoperativos)
                pagos_prestamo.append(
                    Cuotaanualprestamo if ano <= Plazodelprstamoaos else 0
                )
                amortizaciones.append(amortizacion)
                intereses_pagos.append(intereses)
                flujos_caja.append(flujo_ano)

                # Cálculo acumulado para PR (Payback Period)
                flujo_acumulado += flujo_ano
                if not PR_alcanzado and flujo_acumulado >= 0:
                    PR_alcanzado = True
                    PR_ano = ano
                    # Cálculo exacto del PR considerando el flujo dentro del año
                    if ano > 1:
                        flujo_previo = flujo_acumulado - flujo_ano
                        PR_ano = (ano - 1) + (-flujo_previo / flujo_ano)

            # Cálculo VAN y TIR
            VAN = npf.npv(Tasadedescuentodecimal, flujos_caja)
            TIR = npf.irr(flujos_caja) * 100  # En porcentaje

            # Crear DataFrame con el flujo de caja detallado
            df_flujocaja = pd.DataFrame(
                {
                    "Año": años,
                    "Ingresos": ingresos,
                    "G. Fijos": gastos_fijos,
                    "G. Prést.": pagos_prestamo,
                    "Amortización": amortizaciones,
                    "Intereses": intereses_pagos,
                    "Bonificación": bonificaciones[: len(años)],
                    "F. Caja": flujos_caja,
                    "F. Acum.": np.cumsum(flujos_caja),
                }
            )

            # Crear DataFrame con los resultados principales
            df_resultados = pd.DataFrame(
                {
                    "Métrica": [
                        "Período de Recuperación (PR)",
                        "Valor Actual Neto (VAN)",
                        "Tasa Interna de Retorno (TIR)",
                    ],
                    "Valor": [
                        f"{PR_ano:.2f} años" if PR_ano is not None else "No recuperado",
                        f"{VAN:,.2f}",
                        f"{TIR:.2f}\%",
                    ],
                    "Descripción": [
                        "Tiempo requerido para recuperar la inversión inicial",
                        "Valor presente de los flujos de caja futuros",
                        "Tasa de descuento que hace el VAN igual a cero",
                    ],
                }
            )

            # Ajustar formato de números en el DataFrame
            pd.options.display.float_format = "{:,.2f}".format

            # Mostrar resultados
            print("RESULTADOS PRINCIPALES:")
            print(df_resultados.to_string(index=False))
            print("\nDETALLE DEL FLUJO DE CAJA:")
            print(df_flujocaja.to_string(index=False))

            # Opcional: Guardar en archivos Excel
            # df_resultados.to_excel("resultados_financieros.xlsx", sheet_name="Resultados", index=False)
            # df_flujocaja.to_excel("resultados_financieros.xlsx", sheet_name="Flujo de Caja", index=False)

            df_flujocaja.set_index("Año", inplace=True)
            pp.TablaFlujoDeCaja = df_flujocaja.copy()

            df_resultados.set_index("Métrica", inplace=True)
            pp.TablaResultados = df_resultados.copy()

            return pp, uu, vv, xx, yy, zz

    class tabla:

        def T00df2tablas(pp, uu, vv, xx, yy, zz):
            # tabla de todos los df
            for var_name, value in pp.__dict__.items():
                # Verificar si la variable es un DataFrame
                titulo = comun.texfile.separar_y_formatear(var_name)
                if var_name == zz.codigo:
                    titulo = "Parámetros"

                if isinstance(value, pd.DataFrame):
                    # Redondear los valores del DataFrame a 1 decimal
                    value = value.round(1)

                    latex_code = value.to_latex(float_format="%.1f")

                    latex_code_with_resizebox = f"""
                    \\begin{{table}}[h!] \\centering
                        % \\resizebox{{0.4\\textwidth}}{{!}}
                        {{
                        {latex_code}
                        }}
                        \\caption{{{titulo}}}
                    \\end{{table}}
                    """

                    # Asignar el LaTeX a la propiedad correspondiente de uu
                    setattr(uu, f"{var_name}", latex_code_with_resizebox)

            # Convierte todos los df de uu a tablas
            return pp, uu, vv, xx, yy, zz

    class grafico:

        def f01figuraFlujoDeCaja(pp, uu, vv, xx, yy, zz):
            print(f"\n {100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}")

            titulo = sys._getframe().f_code.co_name
            titulo = comun.texfile.separar_y_formatear(titulo)
            # 1. Crear un archivo temporal para la imagen
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image:
                temp_image_path1 = (
                    temp_image.name
                )  # Guardar la ruta del archivo temporal

            # 📌 Gráfico del flujo de caja y acumulado
            fig, ax = plt.subplots(figsize=(6, 5))

            df = pp.TablaFlujoDeCaja.copy()

            # Graficar el DataFrame en el eje ax
            df["F. Caja"].plot(kind="bar", ax=ax, legend=True)
            df["F. Acum."].plot(
                kind="line", ax=ax, legend=True, marker="o", color=["green"]
            )
            ax.grid(axis="y", linestyle="--", alpha=0.7)
            ax.set_title("Flujo de Caja y Acumulado")
            plt.xticks(rotation=50)
            ax.set_ylabel("€")  # Etiqueta en el eje y

            plt.savefig(temp_image_path1, format="png", dpi=300)
            # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional
            latex_code = (
                ""
                + """\\begin{figure}[H] \\centering
                """
                + f"""
                \\includegraphics[width=.5\\textwidth]{{{temp_image_path1}}}
                \\caption{{{titulo}}}
                """
                + """
                \\label{fig:dfassssdfsa}
                \\end{figure}
                """
            )
            setattr(uu, f"{sys._getframe().f_code.co_name}", latex_code)
            return pp, uu, vv, xx, yy, zz

    class imagen:

        def TodasLasImagnesDeLaCarpetaImg(pp, uu, vv, xx, yy, zz):
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            carpetaimagenes = Path(f"{zz.ruta_script}/{zz.codigo}/img")

            if not os.path.isdir(carpetaimagenes):
                return pp, uu, vv, xx, yy, zz
            imagenes = [f.name for f in carpetaimagenes.glob("*.png")] + [
                f.name for f in carpetaimagenes.glob("*.jpg")
            ]

            for imagen in imagenes:
                # _, titulo = comun.texfile.separar_y_mayusculizar(imagen[:-4])
                titulo = imagen[:-4]
                # Envolvemos el código generado con el entorno 'tabularx' para usar el ancho proporcional

                latex_code = (
                    ""
                    + """\\begin{figure}[H] \\centering
                    """
                    + f"""
                    \\includegraphics[width=0.7\\textwidth]{{{Path(f"{zz.ruta_script}/img/{imagen}")}}} 
                    \\caption{{{titulo}}}
                    """
                    + """
                    \\label{fig:dfasdfsa}
                    \\end{figure}
                    """
                )

                setattr(uu, f"imagen{imagen}", latex_code)
            return pp, uu, vv, xx, yy, zz


########################################################################################################
########################################################################################################


class comun:
    class main:

        def main(pp, uu, vv, xx, yy, zz):
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            # segun si es un formulario o un documento nuevo

            pp, uu, vv, xx, yy, zz = comun.main.data2parametros(pp, uu, vv, xx, yy, zz)
            print(pp)
            pp, uu, vv, xx, yy, zz = comun.main.calulos(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.tablas(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.graficos(pp, uu, vv, xx, yy, zz)
            pp, uu, vv, xx, yy, zz = comun.main.imagenes(pp, uu, vv, xx, yy, zz)

            output_buffer = comun.main.replaceLatex(pp, uu, vv, xx, yy, zz)

            # print(uu)
            return output_buffer

        def data2parametros(pp, uu, vv, xx, yy, zz):
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            # Para cada hoja en el JSON
            # Convertir cada valor del diccionario en un DataFrame
            for key, value in zz.data.items():
                key_sanitized = comun.texfile.sanitize_name(
                    key
                )  # Limpiar el nombre de la clave
                df = pd.DataFrame(
                    value[1:], columns=value[0]
                )  # Primera fila como encabezado
                df.set_index(
                    df.columns[0], inplace=True
                )  # Usar la primera columna como índice
                setattr(pp, f"df{key_sanitized}", df)  # Guardar en ss con nombre limpio
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == zz.codigo:
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar índice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{idx_sanitized}",
                                val,
                            )
                # para tener en latex los valores individuales de algunas hojas
                if key_sanitized == "Coordenadas" or key_sanitized == "Participantes":
                    # crea variables por columna y por cada item si la hoja es la del codigo del documento
                    # Crear variables por columna
                    for col in df.columns:
                        col_sanitized = comun.texfile.sanitize_name(
                            col
                        )  # Limpiar nombre de columna
                        # setattr(
                        #     pp, f"dc{key_sanitized}{col_sanitized}", df[[col]]
                        # )  # DataFrame con solo esa columna

                        # Crear variables individuales por celda
                        for idx, val in df[col].items():
                            idx_sanitized = comun.texfile.sanitize_name(
                                idx
                            )  # Limpiar índice (fila)
                            if pd.isna(val):
                                val = "."
                            setattr(
                                pp,
                                f"{key_sanitized}{col_sanitized}{idx_sanitized}",
                                val,
                            )

            return pp, uu, vv, xx, yy, zz

        def calulos(pp, uu, vv, xx, yy, zz):
            """calculos, valores calculados en yy"""
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            listacalculos = [
                m
                for m in dir(codigo_para_replace_en_el_latex.calculo)
                if callable(getattr(codigo_para_replace_en_el_latex.calculo, m))
                and not m.startswith("__")
            ]
            listacalculos = [s for s in listacalculos if not s.startswith("__")]
            for calculo in sorted(listacalculos):
                metodo = getattr(codigo_para_replace_en_el_latex.calculo, calculo, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)

            return pp, uu, vv, xx, yy, zz

        def graficos(pp, uu, vv, xx, yy, zz):
            """graficos"""
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            listafiguras = [
                m
                for m in dir(codigo_para_replace_en_el_latex.grafico)
                if callable(getattr(codigo_para_replace_en_el_latex.grafico, m))
                and not m.startswith("__")
            ]
            listafiguras = [s for s in listafiguras if not s.startswith("__")]
            for grafico in listafiguras:
                metodo = getattr(codigo_para_replace_en_el_latex.grafico, grafico, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def imagenes(pp, uu, vv, xx, yy, zz):
            """imagenes"""
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            listaimagenes = [
                m
                for m in dir(codigo_para_replace_en_el_latex.imagen)
                if callable(getattr(codigo_para_replace_en_el_latex.imagen, m))
                and not m.startswith("__")
            ]
            listaimagenes = [s for s in listaimagenes if not s.startswith("__")]

            for imagen in listaimagenes:
                metodo = getattr(codigo_para_replace_en_el_latex.imagen, imagen, None)
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)
            return pp, uu, vv, xx, yy, zz

        def tablas(pp, uu, vv, xx, yy, zz):
            """tablas"""
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            listatablas = [
                m
                for m in dir(codigo_para_replace_en_el_latex.tabla)
                if callable(getattr(codigo_para_replace_en_el_latex.tabla, m))
                and not m.startswith("__")
            ]
            listatablas = [s for s in listatablas if not s.startswith("__")]
            for tabla in listatablas:
                # print(tabla)
                metodo = getattr(
                    codigo_para_replace_en_el_latex.tabla, tabla, None
                )  # Obtener el método o None si no existe
                pp, uu, vv, xx, yy, zz = metodo(pp, uu, vv, xx, yy, zz)

            return pp, uu, vv, xx, yy, zz

        def replaceLatex(pp, uu, vv, xx, yy, zz):
            """
            Compila un archivo .tex a PDF, reemplaza un texto y devuelve el PDF en memoria.
            :param ruta_tex: Ruta al archivo .tex original.
            :return: Objeto BytesIO con el contenido del PDF.
            """

            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )

            pdf_path = Path(zz.nombre).with_suffix(".pdf")
            ruta_tex = f"{os.path.splitext(pdf_path)[0]}.tex"

            nombre_archivo = os.path.basename(ruta_tex)  # Nombre del archivo sin ruta

            # Crear un directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                ruta_tex_temp = os.path.join(temp_dir, nombre_archivo)

                contenido = file_tex(pp, uu, vv, xx, yy, zz)
                # tabla final con todas los parámetros  y variables
                HojaDeDatos = "\n"

                for var_name, value in pp.__dict__.items():
                    print(f"pp.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\npp. {var_name}\n\npp.{var_name}"

                for var_name, value in uu.__dict__.items():
                    print(f"uu.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\nuu. {var_name}\n\nuu.{var_name}"

                for var_name, value in vv.__dict__.items():
                    print(f"vv.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\nvv. {var_name}\n\nvv.{var_name}"

                for var_name, value in xx.__dict__.items():
                    print(f"xx.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\nxx. {var_name}\n\nxx.{var_name}"

                for var_name, value in yy.__dict__.items():
                    print(f"yy.{var_name}")
                    HojaDeDatos = HojaDeDatos + f"\n\nyy. {var_name}\n\nyy.{var_name}"

                # for var_name, value in zz.__dict__.items():
                #     print(f"zz.{var_name}")
                #     HojaDeDatos = HojaDeDatos + f"\n\nzz. {var_name}\n\nzz.{var_name}"

                contenido = contenido.replace(f"HojaDeDatos", HojaDeDatos)

                if 1:
                    # Reemplazar directamente todas las variables en el texto
                    for var_name, value in pp.__dict__.items():
                        # print(f"pp.{var_name}")
                        contenido = contenido.replace(
                            f"pp.{var_name}",
                            f"{str(value)}",
                        )
                    for var_name, value in uu.__dict__.items():
                        # print(f"uu.{var_name}")
                        contenido = contenido.replace(f"uu.{var_name}", f"{str(value)}")

                    for var_name, value in vv.__dict__.items():
                        # print(f"vv.{var_name}")
                        contenido = contenido.replace(f"vv.{var_name}", f"{str(value)}")

                    for var_name, value in xx.__dict__.items():
                        # print(f"xx.{var_name}")
                        contenido = contenido.replace(f"xx.{var_name}", f"{str(value)}")

                    for var_name, value in yy.__dict__.items():
                        # print(f"yy.{var_name}")
                        contenido = contenido.replace(f"yy.{var_name}", f"{str(value)}")

                    for var_name, value in zz.__dict__.items():
                        # print(f"zz.{var_name}")
                        contenido = contenido.replace(f"zz.{var_name}", f"{str(value)}")

                with open(
                    "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaa.tex", "w", encoding="utf-8"
                ) as file:
                    file.write(contenido)

                # Guardar los cambios en el archivo temporal
                with open(ruta_tex_temp, "w", encoding="utf-8") as file:
                    file.write(contenido)

                # Función para ejecutar pdflatex dos veces
                def ejecutar_pdflatex():
                    print(
                        f"\n{100*'+'}{ __class__.__name__}.{sys._getframe().f_code.co_name}"
                    )
                    if 1:
                        subprocess.run(
                            ["pdflatex", "-output-directory", temp_dir, ruta_tex_temp],
                            check=True,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )

                    return True

                # # para que se cre el indice debe estar descomentado
                # Ejecutar pdflatex dos veces para generar la TOC correctamente
                if not (ejecutar_pdflatex() and ejecutar_pdflatex()):
                    return None  # Si la compilación falla, se detiene el proceso
                # ejecutar_pdflatex()

                # Ruta del PDF generado
                pdf_generado = os.path.join(
                    temp_dir, nombre_archivo.replace(".tex", ".pdf")
                )

                # Leer el PDF en memoria
                with open(pdf_generado, "rb") as file:
                    pdf_en_memoria = BytesIO(file.read())

            return pdf_en_memoria

            #######################

        def xlsxhojaconvalorespordefecto(ruta_script):
            """ """
            print(
                f"\n {100*'+'} { __class__.__name__}.{sys._getframe().f_code.co_name}"
            )
            # si es un nuevo archivo tex copia los parámetros actualizados en el excel
            nombre = os.path.basename(__file__)[:-3]
            codigo, titulo = comun.texfile.separar_y_mayusculizar(nombre)
            #
            ruta_archivo_excel = f"{ruta_script}/{codigo}.xlsx"
            if 1:
                df = pd.DataFrame(parámetros(), columns=["Parámetro", "Valor"])
                # Verificar si el archivo existe
                if not os.path.exists(ruta_archivo_excel):
                    # Crear un nuevo libro de trabajo
                    libro = openpyxl.Workbook()

                    data = parámetros()

                    # Crear un libro de trabajo de Excel
                    libro = openpyxl.Workbook()

                    # Eliminar la hoja predeterminada creada por openpyxl
                    libro.remove(libro.active)

                    # Crear hojas con los nombres de los keys y escribir los datos
                    for sheet_name, sheet_data in data.items():
                        hoja = libro.create_sheet(
                            title=sheet_name
                        )  # Crear la hoja con el nombre del key
                        for row in sheet_data:
                            hoja.append(row)  # Agregar cada fila a la hoja

                    # Guardar el archivo Excel
                    libro.save(ruta_archivo_excel)

                else:
                    # Cargar el archivo Excel existente
                    # print(ruta_archivo_excel)
                    libro = openpyxl.load_workbook(ruta_archivo_excel)

            # Cargar todas las hojas del archivo Excel
            sheets = pd.read_excel(
                ruta_archivo_excel, sheet_name=None, engine="openpyxl", header=0
            )

            # Crear el diccionario data
            data = {
                sheet_name: [df.columns.tolist()]
                + df.values.tolist()  # Incluir cabecera y datos
                for sheet_name, df in sheets.items()
            }

            return data

    class texfile:

        def normalizar_cadena(texto):
            if isinstance(texto, str):  # Asegurar que es un string
                texto = texto.lower()
                texto = "".join(
                    c
                    for c in unicodedata.normalize("NFD", texto)
                    if unicodedata.category(c) != "Mn"
                )
                # Capitalizar la primera letra después de un espacio
                texto = re.sub(r"(\s+)([a-z])", lambda m: m.group(2).upper(), texto)
                # Eliminar espacios
                texto = re.sub(r"\s+", "", texto)
                # Mantener solo caracteres permitidos
                texto = re.sub(r"[^a-zA-Z0-9_]", "", texto)
            return texto

        def limpiar_valores(data):
            for key, values in data.items():
                for i in range(
                    1, len(values)
                ):  # Saltamos la primera fila (encabezados)
                    for j in range(
                        1, len(values[i])
                    ):  # Recorremos todas las columnas (excepto encabezados)
                        if isinstance(
                            values[i][j], str
                        ):  # Verificar si es una cadena antes de dividir
                            valor_limpio = (
                                values[i][j].split(";")[0].strip()
                            )  # Tomar solo el primer valor y limpiar espacios

                            # Intentar convertir a flotante si es posible
                            try:
                                values[i][j] = float(valor_limpio)
                            except ValueError:
                                values[i][
                                    j
                                ] = valor_limpio  # Si no se puede convertir, dejar como está
            return data

        def separar_y_mayusculizar(texto):
            # Usamos una expresión regular para separar las mayúsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            # Separar la primera parte (mantenerla como está)
            primera_parte = partes[0]

            # El resto de las partes lo convertimos a mayúsculas y las unimos con espacios
            resto = " ".join([parte.upper() for parte in partes[1:]])
            # resto = " ".join([parte for parte in partes[1:]])

            return primera_parte, resto

        def separar_y_formatear(texto):
            # Usamos una expresión regular para separar las mayúsculas
            partes = re.findall(r"[A-Z][^A-Z]*", texto)

            if not partes:
                return ""

            # Convertimos todo a minúsculas y capitalizamos solo la primera letra
            resultado = " ".join([parte.lower() for parte in partes])
            return resultado.capitalize()

        # Función para limpiar nombres de variables
        def sanitize_name(name):
            return re.sub(
                r"[^a-zA-Z0-9]", "", str(name)
            )  # Elimina cualquier carácter no alfanumérico


if __name__ == "__main__":

    # Copio la ruta del scripot para crear el excel
    ruta_script = os.path.dirname(
        os.path.abspath(__file__)
    )  # Obtener la carpeta sin el nombre del script
    #  pero vuelvo a cambiar para ser compatible con las direcciones de fastapi y la carpeta static para graficos

    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    os.chdir("../../../../")  # Cambia dos niveles arriba

    print(os.getcwd())  # Imprime el directorio de trabajo actual

    data = comun.main.xlsxhojaconvalorespordefecto(ruta_script)

    nombre_funcion = os.path.basename(__file__)[:-3]
    globals()[nombre_funcion](data)
