#!/usr/bin/env python3

import os
import json
import shutil
from openpyxl import Workbook


# def obtener_contenido_jsx(carpeta, subcarpeta, nombre_archivo):


def catalogojson2jsx():
    # Ruta al archivo JSON

    # Ruta base donde se van a guardar los archivos JSX
    base_destino = "../public/excel"

    # Verificar si el directorio existe y borrarlo
    if os.path.exists(base_destino):
        shutil.rmtree(base_destino)

    ruta_json = "fastapi.json"

    # Leer los datos desde el JSON
    with open(ruta_json, "r", encoding="utf-8") as archivo:
        datos = json.load(archivo)



    # Procesar cada entrada del JSON
    for ruta_original, secciones in datos.items():
        # Cambiar extensión a .xlsx
        ruta_excel = os.path.splitext(ruta_original)[0] + ".xlsx"
        ruta_excel=f"../public/excel/{ruta_excel}"

        # Asegurar que el directorio existe
        os.makedirs(os.path.dirname(ruta_excel), exist_ok=True)

        # Crear un archivo Excel
        wb = Workbook()
        # Eliminar la hoja por defecto que se crea
        wb.remove(wb.active)

        for nombre_hoja, contenido in secciones.items():
            # Crear una hoja nueva por cada sección
            ws = wb.create_sheet(title=nombre_hoja[:31])  # Excel permite máx 31 caracteres en nombres de hoja

            for fila in contenido:
                ws.append(fila)

        # Guardar el archivo Excel
        wb.save(ruta_excel)

        continue

        # Construir ruta
        ruta_directorio = os.path.join(base_destino, grupo, sector)
        ruta_archivo = os.path.join(ruta_directorio, f"{cod}.xlsx")

        # Crear directorios (limpios)
        os.makedirs(ruta_directorio, exist_ok=True)

        ################################

        # Crear un libro de trabajo de Excel
        libro = openpyxl.Workbook()

        # Eliminar la hoja predeterminada creada por openpyxl
        libro.remove(libro.active)

        # Crear hojas con los nombres de los keys y escribir los datos
        for sheet_name, sheet_data in data.items():
            hoja = libro.create_sheet(
                title=sheet_name
            )  # Crear la hoja con el nombre del key
            for row in sheet_data:
                hoja.append(row)  # Agregar cada fila a la hoja

        # Guardar el archivo Excel
        libro.save(ruta_archivo)

    print("Archivos creados correctamente.")


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    catalogojson2jsx()
